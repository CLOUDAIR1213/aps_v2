from __future__ import annotations

import json
import re
from io import BytesIO
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete as sqla_delete
from sqlalchemy import select, or_
from sqlalchemy import update as sqla_update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.production import (
    BusinessRiskIssueState,
    ExportBatch,
    ImportBatch,
    OperationDependency,
    OperationMappingRule,
    Part,
    Personnel,
    ProductionOperation,
    ProductionSchedule,
    ProductionScheduleItem,
    ProductionScheduleItemPersonnelAllocation,
    ProductionScheduleOrderLock,
    ResourceGroup,
    ResourceGroupMember,
    ResourceMachine,
    WorkCenter,
    WorkCenterPersonnel,
    WorkOrder,
)
from app.schemas.production import (
    DispatchAutoAssignAllocation,
    DispatchAutoAssignRequest,
    DispatchAutoAssignResponse,
    DispatchAutoAssignSummary,
    DispatchAutoAssignTaskPreview,
    DispatchResponse,
    DispatchTaskRow,
    ExternalTaskListResponse,
    ExternalTaskRow,
    ExternalTaskUpdate,
    ExternalTaskUpdateResponse,
    ImportCommitRequest,
    ImportCommitResponse,
    ImportIssue,
    OperationMappingRuleRead,
    PersonnelAllocationRead,
    PersonnelBatchAllocationRequest,
    PersonnelBatchAllocationResponse,
    PersonnelBatchAllocationSkippedItem,
    PersonnelAllocationSaveResponse,
    PersonnelAllocationWrite,
    PersonnelImportResponse,
    PersonnelOption,
    PersonnelWorkloadResponse,
    PersonnelWorkloadRow,
    PersonnelWorkloadTask,
    ProductionScheduleItemRead,
    ProductionSchedulingResult,
    ResourceGroupRead,
    ResourceMachineRead,
    ResourceGroupMemberRead,
    ScheduleBoardDateColumn,
    ScheduleBoardDailyCell,
    ScheduleBoardResponse,
    ScheduleBoardRow,
    WorkCenterCreate,
    WorkCenterRead,
    WorkCenterUpdate,
    WorkOrderTicketAllocation,
    WorkOrderTicketResponse,
    WorkOrderTicketRow,
)
from app.services.production_import_service import get_parent_no


WORK_START = time(8, 0)
LUNCH_START = time(12, 0)
LUNCH_END = time(13, 0)
WORK_END = time(17, 0)
WEEKDAYS = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
PERSONNEL_SHEET_CANDIDATES = ("机台人员", "机组人员")
MONITOR_HIDDEN_WORK_CENTERS = {"下料"}


def _is_monitor_hidden_names(*names: str | None) -> bool:
    names = {(name or "").strip() for name in names}
    return bool(names & MONITOR_HIDDEN_WORK_CENTERS)


def is_monitor_hidden_schedule_item(item: ProductionScheduleItem) -> bool:
    operation = item.operation
    center = operation.work_center if operation else None
    return _is_monitor_hidden_names(
        operation.name if operation else None,
        center.name if center else None,
    )


def is_monitor_hidden_operation(operation: ProductionOperation) -> bool:
    return _is_monitor_hidden_names(operation.name, operation.work_center.name)


def normalize_schedule_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(second=0, microsecond=0)
    return value.astimezone().replace(tzinfo=None, second=0, microsecond=0)


def slugify_code(name: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z]+", "-", name).strip("-").upper()
    if cleaned:
        return cleaned[:40]
    return "WC-" + str(abs(hash(name)) % 100000)


def _excel_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _sheet_row_values(sheet, row: int) -> set[str]:
    return {
        _excel_text(sheet.cell(row, col).value)
        for col in range(1, sheet.max_column + 1)
        if _excel_text(sheet.cell(row, col).value)
    }


def _find_personnel_header_row(sheet) -> int | None:
    for row in range(1, min(sheet.max_row, 8) + 1):
        row_values = _sheet_row_values(sheet, row)
        next_row_values = _sheet_row_values(sheet, row + 1)
        if "NO." in row_values and {"工号", "姓名"}.intersection(next_row_values):
            return row
    return None


def _find_personnel_sheet(workbook):
    for sheet_name in PERSONNEL_SHEET_CANDIDATES:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    for sheet in workbook.worksheets:
        if _find_personnel_header_row(sheet) is not None:
            return sheet
    return None


def _merged_block_bounds(sheet, row: int, col: int) -> tuple[int, int]:
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return merged_range.min_col, merged_range.max_col
    return col, col


def _merged_block_value(sheet, row: int, col: int) -> str:
    start_col, _ = _merged_block_bounds(sheet, row, col)
    return _excel_text(sheet.cell(row, start_col).value)


def _personnel_header_blocks(sheet, header_row: int) -> list[tuple[str, int, int]]:
    blocks: list[tuple[str, int, int]] = []
    seen_starts: set[int] = set()
    subheader_row = header_row + 1

    col = 1
    while col <= sheet.max_column:
        block_start, block_end = _merged_block_bounds(sheet, header_row, col)
        if block_start in seen_starts:
            col += 1
            continue
        seen_starts.add(block_start)

        work_center_name = _merged_block_value(sheet, header_row, col)
        if not work_center_name or work_center_name == "NO.":
            col = max(col + 1, block_end + 1)
            continue

        if block_start == block_end:
            current_subheader = _excel_text(sheet.cell(subheader_row, block_start).value)
            next_subheader = _excel_text(sheet.cell(subheader_row, block_start + 1).value)
            if current_subheader in {"工号", "外协"} or (
                not current_subheader and next_subheader in {"姓名", "外协"}
            ):
                block_end = min(block_start + 1, sheet.max_column)

        blocks.append((work_center_name, block_start, block_end))
        col = max(col + 1, block_end + 1)

    return blocks


async def list_work_centers(db: AsyncSession) -> list[WorkCenterRead]:
    result = await db.execute(
        select(WorkCenter)
        .options(selectinload(WorkCenter.machines))
        .order_by(WorkCenter.is_external, WorkCenter.name)
    )
    centers = list(result.scalars().all())
    return [
        WorkCenterRead(
            id=center.id,
            code=center.code,
            name=center.name,
            is_external=center.is_external,
            default_capacity_per_day=center.default_capacity_per_day,
            default_duration_hours=center.default_duration_hours,
            external_capacity_slots=center.external_capacity_slots,
            external_lead_time_hours=center.external_lead_time_hours,
            external_vendor_name=center.external_vendor_name,
            status=center.status,
            description=center.description,
            machine_count=len(center.machines),
            created_at=center.created_at,
            updated_at=center.updated_at,
        )
        for center in centers
    ]


async def list_resource_machines(db: AsyncSession) -> list[ResourceMachineRead]:
    result = await db.execute(select(ResourceMachine).order_by(ResourceMachine.work_center_id, ResourceMachine.code))
    return [ResourceMachineRead.model_validate(machine) for machine in result.scalars().all()]


async def import_personnel_workbook(
    db: AsyncSession,
    content: bytes,
    filename: str,
) -> PersonnelImportResponse:
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=True, keep_vba=True)
    sheet = _find_personnel_sheet(workbook)
    if sheet is None:
        names = "、".join(PERSONNEL_SHEET_CANDIDATES)
        raise ValueError(f"Workbook must contain personnel sheet '{names}', or a sheet with NO./工号/姓名 headers.")

    issues: list[ImportIssue] = []
    imported_people: set[str] = set()
    linked_work_centers: set[int] = set()
    links_created = 0

    header_row = _find_personnel_header_row(sheet)
    if header_row is None:
        raise ValueError(f"Sheet '{sheet.title}' does not contain a valid NO./工号/姓名 personnel header.")

    subheader_row = header_row + 1
    data_start_row = header_row + 2

    for work_center_name, block_start, block_end in _personnel_header_blocks(sheet, header_row):
        subheaders = {
            _excel_text(sheet.cell(subheader_row, col).value)
            for col in range(block_start, block_end + 1)
            if _excel_text(sheet.cell(subheader_row, col).value)
        }
        if "外协" in subheaders and not {"工号", "姓名"}.issubset(subheaders):
            issues.append(
                ImportIssue(
                    severity="info",
                    row=header_row,
                    column=block_start,
                    field=work_center_name,
                    message=f"{work_center_name} 为外协工段，人员导入已跳过。",
                )
            )
            continue

        person_column_pairs: list[tuple[int, int]] = []
        col = block_start
        while col <= block_end:
            left_header = _excel_text(sheet.cell(subheader_row, col).value)
            right_header = _excel_text(sheet.cell(subheader_row, col + 1).value) if col + 1 <= sheet.max_column else ""
            if left_header == "工号" and right_header == "姓名":
                person_column_pairs.append((col, col + 1))
                col += 2
                continue
            col += 1

        if not person_column_pairs:
            issues.append(
                ImportIssue(
                    severity="warning",
                    row=header_row,
                    column=block_start,
                    field=work_center_name,
                    message=f"{work_center_name} 未找到工号/姓名列，已跳过。",
                )
            )
            continue

        center = await ensure_work_center(db, work_center_name, False)
        linked_work_centers.add(center.id)

        for employee_col, name_col in person_column_pairs:
            for row in range(data_start_row, sheet.max_row + 1):
                employee_no = _excel_text(sheet.cell(row, employee_col).value)
                person_name = _excel_text(sheet.cell(row, name_col).value)
                if not person_name and not employee_no:
                    continue
                if not person_name or not employee_no:
                    issues.append(
                        ImportIssue(
                            severity="warning",
                            row=row,
                            column=employee_col,
                            field=work_center_name,
                            message=f"{work_center_name} 第 {row} 行工号或姓名不完整，已跳过。",
                        )
                    )
                    continue

                result = await db.execute(select(Personnel).where(Personnel.employee_no == employee_no))
                person = result.scalars().first()
                if person is None:
                    person = Personnel(employee_no=employee_no, name=person_name)
                    db.add(person)
                    await db.commit()
                    await db.refresh(person)
                elif person.name != person_name:
                    person.name = person_name
                    await db.commit()
                    await db.refresh(person)
                imported_people.add(employee_no)

                link_result = await db.execute(
                    select(WorkCenterPersonnel).where(
                        WorkCenterPersonnel.work_center_id == center.id,
                        WorkCenterPersonnel.person_id == person.id,
                    )
                )
                if link_result.scalars().first() is None:
                    db.add(
                        WorkCenterPersonnel(
                            work_center_id=center.id,
                            person_id=person.id,
                            sort_order=row - data_start_row,
                        )
                    )
                    await db.commit()
                    links_created += 1

    return PersonnelImportResponse(
        imported_people=len(imported_people),
        linked_work_centers=len(linked_work_centers),
        links_created=links_created,
        issues=issues,
    )


async def get_work_center_names(db: AsyncSession) -> set[str]:
    result = await db.execute(select(WorkCenter.name))
    return set(result.scalars().all())


async def ensure_work_center(
    db: AsyncSession,
    name: str,
    is_external: bool = False,
) -> WorkCenter:
    result = await db.execute(select(WorkCenter).where(WorkCenter.name == name))
    existing = result.scalars().first()
    if existing:
        if is_external and not existing.is_external:
            existing.is_external = True
            await db.commit()
            await db.refresh(existing)
        return existing

    code = slugify_code(name)
    suffix = 1
    while True:
        code_result = await db.execute(select(WorkCenter).where(WorkCenter.code == code))
        if code_result.scalars().first() is None:
            break
        suffix += 1
        code = f"{slugify_code(name)}-{suffix}"

    center = WorkCenter(
        code=code,
        name=name,
        is_external=is_external,
        default_capacity_per_day=480,
        default_duration_hours=8,
        external_capacity_slots=1,
    )
    db.add(center)
    await db.commit()
    await db.refresh(center)

    if not center.is_external:
        machine = ResourceMachine(
            work_center_id=center.id,
            code=f"{center.code}-01",
            name=f"{center.name}-01",
            capacity_per_day=center.default_capacity_per_day,
        )
        db.add(machine)
        await db.commit()

    return center


async def create_work_center(db: AsyncSession, payload: WorkCenterCreate) -> WorkCenterRead:
    center = WorkCenter(
        code=payload.code,
        name=payload.name,
        is_external=payload.is_external,
        default_capacity_per_day=payload.default_capacity_per_day,
        default_duration_hours=payload.default_duration_hours,
        external_capacity_slots=max(payload.external_capacity_slots, 1),
        external_lead_time_hours=payload.external_lead_time_hours,
        external_vendor_name=payload.external_vendor_name,
        status=payload.status,
        description=payload.description,
    )
    db.add(center)
    await db.commit()
    await db.refresh(center)

    if not payload.is_external:
        for index in range(max(payload.machine_count, 1)):
            db.add(
                ResourceMachine(
                    work_center_id=center.id,
                    code=f"{center.code}-{index + 1:02d}",
                    name=f"{center.name}-{index + 1:02d}",
                    capacity_per_day=center.default_capacity_per_day,
                )
            )
        await db.commit()
        await db.refresh(center, attribute_names=["machines"])

    return WorkCenterRead(
        id=center.id,
        code=center.code,
        name=center.name,
        is_external=center.is_external,
        default_capacity_per_day=center.default_capacity_per_day,
        default_duration_hours=center.default_duration_hours,
        external_capacity_slots=center.external_capacity_slots,
        external_lead_time_hours=center.external_lead_time_hours,
        external_vendor_name=center.external_vendor_name,
        status=center.status,
        description=center.description,
        machine_count=len(center.machines),
        created_at=center.created_at,
        updated_at=center.updated_at,
    )


async def update_work_center(db: AsyncSession, work_center_id: int, payload: WorkCenterUpdate) -> WorkCenterRead:
    center = await db.get(WorkCenter, work_center_id)
    if not center:
        raise ValueError("工段不存在。")
    update_data = payload.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if field == "code" and value:
            existing = await db.execute(
                select(WorkCenter).where(WorkCenter.code == value, WorkCenter.id != work_center_id)
            )
            if existing.scalars().first():
                raise ValueError(f"工段编码 {value} 已存在。")
        if field == "name" and value:
            existing = await db.execute(
                select(WorkCenter).where(WorkCenter.name == value, WorkCenter.id != work_center_id)
            )
            if existing.scalars().first():
                raise ValueError(f"工段名称 {value} 已存在。")
        if field == "external_capacity_slots":
            value = max(int(value or 1), 1)
        setattr(center, field, value)
    await db.commit()
    await db.refresh(center)
    return WorkCenterRead(
        id=center.id,
        code=center.code,
        name=center.name,
        is_external=center.is_external,
        default_capacity_per_day=center.default_capacity_per_day,
        default_duration_hours=center.default_duration_hours,
        external_capacity_slots=center.external_capacity_slots,
        external_lead_time_hours=center.external_lead_time_hours,
        external_vendor_name=center.external_vendor_name,
        status=center.status,
        description=center.description,
        machine_count=len(center.machines) if center.machines else 0,
        created_at=center.created_at,
        updated_at=center.updated_at,
    )


async def disable_work_center(db: AsyncSession, work_center_id: int) -> WorkCenterRead:
    center = await db.get(WorkCenter, work_center_id, options=[selectinload(WorkCenter.machines)])
    if not center:
        raise ValueError("工段不存在。")
    center.status = "disabled"
    await db.commit()
    await db.refresh(center)
    return WorkCenterRead(
        id=center.id,
        code=center.code,
        name=center.name,
        is_external=center.is_external,
        default_capacity_per_day=center.default_capacity_per_day,
        default_duration_hours=center.default_duration_hours,
        external_capacity_slots=center.external_capacity_slots,
        external_lead_time_hours=center.external_lead_time_hours,
        external_vendor_name=center.external_vendor_name,
        status=center.status,
        description=center.description,
        machine_count=len(center.machines),
        created_at=center.created_at,
        updated_at=center.updated_at,
    )


async def delete_work_center(db: AsyncSession, work_center_id: int) -> None:
    center = await db.get(WorkCenter, work_center_id)
    if not center:
        raise ValueError("工段不存在。")

    operation_ref = await db.scalar(
        select(ProductionOperation.id)
        .where(ProductionOperation.work_center_id == work_center_id)
        .limit(1)
    )
    schedule_ref = await db.scalar(
        select(ProductionScheduleItem.id)
        .where(ProductionScheduleItem.work_center_id == work_center_id)
        .limit(1)
    )
    if operation_ref or schedule_ref:
        raise ValueError("该工段已被工单或历史排产引用，不能物理删除；请使用禁用。")

    machine_ids = list(
        (
            await db.execute(
                select(ResourceMachine.id).where(ResourceMachine.work_center_id == work_center_id)
            )
        ).scalars().all()
    )
    await db.execute(
        sqla_delete(ResourceGroupMember).where(
            ResourceGroupMember.member_type == "work_center",
            ResourceGroupMember.member_id == work_center_id,
        )
    )
    if machine_ids:
        await db.execute(
            sqla_delete(ResourceGroupMember).where(
                ResourceGroupMember.member_type == "machine",
                ResourceGroupMember.member_id.in_(machine_ids),
            )
        )
    await db.execute(
        sqla_delete(OperationMappingRule).where(OperationMappingRule.work_center_id == work_center_id)
    )
    await db.execute(
        sqla_delete(WorkCenterPersonnel).where(WorkCenterPersonnel.work_center_id == work_center_id)
    )
    await db.execute(
        sqla_delete(ResourceMachine).where(ResourceMachine.work_center_id == work_center_id)
    )
    await db.execute(sqla_delete(WorkCenter).where(WorkCenter.id == work_center_id))
    await db.commit()


async def create_machine(db: AsyncSession, payload) -> ResourceMachineRead:
    center = await db.get(WorkCenter, payload.work_center_id)
    if not center:
        raise ValueError("工段不存在。")
    existing = await db.execute(select(ResourceMachine).where(ResourceMachine.code == payload.code))
    if existing.scalars().first():
        raise ValueError(f"设备编码 {payload.code} 已存在。")
    machine = ResourceMachine(
        work_center_id=payload.work_center_id,
        code=payload.code,
        name=payload.name,
        status=payload.status,
        capacity_per_day=payload.capacity_per_day,
    )
    db.add(machine)
    await db.commit()
    await db.refresh(machine)
    return ResourceMachineRead.model_validate(machine)


async def update_machine(db: AsyncSession, machine_id: int, payload) -> ResourceMachineRead:
    machine = await db.get(ResourceMachine, machine_id)
    if not machine:
        raise ValueError("设备不存在。")
    update_data = payload.model_dump(exclude_unset=True)
    if "code" in update_data and update_data["code"]:
        existing = await db.execute(
            select(ResourceMachine).where(
                ResourceMachine.code == update_data["code"],
                ResourceMachine.id != machine_id,
            )
        )
        if existing.scalars().first():
            raise ValueError(f"设备编码 {update_data['code']} 已存在。")
    for field, value in update_data.items():
        setattr(machine, field, value)
    await db.commit()
    await db.refresh(machine)
    return ResourceMachineRead.model_validate(machine)


async def delete_machine(db: AsyncSession, machine_id: int) -> None:
    machine = await db.get(ResourceMachine, machine_id)
    if not machine:
        raise ValueError("设备不存在。")
    schedule_ref = await db.scalar(
        select(ProductionScheduleItem.id)
        .where(ProductionScheduleItem.machine_id == machine_id)
        .limit(1)
    )
    if schedule_ref:
        raise ValueError("该设备已被历史排产引用，不能物理删除；请改为禁用或停机。")
    await db.execute(
        sqla_delete(ResourceGroupMember).where(
            ResourceGroupMember.member_type == "machine",
            ResourceGroupMember.member_id == machine_id,
        )
    )
    await db.execute(sqla_delete(ResourceMachine).where(ResourceMachine.id == machine_id))
    await db.commit()


async def list_operation_mapping_rules(db: AsyncSession) -> list[OperationMappingRuleRead]:
    result = await db.execute(
        select(OperationMappingRule)
        .options(selectinload(OperationMappingRule.work_center))
        .order_by(OperationMappingRule.source_name)
    )
    rules = list(result.scalars().all())
    return [
        OperationMappingRuleRead(
            id=rule.id,
            source_name=rule.source_name,
            normalized_name=rule.normalized_name,
            work_center_id=rule.work_center_id,
            is_external=rule.is_external,
            status=rule.status,
            created_at=rule.created_at,
            updated_at=rule.updated_at,
            work_center_name=rule.work_center.name if rule.work_center else None,
        )
        for rule in rules
    ]


async def create_operation_mapping_rule(db: AsyncSession, payload) -> OperationMappingRuleRead:
    existing = await db.execute(
        select(OperationMappingRule).where(OperationMappingRule.source_name == payload.source_name)
    )
    if existing.scalars().first():
        raise ValueError(f"映射规则 {payload.source_name} 已存在，请使用更新接口。")
    center = await db.get(WorkCenter, payload.work_center_id)
    if not center:
        raise ValueError("关联工段不存在。")
    rule = OperationMappingRule(
        source_name=payload.source_name,
        normalized_name=payload.normalized_name,
        work_center_id=payload.work_center_id,
        is_external=payload.is_external,
        status=payload.status,
    )
    db.add(rule)
    await db.commit()
    await db.refresh(rule)
    return OperationMappingRuleRead(
        id=rule.id,
        source_name=rule.source_name,
        normalized_name=rule.normalized_name,
        work_center_id=rule.work_center_id,
        is_external=rule.is_external,
        status=rule.status,
        created_at=rule.created_at,
        updated_at=rule.updated_at,
        work_center_name=center.name,
    )


async def update_operation_mapping_rule(db: AsyncSession, rule_id: int, payload) -> OperationMappingRuleRead:
    rule = await db.get(OperationMappingRule, rule_id, options=[selectinload(OperationMappingRule.work_center)])
    if not rule:
        raise ValueError("映射规则不存在。")
    update_data = payload.model_dump(exclude_unset=True)
    if "work_center_id" in update_data:
        center = await db.get(WorkCenter, update_data["work_center_id"])
        if not center:
            raise ValueError("关联工段不存在。")
    for field, value in update_data.items():
        setattr(rule, field, value)
    await db.commit()
    await db.refresh(rule)
    return OperationMappingRuleRead(
        id=rule.id,
        source_name=rule.source_name,
        normalized_name=rule.normalized_name,
        work_center_id=rule.work_center_id,
        is_external=rule.is_external,
        status=rule.status,
        created_at=rule.created_at,
        updated_at=rule.updated_at,
        work_center_name=rule.work_center.name if rule.work_center else None,
    )


async def delete_operation_mapping_rule(db: AsyncSession, rule_id: int) -> None:
    rule = await db.get(OperationMappingRule, rule_id)
    if not rule:
        raise ValueError("映射规则不存在。")
    await db.delete(rule)
    await db.commit()


async def list_resource_groups(db: AsyncSession) -> list[ResourceGroupRead]:
    result = await db.execute(
        select(ResourceGroup)
        .options(selectinload(ResourceGroup.members))
        .order_by(ResourceGroup.code)
    )
    return [ResourceGroupRead.model_validate(group) for group in result.scalars().all()]


async def create_resource_group(db: AsyncSession, payload) -> ResourceGroupRead:
    existing = await db.execute(select(ResourceGroup).where(ResourceGroup.code == payload.code))
    if existing.scalars().first():
        raise ValueError(f"资源组编码 {payload.code} 已存在。")
    group = ResourceGroup(
        code=payload.code,
        name=payload.name,
        description=payload.description,
        status=payload.status,
    )
    db.add(group)
    await db.commit()
    await db.refresh(group)
    return ResourceGroupRead.model_validate(group)


async def update_resource_group(db: AsyncSession, group_id: int, payload) -> ResourceGroupRead:
    group = await db.get(ResourceGroup, group_id, options=[selectinload(ResourceGroup.members)])
    if not group:
        raise ValueError("资源组不存在。")
    update_data = payload.model_dump(exclude_unset=True)
    if "code" in update_data and update_data["code"]:
        existing = await db.execute(
            select(ResourceGroup).where(
                ResourceGroup.code == update_data["code"],
                ResourceGroup.id != group_id,
            )
        )
        if existing.scalars().first():
            raise ValueError(f"资源组编码 {update_data['code']} 已存在。")
    for field, value in update_data.items():
        setattr(group, field, value)
    await db.commit()
    await db.refresh(group)
    return ResourceGroupRead.model_validate(group)


async def add_resource_group_member(db: AsyncSession, group_id: int, payload) -> ResourceGroupMemberRead:
    group = await db.get(ResourceGroup, group_id)
    if not group:
        raise ValueError("资源组不存在。")

    # Validate member_type and member_id reference
    model_map = {"work_center": WorkCenter, "machine": ResourceMachine, "personnel": Personnel}
    model = model_map.get(payload.member_type)
    if model is None:
        raise ValueError("member_type 必须是 work_center、machine 或 personnel。")
    member_record = await db.get(model, payload.member_id)
    if not member_record:
        type_labels = {"work_center": "工段", "machine": "设备", "personnel": "人员"}
        raise ValueError(f"{type_labels[payload.member_type]} ID {payload.member_id} 不存在。")

    existing = await db.execute(
        select(ResourceGroupMember).where(
            ResourceGroupMember.group_id == group_id,
            ResourceGroupMember.member_type == payload.member_type,
            ResourceGroupMember.member_id == payload.member_id,
        )
    )
    if existing.scalars().first():
        raise ValueError("该资源已在资源组中。")
    member = ResourceGroupMember(
        group_id=group_id,
        member_type=payload.member_type,
        member_id=payload.member_id,
    )
    db.add(member)
    await db.commit()
    await db.refresh(member)
    return ResourceGroupMemberRead.model_validate(member)


async def remove_resource_group_member(db: AsyncSession, group_id: int, member_id: int) -> None:
    member = await db.get(ResourceGroupMember, member_id)
    if not member or member.group_id != group_id:
        raise ValueError("资源组成员不存在。")
    await db.delete(member)
    await db.commit()


def _is_missing_mapping_issue(issue: ImportIssue) -> bool:
    return issue.severity == "error" and (
        issue.field == "work_center"
        or "尚未配置映射规则" in issue.message
        or "未配置映射" in issue.message
    )


def _is_mapping_issue_covered(issue: ImportIssue, mapping_overrides: dict[str, int]) -> bool:
    return _is_missing_mapping_issue(issue) and any(
        source_name and source_name in issue.message for source_name in mapping_overrides
    )


def _format_import_issue(issue: ImportIssue) -> str:
    location = ""
    if issue.row:
        location = f"R{issue.row}: "
    elif issue.column:
        location = f"C{issue.column}: "
    return f"{location}{issue.message}"


def _sorted_operations(operations: list[ProductionOperation]) -> list[ProductionOperation]:
    return sorted(operations, key=lambda op: (op.seq_no, op.id or 0))


def _build_operation_dependencies(
    parts: list[Any],
    operations_by_part_no: dict[str, list[ProductionOperation]],
) -> tuple[list[OperationDependency], int, int]:
    dependencies: list[OperationDependency] = []
    seen_edges: set[tuple[int, int]] = set()
    sequence_count = 0
    hierarchy_count = 0

    def add_dependency(operation_id: int, depends_on_operation_id: int, kind: str) -> None:
        nonlocal sequence_count, hierarchy_count
        if operation_id == depends_on_operation_id:
            return
        edge = (operation_id, depends_on_operation_id)
        if edge in seen_edges:
            return
        seen_edges.add(edge)
        dependencies.append(
            OperationDependency(
                operation_id=operation_id,
                depends_on_operation_id=depends_on_operation_id,
            )
        )
        if kind == "sequence":
            sequence_count += 1
        else:
            hierarchy_count += 1

    for operations in operations_by_part_no.values():
        sorted_ops = _sorted_operations(operations)
        for previous, current in zip(sorted_ops, sorted_ops[1:], strict=False):
            add_dependency(current.id, previous.id, "sequence")

    part_nos = {part.no for part in parts}
    children_by_parent: dict[str, list[str]] = defaultdict(list)
    for part in parts:
        parent_no = get_parent_no(part.no)
        if parent_no and parent_no in part_nos:
            children_by_parent[parent_no].append(part.no)

    finish_anchor_cache: dict[str, list[ProductionOperation]] = {}

    def finish_anchors(part_no: str, visiting: set[str] | None = None) -> list[ProductionOperation]:
        if part_no in finish_anchor_cache:
            return finish_anchor_cache[part_no]
        visiting = visiting or set()
        if part_no in visiting:
            return []
        visiting.add(part_no)
        own_operations = _sorted_operations(operations_by_part_no.get(part_no, []))
        if own_operations:
            anchors = [own_operations[-1]]
        else:
            anchors = []
            for child_no in children_by_parent.get(part_no, []):
                anchors.extend(finish_anchors(child_no, visiting.copy()))
        finish_anchor_cache[part_no] = anchors
        return anchors

    for parent_no, child_nos in children_by_parent.items():
        parent_operations = _sorted_operations(operations_by_part_no.get(parent_no, []))
        if not parent_operations:
            continue
        parent_first_operation = parent_operations[0]
        for child_no in child_nos:
            for child_anchor in finish_anchors(child_no):
                add_dependency(parent_first_operation.id, child_anchor.id, "hierarchy")

    return dependencies, sequence_count, hierarchy_count


async def commit_import(db: AsyncSession, request: ImportCommitRequest) -> ImportCommitResponse:
    blocking_issues = [
        issue
        for issue in request.preview.issues
        if (
            issue.severity == "error"
            and not (
                (request.create_missing_work_centers and _is_missing_mapping_issue(issue))
                or _is_mapping_issue_covered(issue, request.mapping_overrides)
            )
        )
        or issue.field == "external_default_duration"
    ]
    if blocking_issues:
        issue_text = "；".join(_format_import_issue(issue) for issue in blocking_issues[:5])
        if len(blocking_issues) > 5:
            issue_text = f"{issue_text}；另有 {len(blocking_issues) - 5} 个错误"
        raise ValueError(f"导入预览仍有不可自动处理的错误：{issue_text}")

    existing_order = await db.execute(select(WorkOrder).where(WorkOrder.order_no == request.order.order_no))
    if existing_order.scalars().first():
        raise ValueError(f"Work order {request.order.order_no} already exists.")

    # Load active mapping rules
    rules_result = await db.execute(
        select(OperationMappingRule)
        .where(OperationMappingRule.status == "active")
        .options(selectinload(OperationMappingRule.work_center))
    )
    rule_map = {rule.source_name: rule for rule in rules_result.scalars().all()}

    # User overrides bind an original Excel operation column to an existing work center.
    for source_name, work_center_id in request.mapping_overrides.items():
        center = await db.get(WorkCenter, work_center_id)
        if center is None:
            raise ValueError(f"映射覆盖的工段不存在：{source_name} -> {work_center_id}。")
        if center.status != "active":
            raise ValueError(f"映射覆盖的工段已禁用：{source_name} -> {center.name}。")

        existing_rule_result = await db.execute(
            select(OperationMappingRule)
            .where(OperationMappingRule.source_name == source_name)
            .options(selectinload(OperationMappingRule.work_center))
        )
        existing_rule = existing_rule_result.scalars().first()
        is_external = center.is_external
        if existing_rule:
            existing_rule.work_center_id = center.id
            existing_rule.normalized_name = center.name
            existing_rule.is_external = is_external
            existing_rule.status = "active"
            existing_rule.work_center = center
            rule_map[source_name] = existing_rule
        else:
            rule = OperationMappingRule(
                source_name=source_name,
                normalized_name=center.name,
                work_center_id=center.id,
                is_external=is_external,
                status="active",
            )
            rule.work_center = center
            db.add(rule)
            rule_map[source_name] = rule
    if request.mapping_overrides:
        await db.commit()

    # Validate all operations have mapping rules, or create missing rules when the user chose it.
    unmapped = {op.work_center_name for op in request.preview.operations if op.work_center_name not in rule_map}
    if unmapped:
        if not request.create_missing_work_centers:
            raise ValueError(f"以下工序列尚未配置映射规则：{'、'.join(sorted(unmapped))}。请先在工序映射中配置。")
        for source_name in sorted(unmapped):
            existing_rule_result = await db.execute(
                select(OperationMappingRule)
                .where(OperationMappingRule.source_name == source_name)
                .options(selectinload(OperationMappingRule.work_center))
            )
            existing_rule = existing_rule_result.scalars().first()
            is_external = any(
                op.work_center_name == source_name and op.is_external
                for op in request.preview.operations
            )
            if existing_rule:
                if existing_rule.status != "active":
                    existing_rule.status = "active"
                if existing_rule.is_external != is_external:
                    existing_rule.is_external = is_external
                if existing_rule.work_center and is_external and not existing_rule.work_center.is_external:
                    existing_rule.work_center.is_external = True
                rule_map[source_name] = existing_rule
                continue

            center = await ensure_work_center(db, source_name, is_external=is_external)
            rule = OperationMappingRule(
                source_name=source_name,
                normalized_name=source_name,
                work_center_id=center.id,
                is_external=center.is_external,
                status="active",
            )
            rule.work_center = center
            db.add(rule)
            rule_map[source_name] = rule
        await db.commit()

    center_map: dict[str, WorkCenter] = {}
    for operation in request.preview.operations:
        rule = rule_map[operation.work_center_name]
        center_map[operation.work_center_name] = rule.work_center

    work_order = WorkOrder(**request.order.model_dump(), status="pending")
    db.add(work_order)
    await db.commit()
    await db.refresh(work_order)

    batch = ImportBatch(
        work_order_id=work_order.id,
        source_filename=request.preview.source_filename,
        sheet_name=request.preview.sheet_name,
        parsed_summary_json=json.dumps(request.preview.summary, ensure_ascii=False),
    )
    db.add(batch)
    await db.commit()
    await db.refresh(batch)

    part_map: dict[str, Part] = {}
    for item in request.preview.parts:
        part = Part(
            work_order_id=work_order.id,
            import_batch_id=batch.id,
            no=item.no,
            drawing_no=item.drawing_no,
            name=item.name,
            material=item.material,
            note=item.note,
            quantity=item.quantity,
            source_row=item.source_row,
            is_assembly=item.is_assembly,
        )
        db.add(part)
        part_map[item.no] = part
    await db.commit()

    for item in request.preview.parts:
        part = part_map[item.no]
        await db.refresh(part)
        parent_no = get_parent_no(item.no)
        if parent_no and parent_no in part_map:
            part.parent_part_id = part_map[parent_no].id
    await db.commit()

    operations_by_part_no: dict[str, list[ProductionOperation]] = defaultdict(list)
    for item in request.preview.operations:
        part = part_map.get(item.part_no)
        center = center_map[item.work_center_name]
        if not part:
            continue
        operation = ProductionOperation(
            work_order_id=work_order.id,
            part_id=part.id,
            work_center_id=center.id,
            name=item.work_center_name,
            seq_no=item.seq_no,
            duration_hours=item.duration_hours,
            requirement_note=item.requirement_note,
            source_row=item.source_row,
            source_col=item.source_col,
        )
        db.add(operation)
        operations_by_part_no[item.part_no].append(operation)
    await db.commit()

    all_dependencies, sequence_dependency_count, hierarchy_dependency_count = _build_operation_dependencies(
        request.preview.parts,
        operations_by_part_no,
    )

    if all_dependencies:
        db.add_all(all_dependencies)
        await db.commit()

    return ImportCommitResponse(
        work_order=work_order,
        import_batch_id=batch.id,
        part_count=len(part_map),
        operation_count=sum(len(items) for items in operations_by_part_no.values()),
        dependency_count=len(all_dependencies),
        sequence_dependency_count=sequence_dependency_count,
        hierarchy_dependency_count=hierarchy_dependency_count,
    )


async def list_work_orders(db: AsyncSession) -> list[WorkOrder]:
    result = await db.execute(select(WorkOrder).order_by(WorkOrder.status, WorkOrder.due_date, WorkOrder.id))
    return list(result.scalars().all())


async def delete_work_order(db: AsyncSession, work_order_id: int) -> None:
    work_order = await db.get(WorkOrder, work_order_id)
    if not work_order:
        raise ValueError("订单不存在。")

    operation_ids = select(ProductionOperation.id).where(
        ProductionOperation.work_order_id == work_order_id
    )
    await db.execute(
        sqla_delete(ProductionScheduleItem).where(
            ProductionScheduleItem.work_order_id == work_order_id
        )
    )
    await db.execute(
        sqla_delete(ProductionScheduleOrderLock).where(
            ProductionScheduleOrderLock.work_order_id == work_order_id
        )
    )
    await db.execute(
        sqla_delete(OperationDependency).where(
            or_(
                OperationDependency.operation_id.in_(operation_ids),
                OperationDependency.depends_on_operation_id.in_(operation_ids),
            )
        )
    )
    await db.execute(
        sqla_update(Part)
        .where(Part.work_order_id == work_order_id)
        .values(parent_part_id=None)
    )
    await db.execute(
        sqla_delete(ProductionOperation).where(
            ProductionOperation.work_order_id == work_order_id
        )
    )
    await db.execute(sqla_delete(Part).where(Part.work_order_id == work_order_id))
    await db.execute(sqla_delete(ImportBatch).where(ImportBatch.work_order_id == work_order_id))
    await db.execute(sqla_delete(WorkOrder).where(WorkOrder.id == work_order_id))
    await db.commit()


async def update_operation_requirement_note(
    db: AsyncSession,
    operation_id: int,
    requirement_note: str | None,
) -> ProductionOperation:
    operation = await db.get(ProductionOperation, operation_id)
    if operation is None:
        raise ValueError("工序不存在。")

    normalized_note = requirement_note.strip() if isinstance(requirement_note, str) else requirement_note
    operation.requirement_note = normalized_note or None
    await db.commit()
    await db.refresh(operation)
    return operation


async def list_personnel(db: AsyncSession) -> list[dict]:
    result = await db.execute(
        select(Personnel)
        .options(selectinload(Personnel.work_centers).selectinload(WorkCenterPersonnel.work_center))
        .order_by(Personnel.name)
    )
    people = list(result.scalars().all())
    return [
        {
            "id": p.id,
            "employee_no": p.employee_no,
            "name": p.name,
            "status": p.status,
            "work_centers": [
                {"id": link.work_center.id, "name": link.work_center.name, "code": link.work_center.code}
                for link in p.work_centers
            ],
            "created_at": p.created_at,
            "updated_at": p.updated_at,
        }
        for p in people
    ]


async def delete_personnel(db: AsyncSession, person_id: int) -> None:
    person = await db.get(Personnel, person_id)
    if not person:
        raise ValueError("人员不存在。")
    allocation_result = await db.execute(
        select(ProductionScheduleItemPersonnelAllocation.id)
        .where(ProductionScheduleItemPersonnelAllocation.person_id == person_id)
        .limit(1)
    )
    if allocation_result.scalar_one_or_none() is not None:
        raise ValueError("该人员已有派工分摊记录，不能物理删除；请先处理派工记录或将人员状态改为离职。")
    await db.execute(
        sqla_delete(ResourceGroupMember).where(
            ResourceGroupMember.member_type == "personnel",
            ResourceGroupMember.member_id == person_id,
        )
    )
    await db.execute(
        sqla_delete(WorkCenterPersonnel).where(WorkCenterPersonnel.person_id == person_id)
    )
    await db.execute(sqla_delete(Personnel).where(Personnel.id == person_id))
    await db.commit()


def _serialize_personnel_option(person: Personnel) -> PersonnelOption:
    return PersonnelOption(
        id=person.id,
        employee_no=person.employee_no,
        name=person.name,
        status=person.status,
    )


def _serialize_allocation(allocation: ProductionScheduleItemPersonnelAllocation) -> PersonnelAllocationRead:
    return PersonnelAllocationRead(
        id=allocation.id,
        schedule_item_id=allocation.schedule_item_id,
        person_id=allocation.person_id,
        employee_no=allocation.person.employee_no,
        person_name=allocation.person.name,
        ratio_percent=allocation.ratio_percent,
        planned_minutes=allocation.planned_minutes,
    )


def _dispatch_row(item: ProductionScheduleItem) -> DispatchTaskRow:
    operation = item.operation
    work_order = operation.work_order
    part = operation.part
    center = operation.work_center
    planned_minutes = scheduled_work_minutes(item.start_time, item.end_time)
    allocations = sorted(item.personnel_allocations, key=lambda row: (row.person.name, row.person_id))
    serialized_allocations = [_serialize_allocation(allocation) for allocation in allocations]
    assigned_minutes = sum(allocation.planned_minutes for allocation in allocations)
    if not allocations:
        allocation_status = "unassigned"
    elif abs(sum(allocation.ratio_percent for allocation in allocations) - 100) <= 0.001:
        allocation_status = "assigned"
    else:
        allocation_status = "partial"
    return DispatchTaskRow(
        schedule_item_id=item.id,
        operation_id=item.operation_id,
        work_order_id=item.work_order_id,
        work_center_id=item.work_center_id,
        machine_id=item.machine_id,
        order_no=work_order.order_no,
        customer=work_order.customer,
        drawing_no=part.drawing_no,
        part_no=part.no,
        part_name=part.name,
        operation_name=operation.name,
        requirement_note=operation.requirement_note,
        work_center_name=center.name,
        machine_name=item.machine.name if item.machine else None,
        is_external=item.is_external,
        locked=item.locked,
        planned_start=item.start_time,
        planned_end=item.end_time,
        planned_minutes=planned_minutes,
        assigned_minutes=assigned_minutes,
        allocation_status=allocation_status,
        allocations=serialized_allocations,
    )


def _work_order_ticket_row(item: ProductionScheduleItem) -> WorkOrderTicketRow:
    operation = item.operation
    work_order = operation.work_order
    part = operation.part
    center = operation.work_center
    planned_minutes = scheduled_work_minutes(item.start_time, item.end_time)
    allocations = sorted(item.personnel_allocations, key=lambda row: (row.person.name, row.person_id))
    serialized_allocations = [
        WorkOrderTicketAllocation(
            person_id=allocation.person_id,
            employee_no=allocation.person.employee_no,
            person_name=allocation.person.name,
            ratio_percent=allocation.ratio_percent,
            planned_minutes=allocation.planned_minutes,
        )
        for allocation in allocations
    ]
    if not serialized_allocations:
        allocation_status = "unassigned"
        ticket_status = "pending_dispatch"
    elif abs(sum(allocation.ratio_percent for allocation in serialized_allocations) - 100) <= 0.001:
        allocation_status = "assigned"
        ticket_status = "ready_to_export"
    else:
        allocation_status = "partial"
        ticket_status = "needs_completion"

    return WorkOrderTicketRow(
        schedule_item_id=item.id,
        work_order_id=item.work_order_id,
        work_center_id=item.work_center_id,
        order_no=work_order.order_no,
        customer=work_order.customer,
        drawing_no=part.drawing_no,
        part_no=part.no,
        part_name=part.name,
        operation_name=operation.name,
        requirement_note=operation.requirement_note,
        work_center_name=center.name,
        machine_name=item.machine.name if item.machine else None,
        planned_start=item.start_time,
        planned_end=item.end_time,
        planned_minutes=planned_minutes,
        allocation_status=allocation_status,
        ticket_status=ticket_status,
        allocations=serialized_allocations,
    )


async def list_work_order_tickets(
    db: AsyncSession,
    schedule_id: int,
    work_order_id: int | None = None,
    work_center_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    ticket_status: str | None = None,
) -> WorkOrderTicketResponse:
    schedule = await db.get(ProductionSchedule, schedule_id)
    if schedule is None:
        raise ValueError("排产方案不存在。")

    query = (
        select(ProductionScheduleItem)
        .where(
            ProductionScheduleItem.schedule_id == schedule_id,
            ProductionScheduleItem.is_external == False,
        )
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.machine),
            selectinload(ProductionScheduleItem.personnel_allocations).selectinload(
                ProductionScheduleItemPersonnelAllocation.person
            ),
        )
        .order_by(
            ProductionScheduleItem.start_time,
            ProductionScheduleItem.work_center_id,
            ProductionScheduleItem.sequence_on_resource,
            ProductionScheduleItem.id,
        )
    )
    if work_order_id is not None:
        query = query.where(ProductionScheduleItem.work_order_id == work_order_id)
    if work_center_id is not None:
        query = query.where(ProductionScheduleItem.work_center_id == work_center_id)
    if date_from is not None:
        query = query.where(ProductionScheduleItem.start_time >= datetime.combine(date_from, time.min))
    if date_to is not None:
        query = query.where(ProductionScheduleItem.start_time < datetime.combine(date_to + timedelta(days=1), time.min))

    result = await db.execute(query)
    rows = [
        _work_order_ticket_row(item)
        for item in result.scalars().all()
        if not is_monitor_hidden_schedule_item(item)
    ]

    if ticket_status:
        if ticket_status not in {"pending_dispatch", "needs_completion", "ready_to_export"}:
            raise ValueError("ticket_status must be pending_dispatch, needs_completion or ready_to_export.")
        rows = [row for row in rows if row.ticket_status == ticket_status]

    return WorkOrderTicketResponse(schedule=schedule, tasks=rows)


async def get_dispatch_data(
    db: AsyncSession,
    schedule_id: int,
    work_order_id: int | None = None,
    work_center_id: int | None = None,
    person_id: int | None = None,
    allocation_status: str | None = None,
) -> DispatchResponse:
    schedule = await db.get(ProductionSchedule, schedule_id)
    if schedule is None:
        raise ValueError("排产方案不存在。")

    query = (
        select(ProductionScheduleItem)
        .where(
            ProductionScheduleItem.schedule_id == schedule_id,
            ProductionScheduleItem.is_external == False,
        )
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.machine),
            selectinload(ProductionScheduleItem.personnel_allocations).selectinload(
                ProductionScheduleItemPersonnelAllocation.person
            ),
        )
        .order_by(
            ProductionScheduleItem.start_time,
            ProductionScheduleItem.work_center_id,
            ProductionScheduleItem.sequence_on_resource,
            ProductionScheduleItem.id,
        )
    )
    if work_order_id is not None:
        query = query.where(ProductionScheduleItem.work_order_id == work_order_id)
    if work_center_id is not None:
        query = query.where(ProductionScheduleItem.work_center_id == work_center_id)

    result = await db.execute(query)
    rows = [
        _dispatch_row(item)
        for item in result.scalars().all()
        if not is_monitor_hidden_schedule_item(item)
    ]

    if person_id is not None:
        rows = [
            row
            for row in rows
            if any(allocation.person_id == person_id for allocation in row.allocations)
        ]

    if allocation_status:
        if allocation_status not in {"assigned", "unassigned", "partial"}:
            raise ValueError("allocation_status must be assigned, unassigned or partial.")
        rows = [row for row in rows if row.allocation_status == allocation_status]

    people_result = await db.execute(
        select(Personnel)
        .where(Personnel.status == "active")
        .order_by(Personnel.name, Personnel.employee_no)
    )
    people = [_serialize_personnel_option(person) for person in people_result.scalars().all()]
    return DispatchResponse(schedule=schedule, personnel=people, tasks=rows)


def _allocation_status_for_item(item: ProductionScheduleItem) -> str:
    allocations = item.personnel_allocations or []
    if not allocations:
        return "unassigned"
    if abs(sum(allocation.ratio_percent for allocation in allocations) - 100) <= 0.001:
        return "assigned"
    return "partial"


def _auto_assign_ratio_split(total_ratio: float, count: int) -> list[float]:
    if count <= 1:
        return [round(total_ratio, 2)]
    base = round(total_ratio / count, 2)
    ratios = [base for _ in range(count - 1)]
    ratios.append(round(total_ratio - sum(ratios), 2))
    return ratios


def _auto_assign_summary(tasks: list[DispatchAutoAssignTaskPreview]) -> DispatchAutoAssignSummary:
    processable = [task for task in tasks if not task.skipped]
    return DispatchAutoAssignSummary(
        processable_count=len(processable),
        skipped_count=len(tasks) - len(processable),
        multi_person_count=sum(1 for task in processable if task.multi_person),
        cross_work_center_count=sum(1 for task in processable if task.cross_work_center),
    )


async def _build_auto_assign_preview(
    db: AsyncSession,
    schedule_id: int,
    request: DispatchAutoAssignRequest,
) -> DispatchAutoAssignResponse:
    schedule = await db.get(ProductionSchedule, schedule_id)
    if schedule is None:
        raise ValueError("排产方案不存在。")

    query = (
        select(ProductionScheduleItem)
        .where(
            ProductionScheduleItem.schedule_id == schedule_id,
            ProductionScheduleItem.is_external == False,
        )
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.personnel_allocations).selectinload(
                ProductionScheduleItemPersonnelAllocation.person
            ),
        )
        .order_by(ProductionScheduleItem.start_time, ProductionScheduleItem.id)
    )
    if request.work_order_id is not None:
        query = query.where(ProductionScheduleItem.work_order_id == request.work_order_id)
    if request.work_center_id is not None:
        query = query.where(ProductionScheduleItem.work_center_id == request.work_center_id)

    result = await db.execute(query)
    items = [
        item
        for item in result.scalars().all()
        if not is_monitor_hidden_schedule_item(item)
    ]

    if request.person_id is not None:
        items = [
            item
            for item in items
            if any(allocation.person_id == request.person_id for allocation in item.personnel_allocations)
        ]

    allowed_statuses = {"unassigned", "partial"}
    if request.allocation_status:
        if request.allocation_status not in {"assigned", "unassigned", "partial"}:
            raise ValueError("allocation_status must be assigned, unassigned or partial.")
        allowed_statuses &= {request.allocation_status}

    normalized_query = (request.query or "").strip().lower()
    if normalized_query:
        items = [
            item
            for item in items
            if any(
                normalized_query in str(value).lower()
                for value in [
                    item.operation.work_order.order_no,
                    item.operation.work_order.customer,
                    item.operation.part.drawing_no,
                    item.operation.part.no,
                    item.operation.part.name,
                    item.operation.name,
                    item.operation.work_center.name,
                    " / ".join(
                        f"{allocation.person.name} {allocation.ratio_percent}%"
                        for allocation in item.personnel_allocations
                    ),
                ]
                if value
            )
        ]

    active_people_result = await db.execute(
        select(Personnel)
        .where(Personnel.status == "active")
        .options(selectinload(Personnel.work_centers))
        .order_by(Personnel.name, Personnel.id)
    )
    active_people = list(active_people_result.scalars().all())
    person_center_ids = {
        person.id: {link.work_center_id for link in person.work_centers}
        for person in active_people
    }

    center_links = await _active_personnel_by_work_center(
        db,
        {item.work_center_id for item in items},
    )

    workload_result = await db.execute(
        select(
            ProductionScheduleItemPersonnelAllocation.person_id,
            ProductionScheduleItemPersonnelAllocation.planned_minutes,
        )
        .join(
            ProductionScheduleItem,
            ProductionScheduleItem.id == ProductionScheduleItemPersonnelAllocation.schedule_item_id,
        )
        .where(ProductionScheduleItem.schedule_id == schedule_id)
    )
    workload_minutes: dict[int, int] = defaultdict(int)
    for person_id, planned_minutes in workload_result.all():
        workload_minutes[person_id] += planned_minutes

    def choose_people(center_id: int, count: int, excluded_ids: set[int]) -> tuple[list[Personnel], bool]:
        center_people = [
            link.person
            for link in center_links.get(center_id, [])
            if link.person_id not in excluded_ids
        ]
        source = center_people
        cross_work_center = False
        if not source:
            source = [person for person in active_people if person.id not in excluded_ids]
            cross_work_center = True
        source = sorted(source, key=lambda person: (workload_minutes.get(person.id, 0), person.name, person.id))
        return source[:count], cross_work_center

    previews: list[DispatchAutoAssignTaskPreview] = []
    for item in items:
        status = _allocation_status_for_item(item)
        if status not in allowed_statuses:
            continue

        operation = item.operation
        center = operation.work_center
        existing_allocations = sorted(item.personnel_allocations, key=lambda row: (row.person.name, row.person_id))
        planned_minutes = _operation_person_minutes(operation)
        existing_ratio = sum(allocation.ratio_percent for allocation in existing_allocations)

        base_preview = {
            "schedule_item_id": item.id,
            "order_no": operation.work_order.order_no,
            "drawing_no": operation.part.drawing_no,
            "part_no": operation.part.no,
            "part_name": operation.part.name,
            "operation_name": operation.name,
            "work_center_name": center.name,
            "planned_minutes": planned_minutes,
            "allocation_status": status,
        }

        if any(
            allocation.ratio_percent <= 0 or allocation.ratio_percent > 100
            for allocation in existing_allocations
        ) or existing_ratio > 100:
            previews.append(
                DispatchAutoAssignTaskPreview(
                    **base_preview,
                    skipped=True,
                    skip_reason="已有分摊占比无效，请先手工调整。",
                )
            )
            continue

        remaining_ratio = round(100 - existing_ratio, 6)
        if remaining_ratio <= 0:
            previews.append(
                DispatchAutoAssignTaskPreview(
                    **base_preview,
                    skipped=True,
                    skip_reason="当前任务已无需补充分摊。",
                )
            )
            continue

        existing_person_ids = {allocation.person_id for allocation in existing_allocations}
        desired_total_people = 2 if planned_minutes > 480 else 1
        add_count = max(desired_total_people - len(existing_allocations), 1)
        selected_people, cross_fallback = choose_people(center.id, add_count, existing_person_ids)
        if not selected_people:
            previews.append(
                DispatchAutoAssignTaskPreview(
                    **base_preview,
                    skipped=True,
                    skip_reason="没有可用在职人员。",
                )
            )
            continue

        ratios = _auto_assign_ratio_split(remaining_ratio, len(selected_people))
        final_ratios = [allocation.ratio_percent for allocation in existing_allocations] + ratios
        final_minutes = _allocation_minutes_for_operation(operation, final_ratios)
        preview_allocations: list[DispatchAutoAssignAllocation] = []

        for allocation, minutes in zip(existing_allocations, final_minutes[:len(existing_allocations)]):
            preview_allocations.append(
                DispatchAutoAssignAllocation(
                    person_id=allocation.person_id,
                    employee_no=allocation.person.employee_no,
                    person_name=allocation.person.name,
                    ratio_percent=allocation.ratio_percent,
                    planned_minutes=minutes,
                    cross_work_center=center.id not in person_center_ids.get(allocation.person_id, set()),
                    existing=True,
                )
            )

        new_minutes = final_minutes[len(existing_allocations):]
        for person, ratio, minutes in zip(selected_people, ratios, new_minutes):
            is_cross = center.id not in person_center_ids.get(person.id, set())
            preview_allocations.append(
                DispatchAutoAssignAllocation(
                    person_id=person.id,
                    employee_no=person.employee_no,
                    person_name=person.name,
                    ratio_percent=ratio,
                    planned_minutes=minutes,
                    cross_work_center=is_cross,
                    existing=False,
                )
            )
            workload_minutes[person.id] += minutes

        previews.append(
            DispatchAutoAssignTaskPreview(
                **base_preview,
                allocations=preview_allocations,
                multi_person=len(preview_allocations) > 1,
                cross_work_center=any(allocation.cross_work_center for allocation in preview_allocations) or cross_fallback,
            )
        )

    return DispatchAutoAssignResponse(
        schedule=schedule,
        summary=_auto_assign_summary(previews),
        tasks=previews,
    )


async def preview_auto_assign_dispatch(
    db: AsyncSession,
    schedule_id: int,
    request: DispatchAutoAssignRequest,
) -> DispatchAutoAssignResponse:
    return await _build_auto_assign_preview(db, schedule_id, request)


async def apply_auto_assign_dispatch(
    db: AsyncSession,
    schedule_id: int,
    request: DispatchAutoAssignRequest,
) -> DispatchAutoAssignResponse:
    preview = await _build_auto_assign_preview(db, schedule_id, request)
    processable = [task for task in preview.tasks if not task.skipped]
    if not processable:
        return preview

    item_ids = [task.schedule_item_id for task in processable]
    await db.execute(
        sqla_delete(ProductionScheduleItemPersonnelAllocation)
        .where(ProductionScheduleItemPersonnelAllocation.schedule_item_id.in_(item_ids))
    )
    for task in processable:
        for allocation in task.allocations:
            db.add(
                ProductionScheduleItemPersonnelAllocation(
                    schedule_item_id=task.schedule_item_id,
                    person_id=allocation.person_id,
                    ratio_percent=allocation.ratio_percent,
                    planned_minutes=allocation.planned_minutes,
                )
            )

    await db.commit()
    preview.summary = _auto_assign_summary(preview.tasks)
    return preview


async def _recalculate_schedule_from_item(
    db: AsyncSession,
    anchor_item_id: int,
) -> int:
    anchor = await db.get(ProductionScheduleItem, anchor_item_id)
    if anchor is None:
        raise ValueError("排产明细不存在。")

    items_result = await db.execute(
        select(ProductionScheduleItem)
        .where(ProductionScheduleItem.schedule_id == anchor.schedule_id)
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.personnel_allocations).selectinload(
                ProductionScheduleItemPersonnelAllocation.person
            ),
        )
        .execution_options(populate_existing=True)
    )
    items = list(items_result.scalars().all())
    item_by_operation = {item.operation_id: item for item in items}
    fixed_items = [item for item in items if item.start_time < anchor.start_time and item.id != anchor_item_id]
    pending_by_operation = {
        item.operation_id: item
        for item in items
        if item.id == anchor_item_id or item.start_time >= anchor.start_time
    }
    if not pending_by_operation:
        return 0

    dependency_result = await db.execute(select(OperationDependency))
    dependencies = list(dependency_result.scalars().all())
    dep_map: dict[int, set[int]] = defaultdict(set)
    for dependency in dependencies:
        if dependency.operation_id in item_by_operation:
            if dependency.depends_on_operation_id not in item_by_operation:
                raise ValueError("排产方案存在缺失的前置工序，无法重算。")
            dep_map[dependency.operation_id].add(dependency.depends_on_operation_id)

    completed_end: dict[int, datetime] = {
        item.operation_id: item.end_time
        for item in fixed_items
    }
    work_center_ready: dict[int, datetime] = {}
    external_ready: dict[tuple[int, int], datetime] = {}
    sequence_counter: dict[str, int] = {}

    for item in sorted(fixed_items, key=lambda row: (row.start_time, row.id)):
        if item.is_external:
            _, _, _, resource_key = _schedule_external_on_slot(
                item.operation.work_center,
                item.start_time,
                scheduled_work_hours(item.start_time, item.end_time),
                external_ready,
                anchor.start_time,
                forced_end=item.end_time,
            )
        else:
            resource_key = f"work-center:{item.work_center_id}"
            work_center_ready[item.work_center_id] = max(
                work_center_ready.get(item.work_center_id, item.end_time),
                item.end_time,
            )
        sequence_counter[resource_key] = sequence_counter.get(resource_key, 0) + 1

    base_start = anchor.start_time
    recalculated_count = 0

    while pending_by_operation:
        ready = [
            item
            for item in pending_by_operation.values()
            if dep_map.get(item.operation_id, set()).issubset(completed_end.keys())
        ]
        if not ready:
            raise ValueError("Operation dependency graph contains a cycle or missing dependency.")

        item = sorted(ready, key=lambda row: _operation_priority_key(row.operation))[0]
        operation = item.operation
        center = operation.work_center
        dependency_end = max(
            [completed_end[dep_id] for dep_id in dep_map.get(item.operation_id, set())],
            default=base_start,
        )
        start_floor = max(base_start, dependency_end)

        if center.is_external:
            duration_h = effective_operation_duration_hours(operation) or _external_default_duration_hours(center)
            start_time_val, end_time, _, resource_key = _schedule_external_on_slot(
                center,
                start_floor,
                duration_h,
                external_ready,
                base_start,
                forced_end=_external_task_end_override(item),
            )
            item.external_expected_return_at = end_time
            if item.external_status in {None, ""}:
                item.external_status = "pending"
        else:
            duration_h = effective_operation_duration_hours(operation)
            start_time_val = next_work_time(
                max(start_floor, work_center_ready.get(center.id, base_start))
            )
            end_time = add_work_hours(start_time_val, duration_h)
            work_center_ready[center.id] = end_time
            resource_key = f"work-center:{center.id}"
            allocations = sorted(item.personnel_allocations, key=lambda row: (row.person.name, row.person_id))
            if allocations:
                planned_minutes = _allocation_minutes_for_operation(
                    operation,
                    [allocation.ratio_percent for allocation in allocations],
                )
                for allocation, minutes in zip(allocations, planned_minutes):
                    allocation.planned_minutes = minutes

        sequence_counter[resource_key] = sequence_counter.get(resource_key, 0) + 1
        item.start_time = start_time_val
        item.end_time = end_time
        item.sequence_on_resource = sequence_counter[resource_key]
        item.machine_id = None if not center.is_external else item.machine_id
        completed_end[item.operation_id] = end_time
        pending_by_operation.pop(item.operation_id)
        recalculated_count += 1

    return recalculated_count


async def save_personnel_allocations(
    db: AsyncSession,
    schedule_item_id: int,
    allocations: list[PersonnelAllocationWrite],
) -> PersonnelAllocationSaveResponse:
    item = await db.get(
        ProductionScheduleItem,
        schedule_item_id,
        options=[
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.personnel_allocations).selectinload(
                ProductionScheduleItemPersonnelAllocation.person
            ),
        ],
    )
    if item is None:
        raise ValueError("排产明细不存在。")
    if item.operation.work_center.is_external:
        raise ValueError("外协工段不占用内部人员，不需要派工分摊。")
    if not allocations:
        raise ValueError("至少需要分配一名人员，且占比合计必须为 100%。")

    person_ids = [allocation.person_id for allocation in allocations]
    if len(person_ids) != len(set(person_ids)):
        raise ValueError("同一个人员不能在同一任务中重复分配。")

    ratio_sum = 0.0
    for allocation in allocations:
        if allocation.ratio_percent <= 0 or allocation.ratio_percent > 100:
            raise ValueError("人员占比必须大于 0 且不超过 100%。")
        ratio_sum += allocation.ratio_percent
    if abs(ratio_sum - 100) > 0.001:
        raise ValueError("人员占比合计必须为 100%。")

    people_result = await db.execute(select(Personnel).where(Personnel.id.in_(person_ids)))
    people_by_id = {person.id: person for person in people_result.scalars().all()}
    missing_ids = [person_id for person_id in person_ids if person_id not in people_by_id]
    if missing_ids:
        raise ValueError(f"人员不存在：{', '.join(str(person_id) for person_id in missing_ids)}。")
    inactive = [person.name for person in people_by_id.values() if person.status != "active"]
    if inactive:
        raise ValueError(f"以下人员不是在职状态：{'、'.join(sorted(inactive))}。")
    planned_minutes = _allocation_minutes_for_operation(
        item.operation,
        [allocation.ratio_percent for allocation in allocations],
    )

    await db.execute(
        sqla_delete(ProductionScheduleItemPersonnelAllocation)
        .where(ProductionScheduleItemPersonnelAllocation.schedule_item_id == schedule_item_id)
    )
    for allocation, minutes in zip(allocations, planned_minutes):
        db.add(
            ProductionScheduleItemPersonnelAllocation(
                schedule_item_id=schedule_item_id,
                person_id=allocation.person_id,
                ratio_percent=allocation.ratio_percent,
                planned_minutes=minutes,
            )
        )

    await db.commit()
    refreshed_item_result = await db.execute(
        select(ProductionScheduleItem)
        .where(ProductionScheduleItem.id == schedule_item_id)
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.machine),
            selectinload(ProductionScheduleItem.schedule),
            selectinload(ProductionScheduleItem.personnel_allocations).selectinload(
                ProductionScheduleItemPersonnelAllocation.person
            ),
        )
        .execution_options(populate_existing=True)
    )
    refreshed_item = refreshed_item_result.scalars().one()
    refreshed_allocations = sorted(
        refreshed_item.personnel_allocations,
        key=lambda allocation: (allocation.person.name, allocation.person_id),
    )
    return PersonnelAllocationSaveResponse(
        schedule=refreshed_item.schedule,
        task=_dispatch_row(refreshed_item),
        allocations=[_serialize_allocation(allocation) for allocation in refreshed_allocations],
    )


async def save_batch_personnel_allocations(
    db: AsyncSession,
    schedule_id: int,
    request: PersonnelBatchAllocationRequest,
) -> PersonnelBatchAllocationResponse:
    schedule = await db.get(ProductionSchedule, schedule_id)
    if schedule is None:
        raise ValueError("排产方案不存在。")

    requested_ids = list(dict.fromkeys(request.schedule_item_ids))
    if not requested_ids:
        raise ValueError("至少需要选择一条排产明细。")
    if not request.allocations:
        raise ValueError("至少需要分配一名人员，且占比合计必须为 100%。")

    person_ids = [allocation.person_id for allocation in request.allocations]
    if len(person_ids) != len(set(person_ids)):
        raise ValueError("同一个人员不能在同一任务中重复分配。")

    ratio_sum = 0.0
    for allocation in request.allocations:
        if allocation.ratio_percent <= 0 or allocation.ratio_percent > 100:
            raise ValueError("人员占比必须大于 0 且不超过 100%。")
        ratio_sum += allocation.ratio_percent
    if abs(ratio_sum - 100) > 0.001:
        raise ValueError("人员占比合计必须为 100%。")

    people_result = await db.execute(select(Personnel).where(Personnel.id.in_(person_ids)))
    people_by_id = {person.id: person for person in people_result.scalars().all()}
    missing_ids = [person_id for person_id in person_ids if person_id not in people_by_id]
    if missing_ids:
        raise ValueError(f"人员不存在：{', '.join(str(person_id) for person_id in missing_ids)}。")
    inactive = [person.name for person in people_by_id.values() if person.status != "active"]
    if inactive:
        raise ValueError(f"以下人员不是在职状态：{'、'.join(sorted(inactive))}。")

    result = await db.execute(
        select(ProductionScheduleItem)
        .where(
            ProductionScheduleItem.schedule_id == schedule_id,
            ProductionScheduleItem.id.in_(requested_ids),
        )
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.personnel_allocations).selectinload(
                ProductionScheduleItemPersonnelAllocation.person
            ),
        )
        .order_by(ProductionScheduleItem.start_time, ProductionScheduleItem.id)
    )
    items = list(result.scalars().all())
    items_by_id = {item.id: item for item in items}
    skipped: list[PersonnelBatchAllocationSkippedItem] = []
    processed: list[ProductionScheduleItem] = []

    for schedule_item_id in requested_ids:
        item = items_by_id.get(schedule_item_id)
        if item is None:
            skipped.append(
                PersonnelBatchAllocationSkippedItem(
                    schedule_item_id=schedule_item_id,
                    reason="排产明细不存在或不属于当前方案。",
                )
            )
            continue
        operation = item.operation
        if operation.work_center.is_external:
            skipped.append(
                PersonnelBatchAllocationSkippedItem(
                    schedule_item_id=item.id,
                    order_no=operation.work_order.order_no,
                    operation_name=operation.name,
                    reason="外协工段不占用内部人员。",
                )
            )
            continue
        if not request.overwrite_assigned and _allocation_status_for_item(item) == "assigned":
            skipped.append(
                PersonnelBatchAllocationSkippedItem(
                    schedule_item_id=item.id,
                    order_no=operation.work_order.order_no,
                    operation_name=operation.name,
                    reason="已完整派工，默认不覆盖。",
                )
            )
            continue
        processed.append(item)

    if processed:
        processed_ids = [item.id for item in processed]
        await db.execute(
            sqla_delete(ProductionScheduleItemPersonnelAllocation)
            .where(ProductionScheduleItemPersonnelAllocation.schedule_item_id.in_(processed_ids))
        )
        for item in processed:
            planned_minutes = _allocation_minutes_for_operation(
                item.operation,
                [allocation.ratio_percent for allocation in request.allocations],
            )
            for allocation, minutes in zip(request.allocations, planned_minutes):
                db.add(
                    ProductionScheduleItemPersonnelAllocation(
                        schedule_item_id=item.id,
                        person_id=allocation.person_id,
                        ratio_percent=allocation.ratio_percent,
                        planned_minutes=minutes,
                    )
                )

        await db.commit()
        schedule = await db.get(ProductionSchedule, schedule_id)
    return PersonnelBatchAllocationResponse(
        schedule=schedule,
        processed_count=len(processed),
        skipped_count=len(skipped),
        skipped_items=skipped,
    )


async def get_personnel_workload(db: AsyncSession, schedule_id: int) -> PersonnelWorkloadResponse:
    schedule = await db.get(ProductionSchedule, schedule_id)
    if schedule is None:
        raise ValueError("排产方案不存在。")

    result = await db.execute(
        select(ProductionScheduleItemPersonnelAllocation)
        .join(
            ProductionScheduleItem,
            ProductionScheduleItem.id == ProductionScheduleItemPersonnelAllocation.schedule_item_id,
        )
        .where(ProductionScheduleItem.schedule_id == schedule_id)
        .options(
            selectinload(ProductionScheduleItemPersonnelAllocation.person),
            selectinload(ProductionScheduleItemPersonnelAllocation.schedule_item)
            .selectinload(ProductionScheduleItem.operation)
            .selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItemPersonnelAllocation.schedule_item)
            .selectinload(ProductionScheduleItem.operation)
            .selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItemPersonnelAllocation.schedule_item)
            .selectinload(ProductionScheduleItem.operation)
            .selectinload(ProductionOperation.work_center),
        )
        .order_by(ProductionScheduleItemPersonnelAllocation.person_id, ProductionScheduleItem.start_time)
    )
    grouped: dict[int, dict] = {}
    for allocation in result.scalars().all():
        item = allocation.schedule_item
        if is_monitor_hidden_schedule_item(item):
            continue
        operation = item.operation
        work_order = operation.work_order
        part = operation.part
        center = operation.work_center
        row = grouped.setdefault(
            allocation.person_id,
            {
                "person": allocation.person,
                "planned_minutes": 0,
                "orders": set(),
                "work_centers": set(),
                "tasks": [],
            },
        )
        row["planned_minutes"] += allocation.planned_minutes
        row["orders"].add(work_order.id)
        row["work_centers"].add(center.id)
        row["tasks"].append(
            PersonnelWorkloadTask(
                schedule_item_id=item.id,
                work_order_id=work_order.id,
                order_no=work_order.order_no,
                drawing_no=part.drawing_no,
                part_no=part.no,
                operation_name=operation.name,
                work_center_name=center.name,
                planned_start=item.start_time,
                planned_end=item.end_time,
                ratio_percent=allocation.ratio_percent,
                planned_minutes=allocation.planned_minutes,
            )
        )

    rows: list[PersonnelWorkloadRow] = []
    for data in grouped.values():
        person = data["person"]
        rows.append(
            PersonnelWorkloadRow(
                person_id=person.id,
                employee_no=person.employee_no,
                person_name=person.name,
                task_count=len(data["tasks"]),
                planned_minutes=data["planned_minutes"],
                order_count=len(data["orders"]),
                work_center_count=len(data["work_centers"]),
                tasks=data["tasks"],
            )
        )
    rows.sort(key=lambda row: row.planned_minutes, reverse=True)
    return PersonnelWorkloadResponse(schedule=schedule, rows=rows)


async def export_personnel_workload_to_excel(db: AsyncSession, schedule_id: int) -> tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    response = await get_personnel_workload(db, schedule_id)
    schedule = response.schedule

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D5D8C", end_color="2D5D8C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, col_count: int) -> None:
        for col in range(1, col_count + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def style_body_row(ws, row_idx: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            ws.cell(row_idx, col).border = thin_border

    def set_widths(ws, widths: list[int]) -> None:
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    wb = Workbook()

    summary = wb.active
    summary.title = "人员工时汇总"
    summary_headers = ["人员", "工号", "任务数", "计划工时(小时)", "订单数", "工段数"]
    summary.append(summary_headers)
    style_header(summary, len(summary_headers))
    for row_idx, row in enumerate(response.rows, 2):
        summary.append([
            row.person_name,
            row.employee_no,
            row.task_count,
            round(row.planned_minutes / 60, 2),
            row.order_count,
            row.work_center_count,
        ])
        style_body_row(summary, row_idx, len(summary_headers))
    set_widths(summary, [18, 16, 12, 16, 12, 12])

    detail = wb.create_sheet("任务明细")
    detail_headers = [
        "人员",
        "工号",
        "订单号",
        "图号",
        "零件号",
        "工序",
        "工段",
        "计划开始",
        "计划结束",
        "占比",
        "计划工时(小时)",
    ]
    detail.append(detail_headers)
    style_header(detail, len(detail_headers))
    row_idx = 2
    for row in response.rows:
        for task in row.tasks:
            detail.append([
                row.person_name,
                row.employee_no,
                task.order_no,
                task.drawing_no,
                task.part_no,
                task.operation_name,
                task.work_center_name,
                task.planned_start.strftime("%Y-%m-%d %H:%M"),
                task.planned_end.strftime("%Y-%m-%d %H:%M"),
                f"{task.ratio_percent:g}%",
                round(task.planned_minutes / 60, 2),
            ])
            style_body_row(detail, row_idx, len(detail_headers))
            row_idx += 1
    set_widths(detail, [18, 16, 16, 18, 16, 18, 16, 18, 18, 10, 16])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"人员工时汇总_{schedule.schedule_no or schedule_id}.xlsx"
    return buffer.getvalue(), filename


async def export_work_order_tickets_to_excel(
    db: AsyncSession,
    schedule_id: int,
    work_order_id: int | None = None,
    work_center_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    response = await list_work_order_tickets(
        db,
        schedule_id,
        work_order_id=work_order_id,
        work_center_id=work_center_id,
        date_from=date_from,
        date_to=date_to,
        ticket_status="ready_to_export",
    )
    schedule = response.schedule
    tasks = response.tasks
    if not tasks:
        raise ValueError("当前筛选条件下没有已确认派工任务，无法生成加工单。")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D5D8C", end_color="2D5D8C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, col_count: int) -> None:
        for col in range(1, col_count + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def style_body_row(ws, row_idx: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            cell = ws.cell(row_idx, col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    def set_widths(ws, widths: list[int]) -> None:
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    def hours(minutes: int) -> float:
        return round((minutes or 0) / 60, 2)

    wb = Workbook()
    tickets = wb.active
    tickets.title = "加工单明细"
    ticket_headers = [
        "加工单号",
        "排产方案",
        "订单号",
        "客户",
        "图号",
        "零件号",
        "零件名称",
        "工序",
        "工段",
        "设备",
        "计划开始",
        "计划结束",
        "工时(小时)",
        "分配人员",
        "加工要求",
    ]
    tickets.append(ticket_headers)
    style_header(tickets, len(ticket_headers))
    for row_idx, task in enumerate(tasks, 2):
        ticket_no = f"JG-{schedule.schedule_no}-{task.schedule_item_id}"
        allocations = " / ".join(
            f"{allocation.person_name} {allocation.ratio_percent:g}% ({hours(allocation.planned_minutes)}h)"
            for allocation in task.allocations
        )
        tickets.append([
            ticket_no,
            schedule.schedule_no,
            task.order_no,
            task.customer,
            task.drawing_no,
            task.part_no,
            task.part_name,
            task.operation_name,
            task.work_center_name,
            task.machine_name or "",
            task.planned_start.strftime("%Y-%m-%d %H:%M"),
            task.planned_end.strftime("%Y-%m-%d %H:%M"),
            hours(task.planned_minutes),
            allocations,
            task.requirement_note or "",
        ])
        style_body_row(tickets, row_idx, len(ticket_headers))
    set_widths(tickets, [20, 18, 18, 16, 18, 16, 18, 16, 14, 14, 18, 18, 12, 34, 36])

    detail = wb.create_sheet("人员任务明细")
    detail_headers = [
        "加工单号",
        "人员",
        "工号",
        "分摊比例",
        "分摊工时(小时)",
        "订单号",
        "零件名称",
        "工序",
        "工段",
        "计划开始",
        "计划结束",
    ]
    detail.append(detail_headers)
    style_header(detail, len(detail_headers))
    detail_row = 2
    for task in tasks:
        ticket_no = f"JG-{schedule.schedule_no}-{task.schedule_item_id}"
        for allocation in task.allocations:
            detail.append([
                ticket_no,
                allocation.person_name,
                allocation.employee_no,
                f"{allocation.ratio_percent:g}%",
                hours(allocation.planned_minutes),
                task.order_no,
                task.part_name,
                task.operation_name,
                task.work_center_name,
                task.planned_start.strftime("%Y-%m-%d %H:%M"),
                task.planned_end.strftime("%Y-%m-%d %H:%M"),
            ])
            style_body_row(detail, detail_row, len(detail_headers))
            detail_row += 1
    set_widths(detail, [20, 14, 14, 12, 16, 18, 18, 16, 14, 18, 18])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"加工单_{schedule.schedule_no or schedule_id}_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    export_record = ExportBatch(
        export_type="work_order_ticket",
        schedule_id=schedule_id,
        filename=filename,
        params_json=json.dumps(
            {
                "work_order_id": work_order_id,
                "work_center_id": work_center_id,
                "date_from": date_from.isoformat() if date_from else None,
                "date_to": date_to.isoformat() if date_to else None,
                "schedule_item_ids": [task.schedule_item_id for task in tasks],
            },
            ensure_ascii=False,
        ),
    )
    db.add(export_record)
    await db.commit()
    return buffer.getvalue(), filename


async def export_construction_sheets_to_excel(
    db: AsyncSession,
    schedule_id: int,
    work_order_id: int | None = None,
    work_center_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    schedule = await db.get(ProductionSchedule, schedule_id)
    if schedule is None:
        raise ValueError("排产方案不存在。")

    query = (
        select(ProductionScheduleItem)
        .where(
            ProductionScheduleItem.schedule_id == schedule_id,
            ProductionScheduleItem.is_external == False,
        )
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.machine),
        )
        .order_by(
            ProductionScheduleItem.work_order_id,
            ProductionScheduleItem.part_id,
            ProductionScheduleItem.start_time,
            ProductionScheduleItem.sequence_on_resource,
            ProductionScheduleItem.id,
        )
    )
    if work_order_id is not None:
        query = query.where(ProductionScheduleItem.work_order_id == work_order_id)
    if work_center_id is not None:
        query = query.where(ProductionScheduleItem.work_center_id == work_center_id)
    if date_from is not None:
        query = query.where(ProductionScheduleItem.start_time >= datetime.combine(date_from, time.min))
    if date_to is not None:
        query = query.where(ProductionScheduleItem.start_time < datetime.combine(date_to + timedelta(days=1), time.min))

    result = await db.execute(query)
    items = result.scalars().all()
    if not items:
        raise ValueError("当前筛选条件下没有已排产的内部工序，无法生成施工单。")

    def construction_key(item: ProductionScheduleItem) -> tuple[int, int]:
        return (item.work_order_id, item.part_id)

    grouped_items: dict[tuple[int, int], list[ProductionScheduleItem]] = defaultdict(list)
    for item in items:
        grouped_items[construction_key(item)].append(item)

    def group_sort_key(group: list[ProductionScheduleItem]) -> tuple[str, int, str, str]:
        first = group[0]
        operation = first.operation
        part = operation.part
        return (
            operation.work_order.order_no,
            part.source_row or 0,
            part.no or "",
            part.drawing_no or "",
        )

    ordered_groups = sorted(grouped_items.values(), key=group_sort_key)
    if not ordered_groups:
        raise ValueError("当前筛选条件下没有可导出的施工单。")

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    title_font = Font(name="SimSun", size=18, bold=True)
    company_font = Font(name="SimSun", size=14, bold=True)
    header_font = Font(name="SimSun", size=11, bold=True)
    body_font = Font(name="SimSun", size=11)
    small_font = Font(name="SimSun", size=10)
    barcode_font = Font(name="C39HrP24DhTt", size=22)
    header_fill = PatternFill(start_color="E9EEF5", end_color="E9EEF5", fill_type="solid")
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_border = Border(left=thin, right=thin, top=medium, bottom=thin)

    def safe_sheet_title(base: str, used: set[str]) -> str:
        cleaned = re.sub(r"[:\\/?*\[\]]+", "-", base).strip() or "施工单"
        title = cleaned[:31]
        if title not in used:
            used.add(title)
            return title
        suffix = 2
        while True:
            suffix_text = f"-{suffix}"
            candidate = f"{cleaned[:31 - len(suffix_text)]}{suffix_text}"
            if candidate not in used:
                used.add(candidate)
                return candidate
            suffix += 1

    def fmt_day(value: datetime | None) -> str:
        return value.strftime("%Y-%m-%d") if value else ""

    def fmt_hours(value: float | int | None) -> float:
        return round(float(value or 0), 2)

    def merge(ws, start: str, end: str, value: Any = None) -> None:
        ws.merge_cells(f"{start}:{end}")
        if value is not None:
            ws[start] = value

    def style_range(ws, row_start: int, row_end: int, col_start: int = 1, col_end: int = 13) -> None:
        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row, col)
                cell.border = border
                cell.font = body_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    used_titles: set[str] = set()
    exported_item_ids: list[int] = []
    for group in ordered_groups:
        group.sort(key=lambda item: (item.operation.seq_no, item.start_time, item.sequence_on_resource, item.id))
        first = group[0]
        operation = first.operation
        work_order = operation.work_order
        part = operation.part
        title = safe_sheet_title(f"{work_order.order_no}-{part.drawing_no or part.no}", used_titles)
        ws = wb.create_sheet(title=title)

        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.36
        ws.page_margins.bottom = 0.42
        ws.print_title_rows = "1:6"

        widths = [8, 14, 16, 16, 16, 12, 13, 13, 12, 10, 10, 24, 18]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        merge(ws, "A1", "D1", "上海光灿机械制造有限公司")
        merge(ws, "E1", "J1", "零 件 工 艺 施 工 单")
        ws["K1"] = "图号"
        merge(ws, "L1", "M1", part.drawing_no or part.no or "")
        merge(ws, "A2", "D2", "GC-5-02-1")
        merge(ws, "E2", "J2", f"排产方案：{schedule.schedule_no}")
        ws["K2"] = "名称"
        merge(ws, "L2", "M2", part.name or "")
        ws["A3"] = "材料牌号"
        merge(ws, "B3", "D3", part.material or "")
        ws["E3"] = "材料重量"
        ws["F3"] = part.material_weight or ""
        merge(ws, "G3", "J3", "每件毛坯加工件数")
        ws["K3"] = "订单号，工号"
        merge(ws, "L3", "M3", work_order.order_no)
        ws["A4"] = "毛坯种类"
        merge(ws, "B4", "D4", "")
        ws["E4"] = "订单数量"
        ws["F4"] = work_order.quantity
        merge(ws, "G4", "J4", f"零件数量：{part.quantity}")
        ws["K4"] = "要求完成日期"
        merge(ws, "L4", "M4", fmt_day(work_order.due_date))

        merge(ws, "A5", "A6", "序号")
        merge(ws, "B5", "B6", "工 序")
        merge(ws, "C5", "E6", "工序重点说明，夹具，刀具，量具准备")
        merge(ws, "F5", "F6", "单件 H工时")
        merge(ws, "G5", "G6", "工作日期")
        merge(ws, "H5", "H6", "完工日期")
        merge(ws, "I5", "I6", "操作员")
        merge(ws, "J5", "K5", "检验")
        ws["J6"] = "合格数"
        ws["K6"] = "检验员"
        merge(ws, "L5", "L6", "关键尺寸/备注")
        ws["M5"] = "条码"
        ws["M6"] = "图号，工序"

        style_range(ws, 1, 6)
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws["A1"].font = company_font
        ws["E1"].font = title_font
        for row in range(5, 7):
            for col in range(1, 14):
                cell = ws.cell(row, col)
                cell.font = header_font
                cell.fill = header_fill

        row_idx = 7
        for index, item in enumerate(group, 1):
            op = item.operation
            ticket_no = f"JG-{schedule.schedule_no}-{item.id}"
            exported_item_ids.append(item.id)
            values = {
                1: index,
                2: op.name,
                3: op.requirement_note or "",
                6: fmt_hours(op.duration_hours),
                7: fmt_day(item.start_time),
                8: "",
                9: "",
                10: "",
                11: "",
                12: part.note or "",
                13: f"*{ticket_no}*",
            }
            for col, value in values.items():
                ws.cell(row_idx, col).value = value
            merge(ws, f"C{row_idx}", f"E{row_idx}")
            style_range(ws, row_idx, row_idx)
            for col in range(1, 14):
                ws.cell(row_idx, col).border = section_border if index == 1 else border
            ws.cell(row_idx, 2).font = header_font
            ws.cell(row_idx, 3).font = small_font
            ws.cell(row_idx, 12).font = small_font
            ws.cell(row_idx, 13).font = barcode_font
            ws.cell(row_idx, 3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row_idx, 12).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[row_idx].height = 34
            row_idx += 1

        signature_row = max(row_idx, 25)
        for blank_row in range(row_idx, signature_row):
            style_range(ws, blank_row, blank_row)
            ws.row_dimensions[blank_row].height = 28
        merge(ws, f"A{signature_row}", f"M{signature_row}", "制表：")
        style_range(ws, signature_row, signature_row)
        ws.cell(signature_row, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[signature_row].height = 28
        ws.print_area = f"A1:M{signature_row}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"施工单_{schedule.schedule_no or schedule_id}_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    export_record = ExportBatch(
        export_type="construction_sheet",
        schedule_id=schedule_id,
        filename=filename,
        params_json=json.dumps(
            {
                "work_order_id": work_order_id,
                "work_center_id": work_center_id,
                "date_from": date_from.isoformat() if date_from else None,
                "date_to": date_to.isoformat() if date_to else None,
                "schedule_item_ids": exported_item_ids,
                "sheet_count": len(ordered_groups),
            },
            ensure_ascii=False,
        ),
    )
    db.add(export_record)
    await db.commit()
    return buffer.getvalue(), filename


async def list_pending_operations(db: AsyncSession) -> list[dict]:
    result = await db.execute(
        select(ProductionOperation)
        .join(WorkCenter, ProductionOperation.work_center_id == WorkCenter.id)
        .where(
            ProductionOperation.status.in_(["pending", "scheduled"]),
            WorkCenter.status == "active",
        )
        .options(
            selectinload(ProductionOperation.work_order),
            selectinload(ProductionOperation.part),
            selectinload(ProductionOperation.work_center),
        )
        .order_by(ProductionOperation.work_order_id, ProductionOperation.seq_no, ProductionOperation.id)
    )
    rows = []
    for operation in result.scalars().all():
        if is_monitor_hidden_operation(operation):
            continue
        rows.append(
            {
                "id": operation.id,
                "work_order_id": operation.work_order_id,
                "part_id": operation.part_id,
                "work_center_id": operation.work_center_id,
                "name": operation.name,
                "seq_no": operation.seq_no,
                "duration_hours": operation.duration_hours,
                "requirement_note": operation.requirement_note,
                "part_quantity": operation.part.quantity,
                "effective_duration_hours": effective_operation_duration_hours(operation),
                "status": operation.status,
                "order_no": operation.work_order.order_no,
                "part_no": operation.part.no,
                "drawing_no": operation.part.drawing_no,
                "part_name": operation.part.name,
                "work_center_name": operation.work_center.name,
                "due_date": operation.work_order.due_date,
            }
        )
    return rows


def next_work_time(moment: datetime) -> datetime:
    current = normalize_schedule_datetime(moment)
    while current.weekday() == 6:
        current = datetime.combine(current.date() + timedelta(days=1), WORK_START)
    if current.time() < WORK_START:
        return datetime.combine(current.date(), WORK_START)
    if LUNCH_START <= current.time() < LUNCH_END:
        return datetime.combine(current.date(), LUNCH_END)
    if current.time() >= WORK_END:
        return next_work_time(datetime.combine(current.date() + timedelta(days=1), WORK_START))
    return current


def add_work_hours(start: datetime, hours: float) -> datetime:
    remaining_minutes = max(int(round(hours * 60)), 1)
    current = next_work_time(start)
    while remaining_minutes > 0:
        segment_end_time = LUNCH_START if current.time() < LUNCH_START else WORK_END
        segment_end = datetime.combine(current.date(), segment_end_time)
        available = max(int((segment_end - current).total_seconds() // 60), 0)
        if available <= 0:
            current = next_work_time(current + timedelta(minutes=1))
            continue
        used = min(available, remaining_minutes)
        current += timedelta(minutes=used)
        remaining_minutes -= used
        if remaining_minutes > 0:
            current = next_work_time(current + timedelta(minutes=1))
    return current


def effective_operation_duration_hours(operation: ProductionOperation) -> float:
    quantity = max(operation.part.quantity if operation.part else 1, 1)
    return round((operation.duration_hours or 0) * quantity, 3)


def scheduled_work_hours(start: datetime, end: datetime) -> float:
    if end <= start:
        return 0

    current_day = start.date()
    last_day = end.date()
    total = 0.0
    while current_day <= last_day:
        total += work_hours_on_day(start, end, current_day)
        current_day += timedelta(days=1)
    return round(total, 3)


def scheduled_work_minutes(start: datetime, end: datetime) -> int:
    return max(int(round(scheduled_work_hours(start, end) * 60)), 1)


def _allocation_planned_minutes(total_minutes: int, ratios: list[float]) -> list[int]:
    planned: list[int] = []
    used = 0
    for index, ratio in enumerate(ratios):
        if index == len(ratios) - 1:
            minutes = max(total_minutes - used, 0)
        else:
            minutes = max(int(round(total_minutes * ratio / 100)), 0)
            used += minutes
        planned.append(minutes)
    return planned


def _operation_person_minutes(operation: ProductionOperation) -> int:
    return max(int(round(effective_operation_duration_hours(operation) * 60)), 1)


def _operation_duration_for_people(operation: ProductionOperation, person_count: int) -> float:
    return round(effective_operation_duration_hours(operation) / max(person_count, 1), 3)


def _external_slot_count(center: WorkCenter) -> int:
    return max(int(center.external_capacity_slots or 1), 1)


def _external_default_duration_hours(center: WorkCenter) -> float:
    return float(center.external_lead_time_hours or center.default_duration_hours or 8)


def _external_task_end_override(item: ProductionScheduleItem) -> datetime | None:
    if item.external_status == "returned" and item.external_returned_at:
        return item.external_returned_at
    if item.external_status in {"sent", "exception"} and item.external_expected_return_at:
        return item.external_expected_return_at
    return None


def _schedule_external_on_slot(
    center: WorkCenter,
    start_floor: datetime,
    duration_h: float,
    external_ready: dict[tuple[int, int], datetime],
    base_start: datetime,
    forced_end: datetime | None = None,
) -> tuple[datetime, datetime, int, str]:
    candidates = []
    for slot_index in range(_external_slot_count(center)):
        proposed = max(start_floor, external_ready.get((center.id, slot_index), base_start))
        actual_start = next_work_time(proposed)
        if forced_end and forced_end > actual_start:
            actual_end = forced_end
        else:
            actual_end = add_work_hours(actual_start, duration_h)
        candidates.append((actual_end, actual_start, slot_index))
    end_time, start_time_val, slot_index = min(candidates)
    external_ready[(center.id, slot_index)] = end_time
    return start_time_val, end_time, slot_index, f"external-{center.id}:{slot_index}"


def _allocation_minutes_for_operation(
    operation: ProductionOperation,
    ratios: list[float],
) -> list[int]:
    return _allocation_planned_minutes(_operation_person_minutes(operation), ratios)


def _operation_priority_key(operation: ProductionOperation) -> tuple:
    return (
        -operation.work_order.priority,
        operation.work_order.due_date,
        operation.work_order.created_at,
        operation.part.no,
        operation.seq_no,
        operation.id,
    )


async def _active_personnel_by_work_center(
    db: AsyncSession,
    work_center_ids: set[int],
) -> dict[int, list[WorkCenterPersonnel]]:
    if not work_center_ids:
        return {}
    result = await db.execute(
        select(WorkCenterPersonnel)
        .where(WorkCenterPersonnel.work_center_id.in_(work_center_ids))
        .join(Personnel, WorkCenterPersonnel.person_id == Personnel.id)
        .where(Personnel.status == "active")
        .options(selectinload(WorkCenterPersonnel.person))
        .order_by(
            WorkCenterPersonnel.work_center_id,
            WorkCenterPersonnel.sort_order,
            Personnel.name,
            Personnel.id,
        )
    )
    people_by_center: dict[int, list[WorkCenterPersonnel]] = defaultdict(list)
    for link in result.scalars().all():
        people_by_center[link.work_center_id].append(link)
    return people_by_center


async def run_production_scheduling(
    db: AsyncSession,
    start_time: datetime | None = None,
    work_order_ids: list[int] | None = None,
    base_schedule_id: int | None = None,
    keep_locked: bool = True,
) -> ProductionSchedulingResult:
    # Determine scheduling start time
    if start_time:
        now = next_work_time(start_time)
    else:
        now = next_work_time(datetime.utcnow())

    # Load pending/scheduled operations with optional order filter
    operation_query = (
        select(ProductionOperation)
        .join(WorkCenter, ProductionOperation.work_center_id == WorkCenter.id)
        .where(
            ProductionOperation.status.in_(["pending", "scheduled"]),
            WorkCenter.status == "active",
        )
        .options(
            selectinload(ProductionOperation.work_order),
            selectinload(ProductionOperation.part),
            selectinload(ProductionOperation.work_center).selectinload(WorkCenter.machines),
            selectinload(ProductionOperation.dependencies),
        )
    )
    if work_order_ids:
        operation_query = operation_query.where(
            ProductionOperation.work_order_id.in_(work_order_ids)
        )

    operation_result = await db.execute(operation_query)
    all_operations = list(operation_result.scalars().all())
    if not all_operations:
        raise ValueError("没有找到待排工序。请先导入工单或检查订单范围。")
    # If work_order_ids specified, validate orders exist and are pending
    if work_order_ids:
        order_result = await db.execute(
            select(WorkOrder).where(WorkOrder.id.in_(work_order_ids))
        )
        found_orders = {o.id: o for o in order_result.scalars().all()}
        missing = set(work_order_ids) - set(found_orders.keys())
        if missing:
            raise ValueError(f"以下订单不存在：{', '.join(str(i) for i in sorted(missing))}。")
        non_schedulable = [o for o in found_orders.values() if o.status not in ("pending", "scheduled")]
        if non_schedulable:
            raise ValueError(
                f"以下订单当前状态不可排产：{'、'.join(o.order_no for o in non_schedulable)}。"
                "只有待排或已排状态的订单可以参与排产。"
            )

    # Build dependency graph and validate before creating schedule
    dependency_result = await db.execute(select(OperationDependency))
    dependencies = list(dependency_result.scalars().all())
    pending_by_id = {op.id: op for op in all_operations}
    dep_map: dict[int, set[int]] = defaultdict(set)
    for dependency in dependencies:
        if dependency.operation_id in pending_by_id and dependency.depends_on_operation_id in pending_by_id:
            dep_map[dependency.operation_id].add(dependency.depends_on_operation_id)

    # Check for dependencies pointing to operations on disabled work centers
    blocked_op_ids = {
        d.depends_on_operation_id for d in dependencies
        if d.operation_id in pending_by_id and d.depends_on_operation_id not in pending_by_id
    }
    if blocked_op_ids:
        blocked_ops = await db.execute(
            select(ProductionOperation)
            .where(ProductionOperation.id.in_(blocked_op_ids))
            .options(selectinload(ProductionOperation.work_center))
        )
        blocked_names = {op.work_center.name for op in blocked_ops.scalars().all()}
        affected_orders = {
            op.work_order.order_no for op in pending_by_id.values()
            if any(d.operation_id == op.id and d.depends_on_operation_id in blocked_op_ids for d in dependencies)
        }
        raise ValueError(
            f"工段 {'、'.join(sorted(blocked_names))} 已禁用，但存在依赖其工序的待排任务。"
            f"涉及订单：{'、'.join(sorted(affected_orders))}。请先启用相关工段或调整排产范围。"
        )

    completed_end: dict[int, datetime] = {}
    work_center_ready: dict[int, datetime] = {}
    external_ready: dict[tuple[int, int], datetime] = {}
    sequence_counter: dict[str, int] = {}
    planned_items: list[dict] = []

    while pending_by_id:
        ready = [
            operation
            for operation in pending_by_id.values()
            if dep_map.get(operation.id, set()).issubset(completed_end.keys())
        ]
        if not ready:
            raise ValueError("Operation dependency graph contains a cycle or missing dependency.")

        operation = sorted(ready, key=_operation_priority_key)[0]
        dependency_end = max(
            [completed_end[dep_id] for dep_id in dep_map.get(operation.id, set())],
            default=now,
        )
        start_floor = max(now, dependency_end)
        center = operation.work_center
        if is_monitor_hidden_operation(operation):
            completed_end[operation.id] = start_floor
            operation.status = "scheduled"
            pending_by_id.pop(operation.id)
            continue

        duration_h = effective_operation_duration_hours(operation)

        if center.is_external:
            duration_h = duration_h or _external_default_duration_hours(center)
            start_time_val, end_time, _, resource_key = _schedule_external_on_slot(
                center,
                start_floor,
                duration_h,
                external_ready,
                now,
            )
            machine = None
        else:
            start_time_val = next_work_time(
                max(start_floor, work_center_ready.get(center.id, now))
            )
            end_time = add_work_hours(start_time_val, duration_h)
            work_center_ready[center.id] = end_time
            machine = None
            resource_key = f"work-center:{center.id}"

        sequence_counter[resource_key] = sequence_counter.get(resource_key, 0) + 1
        planned_items.append(
            {
                "operation": operation,
                "operation_id": operation.id,
                "work_order_id": operation.work_order_id,
                "part_id": operation.part_id,
                "work_center_id": operation.work_center_id,
                "machine_id": machine.id if machine else None,
                "start_time": start_time_val,
                "end_time": end_time,
                "sequence_on_resource": sequence_counter[resource_key],
                "is_external": center.is_external,
                "external_status": "pending" if center.is_external else "not_external",
                "external_expected_return_at": end_time if center.is_external else None,
            }
        )
        completed_end[operation.id] = end_time
        pending_by_id.pop(operation.id)

    try:
        run_at = datetime.utcnow()
        schedule = await db.scalar(
            select(ProductionSchedule).where(ProductionSchedule.schedule_no == "PS-CURRENT")
        )
        if schedule is None:
            schedule = await db.scalar(
                select(ProductionSchedule)
                .where(ProductionSchedule.status == "active")
                .order_by(ProductionSchedule.created_at.desc(), ProductionSchedule.id.desc())
                .limit(1)
            )
        if schedule is None:
            schedule = ProductionSchedule(
                schedule_no="PS-CURRENT",
                name="当前统一排产方案",
                status="active",
            )
            db.add(schedule)
            await db.flush()
        else:
            schedule.status = "active"

        schedule.schedule_no = "PS-CURRENT"
        schedule.name = "当前统一排产方案"
        schedule.start_time = now
        schedule.base_schedule_id = None
        schedule.run_params_json = json.dumps(
            {
                "start_time": now.isoformat(),
                "work_order_ids": work_order_ids,
                "mode": "current_schedule_overwrite",
                "run_at": run_at.isoformat(),
            },
            ensure_ascii=False,
        )
        schedule.updated_at = run_at

        existing_item_ids = (
            await db.execute(
                select(ProductionScheduleItem.id).where(ProductionScheduleItem.schedule_id == schedule.id)
            )
        ).scalars().all()
        if existing_item_ids:
            await db.execute(
                sqla_delete(ProductionScheduleItemPersonnelAllocation)
                .where(ProductionScheduleItemPersonnelAllocation.schedule_item_id.in_(existing_item_ids))
            )
        await db.execute(
            sqla_delete(ProductionScheduleItem).where(ProductionScheduleItem.schedule_id == schedule.id)
        )
        await db.execute(
            sqla_delete(ProductionScheduleOrderLock).where(ProductionScheduleOrderLock.schedule_id == schedule.id)
        )
        await db.execute(
            sqla_delete(BusinessRiskIssueState).where(BusinessRiskIssueState.schedule_id == schedule.id)
        )

        for planned in planned_items:
            operation = planned.pop("operation")
            new_item = ProductionScheduleItem(schedule_id=schedule.id, **planned)
            db.add(new_item)
            await db.flush()
            operation.status = "scheduled"

        for op in all_operations:
            op.work_order.status = "scheduled"

        await db.commit()
    except Exception:
        await db.rollback()
        raise

    return await get_production_schedule_result(db, schedule.id)


def _serialize_schedule_item(item: ProductionScheduleItem) -> ProductionScheduleItemRead:
    operation = item.operation
    work_order = operation.work_order
    part = operation.part
    center = operation.work_center
    allocations = sorted(item.personnel_allocations, key=lambda row: (row.person.name, row.person_id))
    return ProductionScheduleItemRead(
        id=item.id,
        schedule_id=item.schedule_id,
        operation_id=item.operation_id,
        work_order_id=item.work_order_id,
        part_id=item.part_id,
        work_center_id=item.work_center_id,
        machine_id=item.machine_id,
        start_time=item.start_time,
        end_time=item.end_time,
        sequence_on_resource=item.sequence_on_resource,
        is_external=item.is_external,
        external_status=item.external_status,
        external_sent_at=item.external_sent_at,
        external_returned_at=item.external_returned_at,
        external_expected_return_at=item.external_expected_return_at,
        external_note=item.external_note,
        locked=item.locked,
        scheduled_duration_hours=scheduled_work_hours(item.start_time, item.end_time),
        order_no=work_order.order_no,
        customer=work_order.customer,
        due_date=work_order.due_date,
        part_no=part.no,
        drawing_no=part.drawing_no,
        part_name=part.name,
        operation_name=operation.name,
        requirement_note=operation.requirement_note,
        work_center_name=center.name,
        machine_name=item.machine.name if item.machine else None,
        machine_code=item.machine.code if item.machine else None,
        allocations=[_serialize_allocation(allocation) for allocation in allocations],
    )


def _build_result(schedule: ProductionSchedule, items: list[ProductionScheduleItem]) -> ProductionSchedulingResult:
    serialized = [
        _serialize_schedule_item(item)
        for item in items
        if not is_monitor_hidden_schedule_item(item)
    ]
    load_map: dict[str, dict] = {}
    order_end_map: dict[int, dict] = {}

    for item in serialized:
        if item.is_external or not item.allocations:
            load_entries = [(f"{item.work_center_id}:external", None, "外协", item.scheduled_duration_hours)]
        else:
            load_entries = [
                (
                    f"{item.work_center_id}:person:{allocation.person_id}",
                    allocation.person_id,
                    allocation.person_name,
                    round(allocation.planned_minutes / 60, 3),
                )
                for allocation in item.allocations
            ]
        for key, person_id, person_name, hours in load_entries:
            load = load_map.setdefault(
                key,
                {
                    "work_center_id": item.work_center_id,
                    "work_center_name": item.work_center_name,
                    "machine_id": item.machine_id,
                    "machine_name": item.machine_name,
                    "person_id": person_id,
                    "person_name": person_name,
                    "resource_name": person_name,
                    "task_count": 0,
                    "hours": 0.0,
                    "is_external": item.is_external,
                },
            )
            load["task_count"] += 1
            load["hours"] += hours

        current = order_end_map.get(item.work_order_id)
        if current is None or item.end_time > current["end_time"]:
            order_end_map[item.work_order_id] = {
                "work_order_id": item.work_order_id,
                "order_no": item.order_no,
                "customer": item.customer,
                "due_date": item.due_date,
                "end_time": item.end_time,
            }

    late_orders = [
        {
            **value,
            "delay_hours": round((value["end_time"] - value["due_date"]).total_seconds() / 3600, 1),
        }
        for value in order_end_map.values()
        if value["end_time"] > value["due_date"]
    ]

    return ProductionSchedulingResult(
        schedule=schedule,
        items=serialized,
        resource_load=sorted(load_map.values(), key=lambda row: row["hours"], reverse=True),
        late_orders=sorted(late_orders, key=lambda row: row["delay_hours"], reverse=True),
    )


async def get_latest_production_schedule_result(db: AsyncSession) -> ProductionSchedulingResult | None:
    result = await db.execute(select(ProductionSchedule).order_by(ProductionSchedule.created_at.desc()))
    schedule = result.scalars().first()
    if not schedule:
        return None
    return await get_production_schedule_result(db, schedule.id)


async def get_production_schedule_result(db: AsyncSession, schedule_id: int) -> ProductionSchedulingResult:
    schedule = await db.get(ProductionSchedule, schedule_id)
    if not schedule:
        raise ValueError("Schedule not found.")
    item_result = await db.execute(
        select(ProductionScheduleItem)
        .where(ProductionScheduleItem.schedule_id == schedule_id)
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.machine),
            selectinload(ProductionScheduleItem.personnel_allocations).selectinload(
                ProductionScheduleItemPersonnelAllocation.person
            ),
        )
        .order_by(
            ProductionScheduleItem.work_center_id,
            ProductionScheduleItem.machine_id,
            ProductionScheduleItem.sequence_on_resource,
        )
    )
    return _build_result(schedule, list(item_result.scalars().all()))


def _external_task_row(item: ProductionScheduleItem) -> ExternalTaskRow:
    operation = item.operation
    part = operation.part
    work_order = operation.work_order
    center = operation.work_center
    return ExternalTaskRow(
        schedule_item_id=item.id,
        schedule_id=item.schedule_id,
        operation_id=item.operation_id,
        work_order_id=item.work_order_id,
        work_center_id=item.work_center_id,
        order_no=work_order.order_no,
        customer=work_order.customer,
        drawing_no=part.drawing_no,
        part_no=part.no,
        part_name=part.name,
        operation_name=operation.name,
        requirement_note=operation.requirement_note,
        work_center_name=center.name,
        vendor_name=center.external_vendor_name,
        external_capacity_slots=_external_slot_count(center),
        planned_send_at=item.start_time,
        expected_return_at=item.external_expected_return_at or item.end_time,
        planned_duration_hours=scheduled_work_hours(item.start_time, item.end_time),
        external_status=item.external_status or "pending",
        external_sent_at=item.external_sent_at,
        external_returned_at=item.external_returned_at,
        external_note=item.external_note,
    )


async def get_external_tasks(
    db: AsyncSession,
    schedule_id: int | None = None,
    work_center_id: int | None = None,
    external_status: str | None = None,
    order_no: str | None = None,
) -> ExternalTaskListResponse:
    if schedule_id is None:
        result = await db.execute(select(ProductionSchedule).order_by(ProductionSchedule.created_at.desc()))
        schedule = result.scalars().first()
    else:
        schedule = await db.get(ProductionSchedule, schedule_id)
    if schedule is None:
        raise ValueError("暂无排产方案，请先执行生产排产。")

    query = (
        select(ProductionScheduleItem)
        .where(
            ProductionScheduleItem.schedule_id == schedule.id,
            ProductionScheduleItem.is_external == True,
        )
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.schedule),
        )
        .order_by(ProductionScheduleItem.start_time, ProductionScheduleItem.id)
    )
    if work_center_id:
        query = query.where(ProductionScheduleItem.work_center_id == work_center_id)
    if external_status:
        query = query.where(ProductionScheduleItem.external_status == external_status)
    if order_no:
        query = (
            query.join(ProductionOperation, ProductionScheduleItem.operation_id == ProductionOperation.id)
            .join(WorkOrder, ProductionOperation.work_order_id == WorkOrder.id)
            .where(WorkOrder.order_no.like(f"%{order_no}%"))
        )

    result = await db.execute(query)
    return ExternalTaskListResponse(
        schedule=schedule,
        tasks=[_external_task_row(item) for item in result.scalars().all()],
    )


async def update_external_task(
    db: AsyncSession,
    schedule_item_id: int,
    payload: ExternalTaskUpdate,
) -> ExternalTaskUpdateResponse:
    item = await db.get(
        ProductionScheduleItem,
        schedule_item_id,
        options=[
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.schedule),
        ],
    )
    if item is None or not item.is_external:
        raise ValueError("外协任务不存在。")

    update_data = payload.model_dump(exclude_unset=True)
    valid_statuses = {"pending", "sent", "returned", "exception"}
    status = update_data.get("external_status")
    if status is not None and status not in valid_statuses:
        raise ValueError("外协状态只能是 pending、sent、returned、exception。")

    now = datetime.utcnow()
    if "external_status" in update_data:
        item.external_status = status or "pending"
        if item.external_status == "sent" and item.external_sent_at is None:
            item.external_sent_at = now
        if item.external_status == "returned" and item.external_returned_at is None:
            item.external_returned_at = now

    for field in ("external_sent_at", "external_returned_at", "external_expected_return_at", "external_note"):
        if field in update_data:
            setattr(item, field, update_data[field])

    if "external_expected_return_at" in update_data and "external_status" not in update_data:
        item.external_status = "sent"
        if item.external_sent_at is None:
            item.external_sent_at = now

    if item.external_status == "returned" and item.external_returned_at:
        item.external_expected_return_at = item.external_returned_at
        item.end_time = item.external_returned_at
    elif item.external_expected_return_at:
        item.end_time = item.external_expected_return_at

    recalculated_count = await _recalculate_schedule_from_item(db, schedule_item_id)
    await db.commit()
    refreshed = await db.get(
        ProductionScheduleItem,
        schedule_item_id,
        options=[
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.schedule),
        ],
    )
    return ExternalTaskUpdateResponse(
        schedule=refreshed.schedule,
        task=_external_task_row(refreshed),
        recalculated_item_count=recalculated_count,
    )


async def get_production_gantt_data(db: AsyncSession, schedule_id: int | None = None) -> list[dict]:
    if schedule_id:
        result = await get_production_schedule_result(db, schedule_id)
    else:
        result = await get_latest_production_schedule_result(db)
    if not result:
        return []

    lanes: dict[str, dict] = {}
    for item in result.items:
        if item.work_center_name in MONITOR_HIDDEN_WORK_CENTERS or item.operation_name in MONITOR_HIDDEN_WORK_CENTERS:
            continue
        if item.is_external or not item.allocations:
            lane_entries = [(f"external:{item.work_center_id}", None, None, item.scheduled_duration_hours)]
        else:
            lane_entries = [
                (
                    f"person:{allocation.person_id}",
                    allocation.person_id,
                    allocation.person_name,
                    round(allocation.planned_minutes / 60, 3),
                )
                for allocation in item.allocations
            ]
        for key, person_id, person_name, duration_hours in lane_entries:
            lane = lanes.setdefault(
                key,
                {
                    "resource_key": key,
                    "work_center_id": item.work_center_id,
                    "work_center_name": item.work_center_name,
                    "machine_id": item.machine_id,
                    "machine_name": item.machine_name,
                    "machine_code": item.machine_code,
                    "person_id": person_id,
                    "person_name": person_name,
                    "is_external": item.is_external,
                    "tasks": [],
                },
            )
            lane["tasks"].append({
                "schedule_item_id": item.id,
                "operation_id": item.operation_id,
                "work_order_id": item.work_order_id,
                "order_no": item.order_no,
                "part_no": item.part_no,
                "drawing_no": item.drawing_no,
                "part_name": item.part_name,
                "task_name": item.operation_name,
                "work_center_name": item.work_center_name,
                "start_time": item.start_time,
                "end_time": item.end_time,
                "scheduled_duration_hours": duration_hours,
                "sequence_on_machine": item.sequence_on_resource,
                "is_external": item.is_external,
                "person_id": person_id,
                "person_name": person_name,
            })
    return list(lanes.values())


def is_workday(day: date) -> bool:
    return day.weekday() != 6


def work_hours_on_day(start: datetime, end: datetime, day: date) -> float:
    if not is_workday(day) or end <= start:
        return 0.0

    total_minutes = 0
    for segment_start_time, segment_end_time in ((WORK_START, LUNCH_START), (LUNCH_END, WORK_END)):
        segment_start = datetime.combine(day, segment_start_time)
        segment_end = datetime.combine(day, segment_end_time)
        overlap_start = max(start, segment_start)
        overlap_end = min(end, segment_end)
        if overlap_end > overlap_start:
            total_minutes += int((overlap_end - overlap_start).total_seconds() // 60)
    return round(total_minutes / 60, 2)


def build_date_columns(start: date, days: int) -> list[ScheduleBoardDateColumn]:
    return [
        ScheduleBoardDateColumn(
            date=(start + timedelta(days=index)).isoformat(),
            weekday=WEEKDAYS[(start + timedelta(days=index)).weekday()],
            is_workday=is_workday(start + timedelta(days=index)),
        )
        for index in range(max(days, 1))
    ]


async def get_default_person_by_work_center(db: AsyncSession) -> dict[int, Personnel]:
    result = await db.execute(
        select(WorkCenterPersonnel)
        .options(selectinload(WorkCenterPersonnel.person))
        .order_by(WorkCenterPersonnel.work_center_id, WorkCenterPersonnel.sort_order, WorkCenterPersonnel.id)
    )
    people: dict[int, Personnel] = {}
    for link in result.scalars().all():
        people.setdefault(link.work_center_id, link.person)
    return people


async def get_schedule_board(
    db: AsyncSession,
    schedule_id: int,
    work_center: str | None = None,
    start_date: date | None = None,
    days: int = 14,
    order_id: int | None = None,
    view_mode: str = "by_work_center",
) -> ScheduleBoardResponse:
    if view_mode not in {"by_work_center", "by_machine", "by_person"}:
        raise ValueError("view_mode must be one of by_work_center, by_machine, by_person.")

    schedule = await db.get(ProductionSchedule, schedule_id)
    if schedule is None:
        raise ValueError("Schedule not found.")

    result = await db.execute(
        select(ProductionScheduleItem)
        .where(ProductionScheduleItem.schedule_id == schedule_id)
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.machine),
            selectinload(ProductionScheduleItem.personnel_allocations).selectinload(
                ProductionScheduleItemPersonnelAllocation.person
            ),
        )
        .order_by(
            ProductionScheduleItem.work_center_id,
            ProductionScheduleItem.machine_id,
            ProductionScheduleItem.start_time,
            ProductionScheduleItem.sequence_on_resource,
        )
    )
    items = [
        item
        for item in result.scalars().all()
        if not is_monitor_hidden_schedule_item(item)
    ]

    if work_center:
        if work_center.isdigit():
            work_center_id = int(work_center)
            items = [item for item in items if item.work_center_id == work_center_id]
        else:
            items = [
                item
                for item in items
                if item.operation.work_center.name == work_center
            ]

    if order_id is not None:
        items = [item for item in items if item.work_order_id == order_id]

    if start_date is None:
        start_date = min((item.start_time.date() for item in items), default=datetime.utcnow().date())

    date_columns = build_date_columns(start_date, days)
    rows: list[ScheduleBoardRow] = []

    for item in items:
        operation = item.operation
        work_order = operation.work_order
        part = operation.part
        center = operation.work_center
        machine_name = item.machine.name if item.machine else None
        allocations = sorted(item.personnel_allocations, key=lambda row: (row.person.name, row.person_id))
        allocation_label = (
            " / ".join(f"{allocation.person.name} {allocation.ratio_percent:g}%" for allocation in allocations)
            if allocations
            else "未派工"
        )

        if view_mode == "by_machine":
            if item.machine_id:
                group_key = f"machine:{item.machine_id}"
                group_label = machine_name or "未分配设备"
            else:
                group_key = f"external:{center.id}"
                group_label = f"{center.name} / 外协"
        elif view_mode == "by_person":
            group_key = ""
            group_label = ""
        else:
            group_key = f"work_center:{center.id}"
            group_label = center.name

        base_daily_hours = [
            work_hours_on_day(item.start_time, item.end_time, date.fromisoformat(column.date))
            for column in date_columns
        ]
        if not any(hours > 0 for hours in base_daily_hours):
            continue

        row_allocations = allocations if view_mode == "by_person" and allocations else [None]
        for allocation in row_allocations:
            ratio = allocation.ratio_percent / 100 if allocation else 1
            if view_mode == "by_person":
                group_key = f"person:{allocation.person_id}" if allocation else "person:unassigned"
                group_label = allocation.person.name if allocation else "未派工"
                person_name = group_label
            else:
                person_name = allocation_label

            daily_cells = [
                ScheduleBoardDailyCell(date=column.date, hours=round(hours * ratio, 2))
                for column, hours in zip(date_columns, base_daily_hours)
            ]

            rows.append(
                ScheduleBoardRow(
                    group_key=group_key,
                    group_label=group_label,
                    schedule_item_id=item.id,
                    operation_id=item.operation_id,
                    work_order_id=item.work_order_id,
                    work_center_id=item.work_center_id,
                    order_no=work_order.order_no,
                    operation_name=operation.name,
                    drawing_no=part.drawing_no,
                    part_no=part.no,
                    part_name=part.name,
                    customer_name=work_order.customer,
                    requirement_note=operation.requirement_note,
                    quantity=part.quantity,
                    duration_hours=round(scheduled_work_hours(item.start_time, item.end_time) * ratio, 2),
                    due_date=work_order.due_date,
                    planned_start=item.start_time,
                    planned_end=item.end_time,
                    machine_name=machine_name,
                    person_name=person_name,
                    is_external=item.is_external,
                    is_late=item.end_time > work_order.due_date,
                    daily_cells=daily_cells,
                )
            )

    rows.sort(key=lambda row: (row.group_label, row.planned_start, row.order_no, row.part_no))
    return ScheduleBoardResponse(
        schedule=schedule,
        view_mode=view_mode,
        date_columns=date_columns,
        rows=rows,
    )


async def lock_order_in_schedule(
    db: AsyncSession,
    schedule_id: int,
    work_order_id: int,
    locked_by: str | None = None,
    note: str | None = None,
) -> dict:
    schedule = await db.get(ProductionSchedule, schedule_id)
    if not schedule:
        raise ValueError("排产方案不存在。")

    work_order = await db.get(WorkOrder, work_order_id)
    if not work_order:
        raise ValueError("工单不存在。")

    # Check if order exists in this schedule
    items_result = await db.execute(
        select(ProductionScheduleItem).where(
            ProductionScheduleItem.schedule_id == schedule_id,
            ProductionScheduleItem.work_order_id == work_order_id,
        )
    )
    items = list(items_result.scalars().all())
    if not items:
        raise ValueError(f"订单 {work_order.order_no} 不在当前排产方案中。")

    # Upsert order lock
    lock_result = await db.execute(
        select(ProductionScheduleOrderLock).where(
            ProductionScheduleOrderLock.schedule_id == schedule_id,
            ProductionScheduleOrderLock.work_order_id == work_order_id,
        )
    )
    lock = lock_result.scalars().first()
    now = datetime.utcnow()
    if lock:
        lock.locked = True
        lock.locked_at = now
        lock.locked_by = locked_by
        lock.note = note
    else:
        lock = ProductionScheduleOrderLock(
            schedule_id=schedule_id,
            work_order_id=work_order_id,
            locked=True,
            locked_at=now,
            locked_by=locked_by,
            note=note,
        )
        db.add(lock)

    # Mark all schedule items for this order as locked
    for item in items:
        item.locked = True
        item.locked_at = now
        item.locked_by = locked_by

    await db.commit()
    await db.refresh(lock)
    return {"schedule_id": schedule_id, "work_order_id": work_order_id, "locked": True, "locked_at": now}


async def unlock_order_in_schedule(
    db: AsyncSession,
    schedule_id: int,
    work_order_id: int,
) -> dict:
    schedule = await db.get(ProductionSchedule, schedule_id)
    if not schedule:
        raise ValueError("排产方案不存在。")

    lock_result = await db.execute(
        select(ProductionScheduleOrderLock).where(
            ProductionScheduleOrderLock.schedule_id == schedule_id,
            ProductionScheduleOrderLock.work_order_id == work_order_id,
        )
    )
    lock = lock_result.scalars().first()
    if lock:
        lock.locked = False

    # Unmark schedule items
    items_result = await db.execute(
        select(ProductionScheduleItem).where(
            ProductionScheduleItem.schedule_id == schedule_id,
            ProductionScheduleItem.work_order_id == work_order_id,
        )
    )
    for item in items_result.scalars().all():
        item.locked = False
        item.locked_at = None
        item.locked_by = None
        item.lock_reason = None

    await db.commit()
    return {"schedule_id": schedule_id, "work_order_id": work_order_id, "locked": False}


async def get_locked_order_ids(db: AsyncSession, schedule_id: int) -> set[int]:
    result = await db.execute(
        select(ProductionScheduleOrderLock).where(
            ProductionScheduleOrderLock.schedule_id == schedule_id,
            ProductionScheduleOrderLock.locked == True,
        )
    )
    return {lock.work_order_id for lock in result.scalars().all()}


async def export_schedule_to_excel(db: AsyncSession, schedule_id: int) -> bytes:
    """Generate an Excel file with 4 sheets for the given schedule."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    schedule = await db.get(ProductionSchedule, schedule_id)
    if not schedule:
        raise ValueError("排产方案不存在。")

    # Get schedule items with full relations
    item_result = await db.execute(
        select(ProductionScheduleItem)
        .where(ProductionScheduleItem.schedule_id == schedule_id)
        .options(
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_order),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.part),
            selectinload(ProductionScheduleItem.operation).selectinload(ProductionOperation.work_center),
            selectinload(ProductionScheduleItem.machine),
        )
        .order_by(ProductionScheduleItem.start_time)
    )
    items = list(item_result.scalars().all())
    if not items:
        raise ValueError("该排产方案没有明细数据。")

    # Build data structures
    items_by_order: dict[int, list[ProductionScheduleItem]] = defaultdict(list)
    load_map: dict[tuple[int, int | None], dict] = {}
    for item in items:
        items_by_order[item.work_order_id].append(item)
        key = (item.work_center_id, item.machine_id)
        center = item.operation.work_center
        load = load_map.setdefault(key, {
            "work_center_name": center.name,
            "machine_name": item.machine.name if item.machine else "外协",
            "busy_minutes": 0,
            "is_external": item.is_external,
            "capacity_per_day": (
                center.default_capacity_per_day
                if item.is_external or item.machine is None
                else item.machine.capacity_per_day
            ),
        })
        load["busy_minutes"] += scheduled_work_minutes(item.start_time, item.end_time)

    locked_order_ids = await get_locked_order_ids(db, schedule_id)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D5D8C", end_color="2D5D8C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    wb = Workbook()

    # Sheet 1: 订单完工表
    ws1 = wb.active
    ws1.title = "订单完工表"
    ws1_headers = [
        "排产方案号", "订单号", "客户", "产品名称", "数量", "优先级",
        "交期", "预计开始", "预计完成", "延期天数", "状态", "锁定", "主要瓶颈"
    ]
    for col, h in enumerate(ws1_headers, 1):
        ws1.cell(1, col, h)
    style_header(ws1, len(ws1_headers))

    row_idx = 2
    for order_id, order_items in items_by_order.items():
        wo = order_items[0].operation.work_order
        planned_start = min(it.start_time for it in order_items)
        planned_end = max(it.end_time for it in order_items)
        delay_days = max((planned_end.date() - wo.due_date.date()).days, 0) if planned_end > wo.due_date else 0
        status = "延期" if planned_end > wo.due_date else "正常"
        is_locked = order_id in locked_order_ids

        # Find bottleneck
        latest_item = max(order_items, key=lambda it: it.end_time)
        bottleneck = latest_item.operation.work_center.name if planned_end > wo.due_date else ""

        ws1.cell(row_idx, 1, schedule.schedule_no)
        ws1.cell(row_idx, 2, wo.order_no)
        ws1.cell(row_idx, 3, wo.customer)
        ws1.cell(row_idx, 4, wo.product_name)
        ws1.cell(row_idx, 5, wo.quantity)
        ws1.cell(row_idx, 6, wo.priority)
        ws1.cell(row_idx, 7, wo.due_date.strftime("%Y-%m-%d"))
        ws1.cell(row_idx, 8, planned_start.strftime("%Y-%m-%d %H:%M"))
        ws1.cell(row_idx, 9, planned_end.strftime("%Y-%m-%d %H:%M"))
        ws1.cell(row_idx, 10, delay_days)
        ws1.cell(row_idx, 11, status)
        ws1.cell(row_idx, 12, "已锁定" if is_locked else "")
        ws1.cell(row_idx, 13, bottleneck)
        for col in range(1, len(ws1_headers) + 1):
            ws1.cell(row_idx, col).border = thin_border
        row_idx += 1

    for col in range(1, len(ws1_headers) + 1):
        ws1.column_dimensions[ws1.cell(1, col).column_letter].width = 16

    # Sheet 2: 设备排班表
    ws2 = wb.create_sheet("设备排班表")
    ws2_headers = [
        "日期", "工段", "设备", "订单号", "客户", "零件图号", "零件名称",
        "工序", "计划开始", "计划结束", "工时(小时)", "外协"
    ]
    for col, h in enumerate(ws2_headers, 1):
        ws2.cell(1, col, h)
    style_header(ws2, len(ws2_headers))

    row_idx = 2
    for item in items:
        op = item.operation
        wo = op.work_order
        part = op.part
        center = op.work_center
        ws2.cell(row_idx, 1, item.start_time.strftime("%Y-%m-%d"))
        ws2.cell(row_idx, 2, center.name)
        ws2.cell(row_idx, 3, item.machine.name if item.machine else "外协")
        ws2.cell(row_idx, 4, wo.order_no)
        ws2.cell(row_idx, 5, wo.customer)
        ws2.cell(row_idx, 6, part.drawing_no)
        ws2.cell(row_idx, 7, part.name)
        ws2.cell(row_idx, 8, op.name)
        ws2.cell(row_idx, 9, item.start_time.strftime("%Y-%m-%d %H:%M"))
        ws2.cell(row_idx, 10, item.end_time.strftime("%Y-%m-%d %H:%M"))
        ws2.cell(row_idx, 11, scheduled_work_hours(item.start_time, item.end_time))
        ws2.cell(row_idx, 12, "是" if item.is_external else "否")
        for col in range(1, len(ws2_headers) + 1):
            ws2.cell(row_idx, col).border = thin_border
        row_idx += 1

    for col in range(1, len(ws2_headers) + 1):
        ws2.column_dimensions[ws2.cell(1, col).column_letter].width = 16

    # Sheet 3: 资源负荷表
    ws3 = wb.create_sheet("资源负荷表")
    ws3_headers = ["工段", "设备", "占用分钟", "可用分钟", "负荷率", "状态"]
    for col, h in enumerate(ws3_headers, 1):
        ws3.cell(1, col, h)
    style_header(ws3, len(ws3_headers))

    # Calculate available minutes
    first_day = min(it.start_time.date() for it in items)
    last_day = max(it.end_time.date() for it in items)
    workdays = _workday_count_for_export(first_day, last_day)

    row_idx = 2
    for key, load in sorted(load_map.items(), key=lambda x: x[1]["busy_minutes"], reverse=True):
        cap_per_day = load["capacity_per_day"]
        available = max(workdays * cap_per_day, 1)
        utilization = round(load["busy_minutes"] / available, 3)
        if utilization >= 0.9:
            status = "瓶颈"
        elif utilization >= 0.6:
            status = "正常"
        else:
            status = "空闲较多"

        ws3.cell(row_idx, 1, load["work_center_name"])
        ws3.cell(row_idx, 2, load["machine_name"])
        ws3.cell(row_idx, 3, load["busy_minutes"])
        ws3.cell(row_idx, 4, available)
        ws3.cell(row_idx, 5, f"{utilization * 100:.1f}%")
        ws3.cell(row_idx, 6, status)
        for col in range(1, len(ws3_headers) + 1):
            ws3.cell(row_idx, col).border = thin_border
        row_idx += 1

    for col in range(1, len(ws3_headers) + 1):
        ws3.column_dimensions[ws3.cell(1, col).column_letter].width = 18

    # Sheet 4: 逾期风险表
    ws4 = wb.create_sheet("逾期风险表")
    ws4_headers = ["订单号", "客户", "交期", "预计完成", "延期天数", "瓶颈资源", "原因", "建议动作"]
    for col, h in enumerate(ws4_headers, 1):
        ws4.cell(1, col, h)
    style_header(ws4, len(ws4_headers))

    row_idx = 2
    for order_items in items_by_order.values():
        wo = order_items[0].operation.work_order
        planned_end = max(it.end_time for it in order_items)
        if planned_end <= wo.due_date:
            continue
        delay_days = max((planned_end.date() - wo.due_date.date()).days, 0)
        latest_item = max(order_items, key=lambda it: it.end_time)
        center_name = latest_item.operation.work_center.name
        load_entry = load_map.get((latest_item.work_center_id, latest_item.machine_id))
        cap_per_day = load_entry["capacity_per_day"] if load_entry else 480
        is_bottleneck = load_entry and (load_entry["busy_minutes"] / max(workdays * cap_per_day, 1)) >= 0.9

        if is_bottleneck:
            reason = f"{center_name}资源负荷较高，导致关键工序等待或排队。"
        else:
            reason = f"{center_name}为最后完成工序，订单整体完工晚于交期。"

        ws4.cell(row_idx, 1, wo.order_no)
        ws4.cell(row_idx, 2, wo.customer)
        ws4.cell(row_idx, 3, wo.due_date.strftime("%Y-%m-%d"))
        ws4.cell(row_idx, 4, planned_end.strftime("%Y-%m-%d %H:%M"))
        ws4.cell(row_idx, 5, delay_days)
        ws4.cell(row_idx, 6, center_name)
        ws4.cell(row_idx, 7, reason)
        ws4.cell(row_idx, 8, "建议调整订单优先级、增加该工段班次、临时外协或检查是否存在设备空闲未利用。")
        for col in range(1, len(ws4_headers) + 1):
            ws4.cell(row_idx, col).border = thin_border
        row_idx += 1

    if row_idx == 2:
        ws4.cell(2, 1, "暂无延期风险")
        for col in range(1, len(ws4_headers) + 1):
            ws4.cell(2, col).border = thin_border

    for col in range(1, len(ws4_headers) + 1):
        ws4.column_dimensions[ws4.cell(1, col).column_letter].width = 20

    # Save export record
    filename = f"排产结果_{schedule.schedule_no}.xlsx"
    export_record = ExportBatch(
        export_type="schedule_result",
        schedule_id=schedule_id,
        filename=filename,
    )
    db.add(export_record)
    await db.commit()

    # Return bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _workday_count_for_export(start_day: date, end_day: date) -> int:
    current = start_day
    count = 0
    while current <= end_day:
        if is_workday(current):
            count += 1
        current += timedelta(days=1)
    return max(count, 1)
