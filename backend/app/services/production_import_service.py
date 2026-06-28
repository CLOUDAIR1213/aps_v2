from __future__ import annotations

from collections import Counter
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.schemas.production import (
    ImportIssue,
    ImportOperationPreview,
    ImportPartPreview,
    ImportPreviewPayload,
)


PRIMARY_SHEET = "焊接件明细"
BASE_COLUMNS = 11
BASE_COLUMN_HEADERS = {
    "NO",
    "图号",
    "名称",
    "材料厚",
    "材料长",
    "材料宽",
    "产品数量(件)",
    "材料",
    "单料重",
    "材料总重",
    "材料费",
}
NOTE_HEADERS = {"备注", "备注1", "工艺备注", "加工要求", "工艺要求", "关键尺寸/备注"}
IGNORED_OPERATION_HEADERS = {"", *NOTE_HEADERS, "flag", "计划日期", "完工日期"}
EXTERNAL_WORK_CENTERS = {"钣金外", "加工外", "表面处理"}


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value: Any, default: int = 1) -> int:
    if value in (None, ""):
        return default
    try:
        return max(int(float(value)), 0)
    except (TypeError, ValueError):
        return default


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _is_assembly_no(no: str) -> bool:
    return bool(no) and "." not in no


def get_parent_no(no: str) -> str | None:
    if "." not in no:
        return None
    return no.rsplit(".", 1)[0]


def _hierarchy_summary(parts: list[ImportPartPreview]) -> dict:
    part_nos = [part.no for part in parts]
    no_counts = Counter(part_nos)
    unique_nos = set(part_nos)
    child_edges = [
        (part.no, part.parent_no)
        for part in parts
        if part.parent_no and part.parent_no in unique_nos
    ]
    parent_nos = {parent_no for _, parent_no in child_edges}
    leaf_count = len([part for part in parts if part.no not in parent_nos])

    return {
        "max_depth": max((part.no.count(".") + 1 for part in parts), default=0),
        "root_count": len([part for part in parts if part.parent_no is None]),
        "leaf_count": leaf_count,
        "parent_child_edge_count": len(child_edges),
        "missing_parent_count": len(
            [part for part in parts if part.parent_no and part.parent_no not in unique_nos]
        ),
        "duplicate_no_count": len([no for no, count in no_counts.items() if count > 1]),
    }


async def parse_work_order_workbook(
    content: bytes,
    filename: str,
    mapping_rules: dict[str, dict] | None = None,
) -> ImportPreviewPayload:
    mapping_rules = mapping_rules or {}
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=True, keep_vba=True)
    issues: list[ImportIssue] = []

    if PRIMARY_SHEET not in workbook.sheetnames:
        raise ValueError(f"Workbook must contain sheet '{PRIMARY_SHEET}'.")

    sheet = workbook[PRIMARY_SHEET]
    headers = [_to_text(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    operation_columns: list[tuple[int, str, int]] = []
    note_columns: list[int] = []

    for col_idx, header in enumerate(headers, start=1):
        if header in NOTE_HEADERS:
            note_columns.append(col_idx)
        if col_idx <= BASE_COLUMNS or header in IGNORED_OPERATION_HEADERS:
            continue
        if header in BASE_COLUMN_HEADERS:
            continue
        operation_columns.append((col_idx, header, len(operation_columns) + 1))
        if header not in mapping_rules:
            issues.append(
                ImportIssue(
                    severity="error",
                    column=col_idx,
                    field="work_center",
                    message=f"工序列 '{header}' 尚未配置映射规则，请先在工序映射中配置。",
                )
            )

    parts: list[ImportPartPreview] = []
    operations: list[ImportOperationPreview] = []
    part_operation_counts: dict[int, int] = {}
    part_hours: dict[int, float] = {}
    external_blank_skipped_count = 0
    external_default_duration_count = 0

    for row_idx in range(2, sheet.max_row + 1):
        no = _to_text(sheet.cell(row_idx, 1).value)
        drawing_no = _to_text(sheet.cell(row_idx, 2).value)
        name = _to_text(sheet.cell(row_idx, 3).value)

        if not no and not drawing_no and not name:
            continue
        if not no and drawing_no == "第一行不输入":
            continue
        if not drawing_no:
            issues.append(
                ImportIssue(
                    severity="error",
                    row=row_idx,
                    field="drawing_no",
                    message="图号为空，该行不会生成有效生产任务。",
                )
            )
            continue
        if not no:
            issues.append(
                ImportIssue(
                    severity="error",
                    row=row_idx,
                    field="no",
                    message=f"图号 {drawing_no} 的 NO 为空，无法建立部件层级。",
                )
            )
            continue

        is_assembly = _is_assembly_no(no)
        row_notes = [
            _to_text(sheet.cell(row_idx, col_idx).value)
            for col_idx in note_columns
            if _to_text(sheet.cell(row_idx, col_idx).value)
        ]
        requirement_note = "\n".join(row_notes) or None

        part = ImportPartPreview(
            no=no,
            drawing_no=drawing_no,
            name=name or drawing_no,
            parent_no=get_parent_no(no),
            material=_to_text(sheet.cell(row_idx, 8).value) or None,
            note=requirement_note,
            quantity=_to_int(sheet.cell(row_idx, 7).value, 1),
            source_row=row_idx,
            is_assembly=is_assembly,
            operation_count=0,
            total_hours=0,
            capacity_hours=0,
        )
        parts.append(part)
        part_operation_counts[row_idx] = 0
        part_hours[row_idx] = 0

        for col_idx, header, seq_no in operation_columns:
            raw_value = sheet.cell(row_idx, col_idx).value
            raw_text = _to_text(raw_value)
            is_external = mapping_rules.get(header, {}).get("is_external", header in EXTERNAL_WORK_CENTERS)

            if raw_text == "":
                if is_external:
                    external_blank_skipped_count += 1
                continue

            duration = _to_float(raw_value)
            if duration is None:
                if is_external:
                    default_h = None
                    if header in mapping_rules:
                        default_h = (
                            mapping_rules[header].get("external_lead_time_hours")
                            or mapping_rules[header].get("default_duration_hours")
                        )
                    if default_h and default_h > 0:
                        duration = float(default_h)
                        external_default_duration_count += 1
                        issues.append(
                            ImportIssue(
                                severity="info",
                                row=row_idx,
                                column=col_idx,
                                field=header,
                                message=(
                                    f"{drawing_no} 的外协工序 '{header}' 使用非数字标记 "
                                    f"'{raw_text}'，已按默认周期 {default_h}h 生成任务。"
                                ),
                            )
                        )
                    else:
                        issues.append(
                            ImportIssue(
                                severity="warning",
                                row=row_idx,
                                column=col_idx,
                                field="external_default_duration",
                                message=(
                                    f"{drawing_no} 的外协工序 '{header}' 使用非数字标记 "
                                    f"'{raw_text}'，但工段没有默认外协周期，已跳过。"
                                ),
                            )
                        )
                        continue
                else:
                    issues.append(
                        ImportIssue(
                            severity="warning",
                            row=row_idx,
                            column=col_idx,
                            field=header,
                            message=f"{drawing_no} 的工序 '{header}' 工时不是数字，已跳过。",
                        )
                    )
                    continue
            if duration is None:
                issues.append(
                    ImportIssue(
                        severity="warning",
                        row=row_idx,
                        column=col_idx,
                        field=header,
                        message=f"{drawing_no} 的工序 '{header}' 工时不是数字，已跳过。",
                    )
                )
                continue
            if duration <= 0:
                continue

            operations.append(
                ImportOperationPreview(
                    part_no=no,
                    drawing_no=drawing_no,
                    part_name=name or drawing_no,
                    work_center_name=header,
                    seq_no=seq_no,
                    duration_hours=round(duration, 3),
                    requirement_note=requirement_note,
                    source_row=row_idx,
                    source_col=col_idx,
                    is_external=is_external,
                    mapped=header in mapping_rules,
                )
            )
            part_operation_counts[row_idx] += 1
            part_hours[row_idx] += duration

    part_no_set = {part.no for part in parts}
    no_counts = Counter(part.no for part in parts)
    duplicate_nos = {no for no, count in no_counts.items() if count > 1}
    duplicate_no_drawing_pairs = {
        pair
        for pair, count in Counter((part.no, part.drawing_no) for part in parts).items()
        if count > 1
    }
    for part in parts:
        part.operation_count = part_operation_counts.get(part.source_row, 0)
        part.total_hours = round(part_hours.get(part.source_row, 0), 3)
        part.capacity_hours = round(part.total_hours * max(part.quantity, 1), 3)
        if (part.no, part.drawing_no) in duplicate_no_drawing_pairs:
            issues.append(
                ImportIssue(
                    severity="error",
                    row=part.source_row,
                    field="no",
                    message=(
                        f"NO {part.no} 与图号 {part.drawing_no} 同时重复，无法区分是否为同一零件重复行。"
                    ),
                )
            )
        elif part.no in duplicate_nos:
            issues.append(
                ImportIssue(
                    severity="warning",
                    row=part.source_row,
                    field="no",
                    message=(
                        f"NO {part.no} 重复，系统已按 Excel 行号拆分为独立零件节点；"
                        "不确定的层级依赖不会自动建立。"
                    ),
                )
            )
        if part.parent_no and part.parent_no not in part_no_set:
            issues.append(
                ImportIssue(
                    severity="warning",
                    row=part.source_row,
                    field="parent_no",
                    message=f"子件 {part.no} 未找到直接父级 {part.parent_no}，无法建立完整层级依赖。",
                )
            )
        elif part.parent_no and no_counts.get(part.parent_no, 0) > 1:
            issues.append(
                ImportIssue(
                    severity="warning",
                    row=part.source_row,
                    field="parent_no",
                    message=(
                        f"子件 {part.no} 的父级 NO {part.parent_no} 存在重复，"
                        "系统不会自动建立该层级依赖。"
                    ),
                )
            )
        if part.operation_count == 0:
            if any(candidate.parent_no == part.no for candidate in parts):
                issues.append(
                    ImportIssue(
                        severity="warning",
                        row=part.source_row,
                        field="no",
                        message=f"父级 {part.no} 没有工序，仅作为层级节点。",
                    )
                )
            else:
                issues.append(
                    ImportIssue(
                        severity="warning",
                        row=part.source_row,
                        field="no",
                        message=f"叶子件 {part.no} 没有任何工序，不会生成排产任务。",
                    )
                )

    part_quantity_by_row = {part.source_row: part.quantity for part in parts}
    summary = {
        "part_count": len(parts),
        "assembly_count": len([part for part in parts if part.is_assembly]),
        "child_part_count": len([part for part in parts if not part.is_assembly]),
        "operation_count": len(operations),
        "work_center_count": len({op.work_center_name for op in operations}),
        "total_hours": round(sum(op.duration_hours for op in operations), 3),
        "total_capacity_hours": round(
            sum(op.duration_hours * max(part_quantity_by_row.get(op.source_row, 1), 1) for op in operations),
            3,
        ),
        "external_task_count": len([op for op in operations if op.is_external]),
        "external_total_hours": round(sum(op.duration_hours for op in operations if op.is_external), 3),
        "external_total_capacity_hours": round(
            sum(
                op.duration_hours * max(part_quantity_by_row.get(op.source_row, 1), 1)
                for op in operations
                if op.is_external
            ),
            3,
        ),
        "external_blank_skipped_count": external_blank_skipped_count,
        "external_default_duration_count": external_default_duration_count,
        "note_count": len([part for part in parts if part.note]),
        "error_count": len([issue for issue in issues if issue.severity == "error"]),
        "warning_count": len([issue for issue in issues if issue.severity == "warning"]),
        "hierarchy": _hierarchy_summary(parts),
    }

    return ImportPreviewPayload(
        source_filename=filename,
        sheet_name=PRIMARY_SHEET,
        parts=parts,
        operations=operations,
        issues=issues,
        summary=summary,
    )
