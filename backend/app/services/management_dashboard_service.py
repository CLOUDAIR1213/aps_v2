from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.production import (
    BusinessRiskIssueState,
    ExportBatch,
    ProductionScheduleItem,
)
from app.schemas.production import (
    ManagementDashboardResponse,
    ManagementDashboardSummary,
    ManagementIssueLink,
    ManagementIssueRow,
    ManagementIssueStateRead,
    ManagementIssueStateUpdate,
)
from app.services.production_analysis_service import (
    _delay_days,
    _load_schedule_items,
    _resolve_schedule,
    _resource_load_rows,
)
from app.services.production_service import is_monitor_hidden_schedule_item


VALID_ISSUE_STATUSES = {"open", "processing", "resolved", "paused"}
RISK_LEVEL_ORDER = {"high": 0, "medium": 1, "low": 2}
RISK_TYPE_LABELS = {
    "order_delay": "订单延期",
    "due_soon": "临近交期",
    "resource_bottleneck": "资源瓶颈",
    "operation_blocking": "关键工序阻塞",
    "external_risk": "外协影响",
}


def _links(schedule_id: int, work_order_id: int | None = None) -> ManagementIssueLink:
    return ManagementIssueLink(
        order_detail=(
            f"/scheduling/orders/{work_order_id}?schedule_id={schedule_id}"
            if work_order_id
            else None
        ),
        schedule_board=f"/scheduling/board/{schedule_id}",
        gantt=f"/gantt?schedule_id={schedule_id}",
    )


def _issue(
    *,
    schedule_id: int,
    risk_type: str,
    risk_level: str,
    title: str,
    reason: str,
    suggestion: str,
    work_order_id: int | None = None,
    order_no: str | None = None,
    customer_name: str | None = None,
    product_name: str | None = None,
    due_date: datetime | None = None,
    planned_start_time: datetime | None = None,
    planned_end_time: datetime | None = None,
    delay_days: int = 0,
    work_center_id: int | None = None,
    work_center_name: str | None = None,
    machine_id: int | None = None,
    machine_name: str | None = None,
    operation_id: int | None = None,
    operation_name: str | None = None,
    utilization: float | None = None,
) -> ManagementIssueRow:
    key_parts = [
        risk_type.replace("_", "-"),
        str(schedule_id),
        str(work_order_id or "resource"),
        str(work_center_id or "center"),
        str(machine_id or "external"),
        str(operation_id or "order"),
    ]
    issue_key = ":".join(key_parts)
    return ManagementIssueRow(
        issue_key=issue_key,
        schedule_id=schedule_id,
        risk_type=risk_type,
        risk_level=risk_level,
        title=title,
        work_order_id=work_order_id,
        order_no=order_no,
        customer_name=customer_name,
        product_name=product_name,
        due_date=due_date,
        planned_start_time=planned_start_time,
        planned_end_time=planned_end_time,
        delay_days=delay_days,
        work_center_id=work_center_id,
        work_center_name=work_center_name,
        machine_id=machine_id,
        machine_name=machine_name,
        operation_id=operation_id,
        operation_name=operation_name,
        utilization=utilization,
        reason=reason,
        suggestion=suggestion,
        links=_links(schedule_id, work_order_id),
    )


async def _load_issue_states(
    db: AsyncSession,
    schedule_id: int,
) -> dict[str, BusinessRiskIssueState]:
    result = await db.execute(
        select(BusinessRiskIssueState).where(BusinessRiskIssueState.schedule_id == schedule_id)
    )
    return {state.issue_key: state for state in result.scalars().all()}


def _merge_issue_states(
    issues: list[ManagementIssueRow],
    states: dict[str, BusinessRiskIssueState],
) -> list[ManagementIssueRow]:
    for issue in issues:
        state = states.get(issue.issue_key)
        if state:
            issue.status = state.status
            issue.note = state.note
            issue.updated_at = state.updated_at
    return issues


def _items_in_horizon(
    items: list[ProductionScheduleItem],
    horizon_end: date,
) -> list[ProductionScheduleItem]:
    return [
        item
        for item in items
        if item.operation.work_order.due_date.date() <= horizon_end
        or item.start_time.date() <= horizon_end
        or item.end_time.date() <= horizon_end
    ]


def _build_order_issues(
    schedule_id: int,
    items: list[ProductionScheduleItem],
    today: date,
    horizon_end: date,
) -> list[ManagementIssueRow]:
    issues: list[ManagementIssueRow] = []
    items_by_order: dict[int, list[ProductionScheduleItem]] = defaultdict(list)
    for item in items:
        items_by_order[item.work_order_id].append(item)

    for order_items in items_by_order.values():
        visible_order_items = [
            item for item in order_items if not is_monitor_hidden_schedule_item(item)
        ]
        if not visible_order_items:
            continue
        work_order = visible_order_items[0].operation.work_order
        if work_order.due_date.date() > horizon_end:
            continue

        planned_start = min(item.start_time for item in visible_order_items)
        planned_end = max(item.end_time for item in visible_order_items)
        delay_days = _delay_days(planned_end, work_order.due_date)
        latest_item = max(visible_order_items, key=lambda item: item.end_time)
        due_margin_days = (work_order.due_date.date() - planned_end.date()).days
        due_in_days = (work_order.due_date.date() - today).days

        if delay_days > 0:
            issues.append(
                _issue(
                    schedule_id=schedule_id,
                    risk_type="order_delay",
                    risk_level="high",
                    title=f"{work_order.order_no} 预计延期 {delay_days} 天",
                    work_order_id=work_order.id,
                    order_no=work_order.order_no,
                    customer_name=work_order.customer,
                    product_name=work_order.product_name,
                    due_date=work_order.due_date,
                    planned_start_time=planned_start,
                    planned_end_time=planned_end,
                    delay_days=delay_days,
                    work_center_id=latest_item.work_center_id,
                    work_center_name=latest_item.operation.work_center.name,
                    machine_id=latest_item.machine_id,
                    machine_name=latest_item.machine.name if latest_item.machine else "外协",
                    operation_id=latest_item.operation_id,
                    operation_name=latest_item.operation.name,
                    reason=(
                        f"订单预计 {planned_end:%Y-%m-%d %H:%M} 完成，晚于交期 "
                        f"{work_order.due_date:%Y-%m-%d %H:%M}。"
                    ),
                    suggestion="优先检查最后完成工序、瓶颈资源和是否需要调整交期承诺。",
                )
            )
        elif due_in_days <= 3 or due_margin_days <= 3:
            issues.append(
                _issue(
                    schedule_id=schedule_id,
                    risk_type="due_soon",
                    risk_level="medium" if due_margin_days <= 3 else "low",
                    title=f"{work_order.order_no} 临近交期",
                    work_order_id=work_order.id,
                    order_no=work_order.order_no,
                    customer_name=work_order.customer,
                    product_name=work_order.product_name,
                    due_date=work_order.due_date,
                    planned_start_time=planned_start,
                    planned_end_time=planned_end,
                    delay_days=0,
                    work_center_id=latest_item.work_center_id,
                    work_center_name=latest_item.operation.work_center.name,
                    machine_id=latest_item.machine_id,
                    machine_name=latest_item.machine.name if latest_item.machine else "外协",
                    operation_id=latest_item.operation_id,
                    operation_name=latest_item.operation.name,
                    reason=f"订单距离交期剩余 {max(due_in_days, 0)} 天，计划完成余量 {max(due_margin_days, 0)} 天。",
                    suggestion="保持跟踪，不建议随意插单；如有异常应优先保护该订单资源窗口。",
                )
            )

        critical_items = [
            item
            for item in order_items
            if not is_monitor_hidden_schedule_item(item)
            and (
                item.end_time > work_order.due_date
                or item.end_time.date() >= work_order.due_date.date() - timedelta(days=1)
            )
        ]
        for item in critical_items[:3]:
            if item.is_external:
                continue
            risk_level = "high" if item.end_time > work_order.due_date else "medium"
            issues.append(
                _issue(
                    schedule_id=schedule_id,
                    risk_type="operation_blocking",
                    risk_level=risk_level,
                    title=f"{item.operation.name} 影响 {work_order.order_no} 交付",
                    work_order_id=work_order.id,
                    order_no=work_order.order_no,
                    customer_name=work_order.customer,
                    product_name=work_order.product_name,
                    due_date=work_order.due_date,
                    planned_start_time=item.start_time,
                    planned_end_time=item.end_time,
                    delay_days=_delay_days(item.end_time, work_order.due_date),
                    work_center_id=item.work_center_id,
                    work_center_name=item.operation.work_center.name,
                    machine_id=item.machine_id,
                    machine_name=item.machine.name if item.machine else None,
                    operation_id=item.operation_id,
                    operation_name=item.operation.name,
                    reason="关键内部工序计划完成时间贴近或晚于订单交期，后续缓冲不足。",
                    suggestion="优先确认该工序是否可提前、换机台或拆分处理。",
                )
            )

        for item in [value for value in order_items if value.is_external]:
            near_due = (work_order.due_date.date() - item.end_time.date()).days <= 3
            delayed = item.end_time > work_order.due_date
            if not delayed and not near_due:
                continue
            risk_level = "high" if delayed else "medium"
            issues.append(
                _issue(
                    schedule_id=schedule_id,
                    risk_type="external_risk",
                    risk_level=risk_level,
                    title=f"{work_order.order_no} 外协工序影响交付",
                    work_order_id=work_order.id,
                    order_no=work_order.order_no,
                    customer_name=work_order.customer,
                    product_name=work_order.product_name,
                    due_date=work_order.due_date,
                    planned_start_time=item.start_time,
                    planned_end_time=item.end_time,
                    delay_days=_delay_days(item.end_time, work_order.due_date),
                    work_center_id=item.work_center_id,
                    work_center_name=item.operation.work_center.name,
                    machine_id=None,
                    machine_name="外协",
                    operation_id=item.operation_id,
                    operation_name=item.operation.name,
                    reason="外协工序位于交付关键路径或接近交期完成，现场可控性较低。",
                    suggestion="尽快确认外协交付承诺，并预留回厂后的内部工序时间。",
                )
            )

    return issues


def _build_resource_issues(
    schedule_id: int,
    items: list[ProductionScheduleItem],
) -> list[ManagementIssueRow]:
    load = _resource_load_rows(_SchedulePlaceholder(schedule_id), items)
    issues: list[ManagementIssueRow] = []
    for resource in load.resources:
        if resource.utilization < 0.8:
            continue
        risk_level = "high" if resource.utilization >= 0.9 else "medium"
        issues.append(
            _issue(
                schedule_id=schedule_id,
                risk_type="resource_bottleneck",
                risk_level=risk_level,
                title=f"{resource.work_center_name} 负荷率 {resource.utilization:.0%}",
                work_center_id=resource.work_center_id,
                work_center_name=resource.work_center_name,
                machine_id=resource.machine_id,
                machine_name=resource.machine_name,
                utilization=resource.utilization,
                reason=(
                    f"{resource.machine_name} 占用 {resource.busy_minutes} 分钟，"
                    f"可用 {resource.available_minutes} 分钟，负荷率 {resource.utilization:.0%}。"
                ),
                suggestion="检查是否需要调整订单优先级、增加班次、换设备或临时外协。",
            )
        )
    return issues


class _SchedulePlaceholder:
    def __init__(self, schedule_id: int):
        self.id = schedule_id
        self.schedule_no = ""
        self.name = ""
        self.status = "draft"
        self.start_time = None
        self.base_schedule_id = None
        self.created_at = datetime.utcnow()
        self.updated_at = datetime.utcnow()


def _filter_issues(
    issues: list[ManagementIssueRow],
    risk_level: str | None,
    risk_type: str | None,
    customer: str | None,
    status: str | None,
) -> list[ManagementIssueRow]:
    filtered = issues
    if risk_level:
        filtered = [issue for issue in filtered if issue.risk_level == risk_level]
    if risk_type:
        filtered = [issue for issue in filtered if issue.risk_type == risk_type]
    if customer:
        filtered = [issue for issue in filtered if issue.customer_name == customer]
    if status:
        filtered = [issue for issue in filtered if issue.status == status]
    return filtered


def _sort_issues(issues: list[ManagementIssueRow]) -> list[ManagementIssueRow]:
    return sorted(
        issues,
        key=lambda issue: (
            RISK_LEVEL_ORDER.get(issue.risk_level, 9),
            issue.due_date or datetime.max,
            -(issue.delay_days or 0),
            issue.issue_key,
        ),
    )


def _summary(
    schedule_id: int,
    schedule_no: str,
    horizon_days: int,
    issues: list[ManagementIssueRow],
) -> ManagementDashboardSummary:
    risk_counts = Counter(issue.risk_level for issue in issues)
    type_counts = Counter(issue.risk_type for issue in issues)
    status_counts = Counter(issue.status for issue in issues)
    delayed_order_ids = {
        issue.work_order_id
        for issue in issues
        if issue.risk_type == "order_delay" and issue.work_order_id
    }
    due_soon_order_ids = {
        issue.work_order_id
        for issue in issues
        if issue.risk_type == "due_soon" and issue.work_order_id
    }
    return ManagementDashboardSummary(
        schedule_id=schedule_id,
        schedule_no=schedule_no,
        horizon_days=horizon_days,
        total_issues=len(issues),
        high_risk_issues=risk_counts["high"],
        medium_risk_issues=risk_counts["medium"],
        low_risk_issues=risk_counts["low"],
        delayed_orders=len(delayed_order_ids),
        due_soon_orders=len(due_soon_order_ids),
        bottleneck_resources=type_counts["resource_bottleneck"],
        external_risks=type_counts["external_risk"],
        open_issues=status_counts["open"],
        processing_issues=status_counts["processing"],
        resolved_issues=status_counts["resolved"],
        paused_issues=status_counts["paused"],
    )


async def get_management_dashboard(
    db: AsyncSession,
    schedule_id: int | None = None,
    horizon_days: int = 30,
    risk_level: str | None = None,
    risk_type: str | None = None,
    customer: str | None = None,
    status: str | None = None,
) -> ManagementDashboardResponse:
    schedule = await _resolve_schedule(db, schedule_id)
    all_items = await _load_schedule_items(db, schedule.id)
    today = date.today()
    horizon_end = today + timedelta(days=max(horizon_days, 1))
    items = _items_in_horizon(all_items, horizon_end)

    issues = _build_order_issues(schedule.id, items, today, horizon_end)
    issues.extend(_build_resource_issues(schedule.id, items))
    issues = _sort_issues(_merge_issue_states(issues, await _load_issue_states(db, schedule.id)))

    customers = sorted({issue.customer_name for issue in issues if issue.customer_name})
    risk_types = sorted({issue.risk_type for issue in issues})
    statuses = sorted({issue.status for issue in issues})
    filtered = _sort_issues(_filter_issues(issues, risk_level, risk_type, customer, status))

    return ManagementDashboardResponse(
        schedule=schedule,
        summary=_summary(schedule.id, schedule.schedule_no, horizon_days, filtered),
        issues=filtered,
        delivery_risks=[
            issue for issue in filtered if issue.risk_type in {"order_delay", "due_soon"}
        ],
        resource_risks=[
            issue for issue in filtered if issue.risk_type == "resource_bottleneck"
        ],
        operation_risks=[
            issue for issue in filtered if issue.risk_type == "operation_blocking"
        ],
        external_risks=[
            issue for issue in filtered if issue.risk_type == "external_risk"
        ],
        customers=customers,
        risk_types=risk_types,
        statuses=statuses,
    )


async def update_management_issue_state(
    db: AsyncSession,
    payload: ManagementIssueStateUpdate,
) -> ManagementIssueStateRead:
    if payload.status not in VALID_ISSUE_STATUSES:
        raise ValueError("问题状态必须是 open、processing、resolved 或 paused。")
    schedule = await _resolve_schedule(db, payload.schedule_id)
    result = await db.execute(
        select(BusinessRiskIssueState).where(
            BusinessRiskIssueState.schedule_id == schedule.id,
            BusinessRiskIssueState.issue_key == payload.issue_key,
        )
    )
    state = result.scalars().first()
    now = datetime.utcnow()
    if state is None:
        state = BusinessRiskIssueState(
            schedule_id=schedule.id,
            issue_key=payload.issue_key,
            status=payload.status,
            note=payload.note,
            created_at=now,
            updated_at=now,
        )
        db.add(state)
    else:
        state.status = payload.status
        state.note = payload.note
        state.updated_at = now

    await db.commit()
    await db.refresh(state)
    return ManagementIssueStateRead.model_validate(state)


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    fill = PatternFill("solid", fgColor="DDEAE4")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _append_issue_rows(ws, issues: list[ManagementIssueRow]) -> None:
    for issue in issues:
        ws.append(
            [
                RISK_TYPE_LABELS.get(issue.risk_type, issue.risk_type),
                issue.risk_level,
                issue.title,
                issue.order_no or "",
                issue.customer_name or "",
                issue.due_date.strftime("%Y-%m-%d %H:%M") if issue.due_date else "",
                issue.planned_end_time.strftime("%Y-%m-%d %H:%M") if issue.planned_end_time else "",
                issue.delay_days,
                issue.work_center_name or "",
                issue.machine_name or "",
                issue.operation_name or "",
                issue.utilization if issue.utilization is not None else "",
                issue.reason,
                issue.suggestion,
                issue.status,
                issue.note or "",
            ]
        )


async def export_management_dashboard_to_excel(
    db: AsyncSession,
    schedule_id: int | None = None,
    horizon_days: int = 30,
    risk_level: str | None = None,
    risk_type: str | None = None,
    customer: str | None = None,
    status: str | None = None,
) -> tuple[bytes, str]:
    dashboard = await get_management_dashboard(
        db,
        schedule_id=schedule_id,
        horizon_days=horizon_days,
        risk_level=risk_level,
        risk_type=risk_type,
        customer=customer,
        status=status,
    )
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "看板摘要"
    _write_header(ws1, ["指标", "值"])
    summary_rows = dashboard.summary.model_dump()
    for key, value in summary_rows.items():
        ws1.append([key, value])

    headers = [
        "风险类型",
        "风险等级",
        "标题",
        "订单号",
        "客户",
        "交期",
        "预计完成",
        "延期天数",
        "工段",
        "设备",
        "工序",
        "负荷率",
        "原因",
        "建议动作",
        "处理状态",
        "备注",
    ]
    sheets = [
        ("订单交付风险", dashboard.delivery_risks),
        ("资源瓶颈", dashboard.resource_risks),
        ("关键工序阻塞", dashboard.operation_risks),
        ("外协风险", dashboard.external_risks),
        ("全部问题状态", dashboard.issues),
    ]
    for title, issues in sheets:
        ws = wb.create_sheet(title)
        _write_header(ws, headers)
        _append_issue_rows(ws, issues)

    for ws in wb.worksheets:
        for column_cells in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 44)
            ws.column_dimensions[column_cells[0].column_letter].width = width

    output = BytesIO()
    wb.save(output)

    filename = f"经营问题看板_{dashboard.schedule.schedule_no}_{date.today():%Y%m%d}.xlsx"
    db.add(
        ExportBatch(
            export_type="management_dashboard",
            schedule_id=dashboard.schedule.id,
            filename=filename,
            params_json=json.dumps(
                {
                    "horizon_days": horizon_days,
                    "risk_level": risk_level,
                    "risk_type": risk_type,
                    "customer": customer,
                    "status": status,
                },
                ensure_ascii=False,
            ),
        )
    )
    await db.commit()
    return output.getvalue(), filename
