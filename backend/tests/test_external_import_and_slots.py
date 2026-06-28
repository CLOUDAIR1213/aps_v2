import asyncio
from io import BytesIO
from types import SimpleNamespace
import unittest

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.models.production import (
    ImportBatch,
    OperationMappingRule,
    Part,
    ProductionOperation,
    ProductionSchedule,
    ProductionScheduleItem,
    ProductionScheduleItemPersonnelAllocation,
    WorkCenter,
    WorkOrder,
)
from app.schemas.production import ImportCommitRequest, WorkOrderCreate
from app.services.production_import_service import parse_work_order_workbook
from app.services.production_service import _schedule_external_on_slot, commit_import, run_production_scheduling


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "焊接件明细"
    headers = [
        "NO",
        "图号",
        "名称",
        "材料厚",
        "材料长",
        "材料宽",
        "产品数量(件)",
        "材料",
        "单料重",
        "材料总重",
        "材料费",
        "钣金外",
    ]
    while len(headers) < 41:
        headers.append("")
    headers[40] = "备注1"
    for col, header in enumerate(headers, start=1):
        sheet.cell(1, col).value = header
    for row_index, row in enumerate(rows, start=2):
        for col, value in row.items():
            sheet.cell(row_index, col).value = value
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class ExternalImportTests(unittest.TestCase):
    def test_note_header_is_exposed_to_part_and_operations(self):
        content = workbook_bytes([{1: "1.1", 2: "DWG-1", 3: "noted", 7: 1, 11: 99, 12: 2, 41: "急件，先下料"}])
        preview = asyncio.run(
            parse_work_order_workbook(content, "notes.xlsx", {"钣金外": {"is_external": True}})
        )

        self.assertEqual(preview.parts[0].note, "急件，先下料")
        self.assertEqual(preview.operations[0].requirement_note, "急件，先下料")
        self.assertFalse(any(issue.field == "材料费" for issue in preview.issues))
        self.assertFalse(any(issue.field == "备注1" for issue in preview.issues))
        self.assertEqual(preview.summary["note_count"], 1)

    def test_external_blank_skips_comma_decimal_and_marker_uses_default(self):
        content = workbook_bytes(
            [
                {1: "1.1", 2: "DWG-1", 3: "blank", 7: 1, 12: None},
                {1: "1.2", 2: "DWG-2", 3: "comma", 7: 1, 12: "2,6"},
                {1: "1.3", 2: "DWG-3", 3: "marker", 7: 2, 12: "√"},
            ]
        )
        preview = asyncio.run(
            parse_work_order_workbook(
                content,
                "external.xlsx",
                {"钣金外": {"is_external": True, "default_duration_hours": 8}},
            )
        )

        self.assertEqual(len(preview.operations), 2)
        self.assertEqual([operation.duration_hours for operation in preview.operations], [2.6, 8])
        self.assertEqual(preview.summary["external_task_count"], 2)
        self.assertEqual(preview.summary["external_blank_skipped_count"], 1)
        self.assertEqual(preview.summary["external_default_duration_count"], 1)
        self.assertEqual(preview.summary["external_total_capacity_hours"], 18.6)

    def test_external_marker_without_default_warns_and_skips(self):
        content = workbook_bytes([{1: "1.1", 2: "DWG-1", 3: "marker", 7: 1, 12: "外协"}])
        preview = asyncio.run(
            parse_work_order_workbook(
                content,
                "external.xlsx",
                {"钣金外": {"is_external": True, "default_duration_hours": None}},
            )
        )

        self.assertEqual(len(preview.operations), 0)
        self.assertTrue(
            any(issue.field == "external_default_duration" and issue.severity == "warning" for issue in preview.issues)
        )

    def test_commit_import_persists_part_note_and_operation_requirement(self):
        async def run_test():
            engine = create_async_engine(
                "sqlite+aiosqlite:///:memory:",
                connect_args={"check_same_thread": False},
                poolclass=StaticPool,
            )
            async with engine.begin() as conn:
                await conn.run_sync(Base.metadata.create_all)

            session_factory = async_sessionmaker(engine, expire_on_commit=False)
            async with session_factory() as session:
                center = WorkCenter(code="EXT", name="钣金外", is_external=True)
                session.add(center)
                await session.commit()
                await session.refresh(center)
                session.add(
                    OperationMappingRule(
                        source_name="钣金外",
                        normalized_name="钣金外",
                        work_center_id=center.id,
                        is_external=True,
                        status="active",
                    )
                )
                await session.commit()

                content = workbook_bytes(
                    [{1: "1.1", 2: "DWG-1", 3: "noted", 7: 1, 12: 2, 41: "急件，先下料"}]
                )
                preview = await parse_work_order_workbook(
                    content,
                    "notes.xlsx",
                    {"钣金外": {"is_external": True}},
                )
                await commit_import(
                    session,
                    ImportCommitRequest(
                        order=WorkOrderCreate(
                            order_no="WO-NOTE-1",
                            customer="FUBEI",
                            product_name="焊接结构件",
                            quantity=1,
                            priority=1,
                            due_date=__import__("datetime").datetime(2026, 6, 30, 8, 0),
                        ),
                        preview=preview,
                        create_missing_work_centers=False,
                    ),
                )

                saved_part = (await session.execute(select(Part))).scalars().one()
                saved_operation = (await session.execute(select(ProductionOperation))).scalars().one()
                self.assertEqual(saved_part.note, "急件，先下料")
                self.assertEqual(saved_operation.requirement_note, "急件，先下料")

            await engine.dispose()

        asyncio.run(run_test())


class ExternalSlotTests(unittest.TestCase):
    def test_external_capacity_two_slots_allows_parallel_start(self):
        center = SimpleNamespace(id=1, external_capacity_slots=2)
        external_ready = {}
        base_start = __import__("datetime").datetime(2026, 5, 25, 8, 0)

        first_start, first_end, first_slot, _ = _schedule_external_on_slot(
            center, base_start, 4, external_ready, base_start
        )
        second_start, second_end, second_slot, _ = _schedule_external_on_slot(
            center, base_start, 4, external_ready, base_start
        )

        self.assertEqual(first_start, base_start)
        self.assertEqual(second_start, base_start)
        self.assertNotEqual(first_slot, second_slot)
        self.assertEqual(first_end, second_end)


class WorkCenterSchedulingTests(unittest.TestCase):
    def test_internal_work_center_without_personnel_can_schedule_without_default_allocation(self):
        async def run_test():
            engine = create_async_engine(
                "sqlite+aiosqlite:///:memory:",
                connect_args={"check_same_thread": False},
                poolclass=StaticPool,
            )
            async with engine.begin() as conn:
                await conn.run_sync(Base.metadata.create_all)

            session_factory = async_sessionmaker(engine, expire_on_commit=False)
            async with session_factory() as session:
                center = WorkCenter(code="WC-SP", name="喷丸", is_external=False)
                order = WorkOrder(
                    order_no="WO-SP-1",
                    customer="FUBEI",
                    product_name="焊接结构件",
                    quantity=1,
                    priority=1,
                    due_date=__import__("datetime").datetime(2026, 6, 30, 8, 0),
                    status="pending",
                )
                session.add_all([center, order])
                await session.commit()
                await session.refresh(center)
                await session.refresh(order)

                batch = ImportBatch(
                    work_order_id=order.id,
                    source_filename="manual.xlsx",
                    sheet_name="焊接件明细",
                    parsed_summary_json="{}",
                )
                session.add(batch)
                await session.commit()
                await session.refresh(batch)

                part = Part(
                    work_order_id=order.id,
                    import_batch_id=batch.id,
                    no="1.1",
                    drawing_no="DWG-SP",
                    name="喷丸件",
                    quantity=1,
                    source_row=2,
                    is_assembly=False,
                )
                session.add(part)
                await session.commit()
                await session.refresh(part)

                session.add(
                    ProductionOperation(
                        work_order_id=order.id,
                        part_id=part.id,
                        work_center_id=center.id,
                        name="喷丸",
                        seq_no=1,
                        duration_hours=2,
                        source_row=2,
                        source_col=12,
                    )
                )
                await session.commit()

                result = await run_production_scheduling(
                    session,
                    start_time=__import__("datetime").datetime(2026, 6, 5, 8, 0),
                    work_order_ids=[order.id],
                )

                self.assertEqual(len(result.items), 1)
                self.assertEqual(result.items[0].work_center_name, "喷丸")
                self.assertEqual(result.items[0].allocations, [])
                schedule_items = (await session.execute(select(ProductionScheduleItem))).scalars().all()
                allocations = (await session.execute(select(ProductionScheduleItemPersonnelAllocation))).scalars().all()
                self.assertEqual(len(schedule_items), 1)
                self.assertEqual(len(allocations), 0)

            await engine.dispose()

        asyncio.run(run_test())

    def test_scheduling_reuses_single_current_schedule(self):
        async def run_test():
            engine = create_async_engine(
                "sqlite+aiosqlite:///:memory:",
                connect_args={"check_same_thread": False},
                poolclass=StaticPool,
            )
            async with engine.begin() as conn:
                await conn.run_sync(Base.metadata.create_all)

            session_factory = async_sessionmaker(engine, expire_on_commit=False)
            async with session_factory() as session:
                center = WorkCenter(code="WC-ONE", name="统一工段", is_external=False)
                order = WorkOrder(
                    order_no="WO-ONE-1",
                    customer="FUBEI",
                    product_name="焊接结构件",
                    quantity=1,
                    priority=1,
                    due_date=__import__("datetime").datetime(2026, 6, 30, 8, 0),
                    status="pending",
                )
                session.add_all([center, order])
                await session.commit()
                await session.refresh(center)
                await session.refresh(order)

                batch = ImportBatch(
                    work_order_id=order.id,
                    source_filename="manual.xlsx",
                    sheet_name="焊接件明细",
                    parsed_summary_json="{}",
                )
                session.add(batch)
                await session.commit()
                await session.refresh(batch)

                part = Part(
                    work_order_id=order.id,
                    import_batch_id=batch.id,
                    no="1.1",
                    drawing_no="DWG-ONE",
                    name="统一排产件",
                    quantity=1,
                    source_row=2,
                    is_assembly=False,
                )
                session.add(part)
                await session.commit()
                await session.refresh(part)

                session.add(
                    ProductionOperation(
                        work_order_id=order.id,
                        part_id=part.id,
                        work_center_id=center.id,
                        name="喷丸",
                        seq_no=1,
                        duration_hours=2,
                        source_row=2,
                        source_col=12,
                    )
                )
                await session.commit()

                first = await run_production_scheduling(
                    session,
                    start_time=__import__("datetime").datetime(2026, 6, 5, 8, 0),
                    work_order_ids=[order.id],
                )
                second = await run_production_scheduling(
                    session,
                    start_time=__import__("datetime").datetime(2026, 6, 6, 8, 0),
                    work_order_ids=[order.id],
                )

                schedules = (await session.execute(select(ProductionSchedule))).scalars().all()
                schedule_items = (await session.execute(select(ProductionScheduleItem))).scalars().all()
                self.assertEqual(first.schedule.id, second.schedule.id)
                self.assertEqual(second.schedule.schedule_no, "PS-CURRENT")
                self.assertEqual(len(schedules), 1)
                self.assertEqual(len(schedule_items), 1)
                self.assertEqual(schedule_items[0].schedule_id, second.schedule.id)
                self.assertEqual(schedule_items[0].start_time, __import__("datetime").datetime(2026, 6, 6, 8, 0))

            await engine.dispose()

        asyncio.run(run_test())


if __name__ == "__main__":
    unittest.main()
