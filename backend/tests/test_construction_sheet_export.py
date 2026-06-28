import asyncio
from datetime import datetime, timedelta
from io import BytesIO
import unittest

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.models.production import (
    ExportBatch,
    Part,
    ProductionOperation,
    ProductionSchedule,
    ProductionScheduleItem,
    WorkCenter,
    WorkOrder,
)
from app.services.production_service import export_construction_sheets_to_excel


async def build_session():
    engine = create_async_engine(
        "sqlite+aiosqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    return engine, session_factory


class ConstructionSheetExportTests(unittest.TestCase):
    def test_export_groups_operations_by_part_without_dispatch(self):
        async def run_test():
            engine, session_factory = await build_session()
            try:
                async with session_factory() as session:
                    center = WorkCenter(code="WC-WELD", name="焊接", status="active")
                    order = WorkOrder(
                        order_no="WO-001",
                        customer="客户A",
                        product_name="测试产品",
                        quantity=5,
                        due_date=datetime(2026, 7, 1, 17, 0),
                    )
                    schedule = ProductionSchedule(schedule_no="PS-TEST", name="测试排产", status="published")
                    session.add_all([center, order, schedule])
                    await session.flush()

                    part_a = Part(
                        work_order_id=order.id,
                        no="1.1",
                        drawing_no="DWG-A",
                        name="支架A",
                        material="Q235",
                        note="关键尺寸A",
                        quantity=2,
                        material_weight=1.25,
                        source_row=2,
                    )
                    part_b = Part(
                        work_order_id=order.id,
                        no="1.2",
                        drawing_no="DWG-B",
                        name="支架B",
                        quantity=1,
                        material_weight=0,
                        source_row=3,
                    )
                    session.add_all([part_a, part_b])
                    await session.flush()

                    op_a1 = ProductionOperation(
                        work_order_id=order.id,
                        part_id=part_a.id,
                        work_center_id=center.id,
                        name="拼装",
                        seq_no=1,
                        duration_hours=1.5,
                        requirement_note="夹具A",
                        source_row=2,
                        source_col=12,
                    )
                    op_a2 = ProductionOperation(
                        work_order_id=order.id,
                        part_id=part_a.id,
                        work_center_id=center.id,
                        name="焊接",
                        seq_no=2,
                        duration_hours=2,
                        source_row=2,
                        source_col=13,
                    )
                    op_b1 = ProductionOperation(
                        work_order_id=order.id,
                        part_id=part_b.id,
                        work_center_id=center.id,
                        name="打磨",
                        seq_no=1,
                        duration_hours=0.5,
                        source_row=3,
                        source_col=14,
                    )
                    session.add_all([op_a1, op_a2, op_b1])
                    await session.flush()

                    start = datetime(2026, 6, 29, 8, 0)
                    items = [
                        ProductionScheduleItem(
                            schedule_id=schedule.id,
                            operation_id=op_a1.id,
                            work_order_id=order.id,
                            part_id=part_a.id,
                            work_center_id=center.id,
                            start_time=start,
                            end_time=start + timedelta(hours=1),
                            sequence_on_resource=1,
                        ),
                        ProductionScheduleItem(
                            schedule_id=schedule.id,
                            operation_id=op_a2.id,
                            work_order_id=order.id,
                            part_id=part_a.id,
                            work_center_id=center.id,
                            start_time=start + timedelta(hours=1),
                            end_time=start + timedelta(hours=3),
                            sequence_on_resource=2,
                        ),
                        ProductionScheduleItem(
                            schedule_id=schedule.id,
                            operation_id=op_b1.id,
                            work_order_id=order.id,
                            part_id=part_b.id,
                            work_center_id=center.id,
                            start_time=start + timedelta(hours=3),
                            end_time=start + timedelta(hours=4),
                            sequence_on_resource=3,
                        ),
                    ]
                    session.add_all(items)
                    await session.commit()

                    content, filename = await export_construction_sheets_to_excel(session, schedule.id)
                    workbook = load_workbook(BytesIO(content))

                    self.assertTrue(filename.startswith("施工单_PS-TEST_"))
                    self.assertEqual(workbook.sheetnames, ["WO-001-DWG-A", "WO-001-DWG-B"])
                    sheet = workbook["WO-001-DWG-A"]
                    self.assertEqual(sheet["E1"].value, "零 件 工 艺 施 工 单")
                    self.assertEqual(sheet["L1"].value, "DWG-A")
                    self.assertEqual(sheet["L2"].value, "支架A")
                    self.assertEqual(sheet["B7"].value, "拼装")
                    self.assertEqual(sheet["C7"].value, "夹具A")
                    self.assertEqual(sheet["F7"].value, 1.5)
                    self.assertEqual(sheet["L7"].value, "关键尺寸A")
                    self.assertEqual(sheet["M7"].value, f"*JG-PS-TEST-{items[0].id}*")
                    self.assertEqual(sheet.print_title_rows, "$1:$6")

                    export_record = await session.scalar(
                        select(ExportBatch).where(ExportBatch.export_type == "construction_sheet")
                    )
                    self.assertIsNotNone(export_record)
                    self.assertIn("\"sheet_count\": 2", export_record.params_json)
            finally:
                await engine.dispose()

        asyncio.run(run_test())

    def test_export_extends_print_area_when_operations_exceed_first_page(self):
        async def run_test():
            engine, session_factory = await build_session()
            try:
                async with session_factory() as session:
                    center = WorkCenter(code="WC-MACH", name="机加工", status="active")
                    order = WorkOrder(
                        order_no="WO-OVER",
                        customer="客户B",
                        product_name="长工艺产品",
                        quantity=1,
                        due_date=datetime(2026, 7, 2, 17, 0),
                    )
                    schedule = ProductionSchedule(schedule_no="PS-LONG", name="长工艺排产", status="published")
                    session.add_all([center, order, schedule])
                    await session.flush()

                    part = Part(
                        work_order_id=order.id,
                        no="2.1",
                        drawing_no="DWG-LONG",
                        name="长工艺零件",
                        quantity=1,
                        material_weight=0,
                        source_row=2,
                    )
                    session.add(part)
                    await session.flush()

                    start = datetime(2026, 6, 29, 8, 0)
                    for index in range(20):
                        operation = ProductionOperation(
                            work_order_id=order.id,
                            part_id=part.id,
                            work_center_id=center.id,
                            name=f"工序{index + 1}",
                            seq_no=index + 1,
                            duration_hours=1,
                            source_row=2,
                            source_col=12 + index,
                        )
                        session.add(operation)
                        await session.flush()
                        session.add(
                            ProductionScheduleItem(
                                schedule_id=schedule.id,
                                operation_id=operation.id,
                                work_order_id=order.id,
                                part_id=part.id,
                                work_center_id=center.id,
                                start_time=start + timedelta(hours=index),
                                end_time=start + timedelta(hours=index + 1),
                                sequence_on_resource=index + 1,
                            )
                        )
                    await session.commit()

                    content, _filename = await export_construction_sheets_to_excel(session, schedule.id)
                    sheet = load_workbook(BytesIO(content))["WO-OVER-DWG-LONG"]

                    self.assertEqual(sheet["B26"].value, "工序20")
                    self.assertEqual(sheet["A27"].value, "制表：")
                    self.assertEqual(str(sheet.print_area), "'WO-OVER-DWG-LONG'!$A$1:$M$27")
            finally:
                await engine.dispose()

        asyncio.run(run_test())


if __name__ == "__main__":
    unittest.main()
