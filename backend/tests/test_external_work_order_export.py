import asyncio
from datetime import datetime, timedelta
from io import BytesIO
from zipfile import ZipFile
import unittest

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.models.production import (
    ExportBatch,
    Part,
    ProductionOperation,
    ProductionSchedule,
    ProductionScheduleItem,
    WorkCenter,
    WorkOrder,
)
from app.services.production_service import export_external_work_orders_to_excel


async def build_session():
    engine = create_async_engine(
        "sqlite+aiosqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    return engine, session_factory


async def seed_external_task(
    session,
    *,
    schedule,
    center,
    order,
    part_no,
    drawing_no,
    part_name,
    operation_name,
    seq_no,
    start_time,
    expected_return_at,
    requirement_note=None,
    external_status="pending",
    external_note=None,
):
    part = Part(
        work_order_id=order.id,
        no=part_no,
        drawing_no=drawing_no,
        name=part_name,
        quantity=1,
        material_weight=0,
        source_row=seq_no + 1,
    )
    session.add(part)
    await session.flush()
    operation = ProductionOperation(
        work_order_id=order.id,
        part_id=part.id,
        work_center_id=center.id,
        name=operation_name,
        seq_no=seq_no,
        duration_hours=8,
        requirement_note=requirement_note,
        source_row=part.source_row,
        source_col=20 + seq_no,
    )
    session.add(operation)
    await session.flush()
    item = ProductionScheduleItem(
        schedule_id=schedule.id,
        operation_id=operation.id,
        work_order_id=order.id,
        part_id=part.id,
        work_center_id=center.id,
        start_time=start_time,
        end_time=expected_return_at,
        sequence_on_resource=seq_no,
        is_external=True,
        external_status=external_status,
        external_expected_return_at=expected_return_at,
        external_note=external_note,
    )
    session.add(item)
    await session.flush()
    return item


class ExternalWorkOrderExportTests(unittest.TestCase):
    def test_export_groups_by_vendor_and_uses_template_print_rules(self):
        async def run_test():
            engine, session_factory = await build_session()
            try:
                async with session_factory() as session:
                    schedule = ProductionSchedule(schedule_no="PS-WX", name="外协排产", status="published")
                    center_a = WorkCenter(
                        code="EXT-A",
                        name="钣金外",
                        is_external=True,
                        external_vendor_name="供应商A",
                    )
                    center_blank = WorkCenter(
                        code="EXT-B",
                        name="表面处理",
                        is_external=True,
                        external_vendor_name=None,
                    )
                    order = WorkOrder(
                        order_no="WO-WX-1",
                        customer="客户A",
                        product_name="外协产品",
                        quantity=1,
                        due_date=datetime(2026, 7, 10, 17, 0),
                    )
                    session.add_all([schedule, center_a, center_blank, order])
                    await session.flush()

                    start = datetime(2026, 6, 29, 8, 0)
                    first_item = None
                    for index in range(37):
                        item = await seed_external_task(
                            session,
                            schedule=schedule,
                            center=center_a,
                            order=order,
                            part_no=f"1.{index + 1}",
                            drawing_no=f"DWG-A-{index + 1}",
                            part_name=f"外协件A{index + 1}",
                            operation_name="钣金外",
                            seq_no=index + 1,
                            start_time=start + timedelta(hours=index),
                            expected_return_at=start + timedelta(days=2, hours=index),
                            requirement_note="激光落料",
                        )
                        first_item = first_item or item
                    await seed_external_task(
                        session,
                        schedule=schedule,
                        center=center_blank,
                        order=order,
                        part_no="2.1",
                        drawing_no="DWG-B-1",
                        part_name="外协件B1",
                        operation_name="表面处理",
                        seq_no=21,
                        start_time=start,
                        expected_return_at=start + timedelta(days=3),
                        external_status="sent",
                        external_note="随货带检验记录",
                    )
                    await session.commit()

                    content, filename = await export_external_work_orders_to_excel(session, schedule.id)
                    workbook = load_workbook(BytesIO(content))

                    self.assertTrue(filename.startswith("外协工单_PS-WX_"))
                    self.assertEqual(set(workbook.sheetnames), {"供应商A", "供应商A-2", "未指定供应商"})
                    vendor_sheet = workbook["供应商A"]
                    self.assertEqual(vendor_sheet["E1"].value, "外 协 加 工 委 托 单")
                    self.assertEqual(vendor_sheet["L2"].value, "供应商A")
                    self.assertEqual(vendor_sheet["F4"].value, 37)
                    self.assertEqual(vendor_sheet["B7"].value, "WO-WX-1")
                    self.assertEqual(vendor_sheet["C7"].value, "DWG-A-1 / 1.1 / 外协件A1")
                    self.assertEqual(vendor_sheet["G7"].value, "钣金外")
                    self.assertEqual(vendor_sheet["L7"].value, "激光落料")
                    self.assertEqual(vendor_sheet["M7"].value, f"*WX-{first_item.id}*")
                    self.assertEqual(str(vendor_sheet.print_area), "'供应商A'!$A$1:$M$44")
                    self.assertFalse(vendor_sheet.row_dimensions[26].hidden)
                    self.assertFalse(vendor_sheet.row_dimensions[44].hidden)

                    continuation_sheet = workbook["供应商A-2"]
                    self.assertEqual(continuation_sheet["L2"].value, "供应商A")
                    self.assertEqual(continuation_sheet["A7"].value, 37)
                    self.assertEqual(continuation_sheet["C7"].value, "DWG-A-37 / 1.37 / 外协件A37")
                    self.assertEqual(str(continuation_sheet.print_area), "'供应商A-2'!$A$1:$M$25")
                    self.assertTrue(continuation_sheet.row_dimensions[26].hidden)
                    self.assertTrue(continuation_sheet.row_dimensions[44].hidden)

                    blank_sheet = workbook["未指定供应商"]
                    self.assertEqual(blank_sheet["L2"].value, "未指定供应商")
                    self.assertEqual(str(blank_sheet.print_area), "'未指定供应商'!$A$1:$M$25")
                    self.assertTrue(blank_sheet.row_dimensions[26].hidden)
                    self.assertTrue(blank_sheet.row_dimensions[44].hidden)
                    self.assertEqual(blank_sheet["J7"].value, "已发出")
                    self.assertEqual(blank_sheet["L7"].value, "随货带检验记录")

                    with ZipFile(BytesIO(content)) as archive:
                        external_links = [
                            name for name in archive.namelist()
                            if "externalLink" in name or "externalLinks" in name
                        ]
                    self.assertEqual(external_links, [])

                    export_record = await session.scalar(
                        select(ExportBatch).where(ExportBatch.export_type == "external_work_order")
                    )
                    self.assertIsNotNone(export_record)
                    self.assertIn("\"vendor_count\": 2", export_record.params_json)
                    self.assertIn("\"sheet_count\": 3", export_record.params_json)
            finally:
                await engine.dispose()

        asyncio.run(run_test())

    def test_export_raises_when_no_external_tasks_match(self):
        async def run_test():
            engine, session_factory = await build_session()
            try:
                async with session_factory() as session:
                    schedule = ProductionSchedule(schedule_no="PS-EMPTY", name="空外协", status="published")
                    session.add(schedule)
                    await session.commit()

                    with self.assertRaisesRegex(ValueError, "没有外协任务"):
                        await export_external_work_orders_to_excel(session, schedule.id)
            finally:
                await engine.dispose()

        asyncio.run(run_test())


if __name__ == "__main__":
    unittest.main()
