from __future__ import annotations

import argparse
import asyncio
import json
import sys
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import selectinload


BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_DIR.parent
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.database import AsyncSessionLocal  # noqa: E402
from app.models.production import (  # noqa: E402
    OperationMappingRule,
    ResourceMachine,
    WorkCenter,
    WorkOrder,
)
from app.schemas.production import ImportCommitRequest, WorkOrderCreate  # noqa: E402
from app.services.management_dashboard_service import (  # noqa: E402
    export_management_dashboard_to_excel,
    get_management_dashboard,
    update_management_issue_state,
)
from app.schemas.production import ManagementIssueStateUpdate  # noqa: E402
from app.services.production_import_service import (  # noqa: E402
    BASE_COLUMNS,
    IGNORED_OPERATION_HEADERS,
    PRIMARY_SHEET,
    parse_work_order_workbook,
)
from app.services.production_analysis_service import (  # noqa: E402
    _load_schedule_items,
    get_production_scheduling_overview,
)
from app.services.production_service import commit_import, run_production_scheduling  # noqa: E402


DEFAULT_SOURCE = Path(
    r"C:\Users\48295\Desktop\20260425给奕航\工艺表\上海FUBEI-20260131-工艺，改5m，1.6立车.xlsm"
)
DEFAULT_OUTPUT_DIR = REPO_ROOT / "outputs" / "management_dashboard_acceptance"
EXPECTED_BASE_HEADERS = [
    "NO",
    "图号",
    "名称",
    "材料厚",
    "材料长",
    "材料宽",
    "产品数量(件)",
    "材料",
    "单料重",
    "材料总重",
    "材料费",
]
EXTERNAL_HEADERS = {"钣金外", "加工外", "表面处理"}
OPERATION_CODES = {
    "下料": "CUT",
    "整形": "STRAIGHTEN",
    "拼装": "ASSEMBLY",
    "焊接": "WELD",
    "打磨": "GRIND",
    "热处理": "HEAT",
    "喷丸": "SHOT",
    "底漆": "PRIMER",
    "车6150": "LATHE6150",
    "车6180": "LATHE6180",
    "立车1.6": "VLATHE16",
    "1m平磨": "GRIND1M",
    "小普铣": "MILL-S",
    "侧铣": "SIDE-MILL",
    "数车50": "CNC50",
    "1米立加": "VMC1M",
    "3m龙门": "GANTRY3M",
    "5m龙门": "GANTRY5M",
    "6m龙门": "GANTRY6M",
    "划线": "MARK",
    "3050钻": "DRILL3050",
    "气攻": "TAP",
    "去毛清理": "DEBURR",
    "面漆": "TOPCOAT",
    "钣金外": "EXT-SHEET",
    "加工外": "EXT-MACH",
    "表面处理": "EXT-SURFACE",
    "检验": "INSPECT",
    "装配": "FINAL-ASSY",
}
CAPACITY_MINUTES = {
    "焊接": 120,
    "5m龙门": 120,
    "拼装": 240,
}


@dataclass(frozen=True)
class Scenario:
    order_no: str
    customer: str
    product_name: str
    priority: int
    due_days: int
    due_hour: int = 17
    duration_scale: float = 1.0
    operation_multipliers: dict[str, float] | None = None
    expected_focus: str = ""


SCENARIOS = [
    Scenario(
        order_no="FUBEI-DUE-002",
        customer="FUBEI-临近交期",
        product_name="上托架机构右固定-临近交期",
        priority=10,
        due_days=3,
        duration_scale=0.02,
        expected_focus="due_soon",
    ),
    Scenario(
        order_no="FUBEI-DELAY-001",
        customer="FUBEI-延期",
        product_name="上托架机构右固定-延期",
        priority=8,
        due_days=2,
        duration_scale=1.2,
        expected_focus="order_delay",
    ),
    Scenario(
        order_no="FUBEI-BOTTLENECK-003",
        customer="FUBEI-瓶颈",
        product_name="上托架机构右固定-资源瓶颈",
        priority=5,
        due_days=24,
        operation_multipliers={"焊接": 3.0, "5m龙门": 3.0},
        expected_focus="resource_bottleneck",
    ),
    Scenario(
        order_no="FUBEI-BLOCK-004",
        customer="FUBEI-阻塞",
        product_name="上托架机构右固定-关键工序阻塞",
        priority=6,
        due_days=12,
        operation_multipliers={"拼装": 2.0, "焊接": 2.0},
        expected_focus="operation_blocking",
    ),
    Scenario(
        order_no="FUBEI-EXT-005",
        customer="FUBEI-外协",
        product_name="上托架机构右固定-外协影响",
        priority=7,
        due_days=8,
        operation_multipliers={"钣金外": 4.0},
        expected_focus="external_risk",
    ),
]


def _to_text(value) -> str:
    return "" if value is None else str(value).strip()


def _is_number(value) -> bool:
    if value in (None, "") or isinstance(value, bool):
        return False
    try:
        float(value)
    except (TypeError, ValueError):
        return False
    return True


def load_primary_headers(source: Path) -> list[str]:
    workbook = load_workbook(source, read_only=True, data_only=True, keep_vba=True)
    if PRIMARY_SHEET not in workbook.sheetnames:
        raise ValueError(f"源文件缺少 Sheet：{PRIMARY_SHEET}")
    sheet = workbook[PRIMARY_SHEET]
    headers = [_to_text(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    if headers[:BASE_COLUMNS] != EXPECTED_BASE_HEADERS:
        raise ValueError(
            "源文件前 11 列不符合当前导入解析规则："
            f"实际={headers[:BASE_COLUMNS]}，期望={EXPECTED_BASE_HEADERS}"
        )
    return headers


def operation_headers(headers: list[str]) -> list[str]:
    return [
        header
        for header in headers[BASE_COLUMNS:]
        if header not in IGNORED_OPERATION_HEADERS
    ]


async def _unique_code(db, preferred: str, existing_id: int | None = None) -> str:
    candidate = preferred[:50]
    index = 1
    while True:
        result = await db.execute(select(WorkCenter).where(WorkCenter.code == candidate))
        row = result.scalars().first()
        if row is None or row.id == existing_id:
            return candidate
        suffix = f"-{index}"
        candidate = f"{preferred[:50 - len(suffix)]}{suffix}"
        index += 1


async def prepare_foundation_data(headers: list[str]) -> dict[str, int]:
    center_ids: dict[str, int] = {}
    async with AsyncSessionLocal() as db:
        for header in operation_headers(headers):
            is_external = header in EXTERNAL_HEADERS
            capacity = CAPACITY_MINUTES.get(header, 480)
            default_duration = 0 if is_external else 8
            result = await db.execute(select(WorkCenter).where(WorkCenter.name == header))
            center = result.scalars().first()
            if center is None:
                center = WorkCenter(
                    code=await _unique_code(db, f"ACC-{OPERATION_CODES.get(header, header.upper())}"),
                    name=header,
                    is_external=is_external,
                    default_capacity_per_day=capacity,
                    default_duration_hours=default_duration,
                    status="active",
                    description="经营问题看板真实工艺表验收数据",
                )
                db.add(center)
                await db.flush()
            else:
                center.is_external = is_external
                center.default_capacity_per_day = capacity
                center.default_duration_hours = default_duration
                center.status = "active"
                if not center.description:
                    center.description = "经营问题看板真实工艺表验收数据"
            await db.flush()
            center_ids[header] = center.id

            if not is_external:
                machines = list(
                    (
                        await db.execute(
                            select(ResourceMachine)
                            .where(ResourceMachine.work_center_id == center.id)
                            .order_by(ResourceMachine.id)
                        )
                    ).scalars().all()
                )
                if machines:
                    machine = machines[0]
                    machine.status = "active"
                    machine.capacity_per_day = capacity
                else:
                    machine_code = f"{center.code}-01"
                    machine = ResourceMachine(
                        work_center_id=center.id,
                        code=machine_code[:50],
                        name=f"{header}-01",
                        status="active",
                        capacity_per_day=capacity,
                    )
                    db.add(machine)

            rule_result = await db.execute(
                select(OperationMappingRule).where(OperationMappingRule.source_name == header)
            )
            rule = rule_result.scalars().first()
            if rule is None:
                db.add(
                    OperationMappingRule(
                        source_name=header,
                        normalized_name=header,
                        work_center_id=center.id,
                        is_external=is_external,
                        status="active",
                    )
                )
            else:
                rule.normalized_name = header
                rule.work_center_id = center.id
                rule.is_external = is_external
                rule.status = "active"

        await db.commit()
    return center_ids


def generate_scenario_workbook(source: Path, output_dir: Path, scenario: Scenario, headers: list[str]) -> Path:
    workbook = load_workbook(source, read_only=False, data_only=True, keep_vba=False)
    if PRIMARY_SHEET not in workbook.sheetnames:
        raise ValueError(f"源文件缺少 Sheet：{PRIMARY_SHEET}")
    for sheet in list(workbook.worksheets):
        if sheet.title != PRIMARY_SHEET:
            workbook.remove(sheet)
    sheet = workbook[PRIMARY_SHEET]

    header_by_col = {
        col: _to_text(sheet.cell(1, col).value)
        for col in range(BASE_COLUMNS + 1, sheet.max_column + 1)
    }
    multipliers = scenario.operation_multipliers or {}
    for row in range(2, sheet.max_row + 1):
        no = sheet.cell(row, 1).value
        drawing = sheet.cell(row, 2).value
        name = sheet.cell(row, 3).value
        if not no and not drawing and not name:
            continue
        if drawing == "第一行不输入":
            continue
        for col, header in header_by_col.items():
            if header in IGNORED_OPERATION_HEADERS:
                continue
            cell = sheet.cell(row, col)
            if not _is_number(cell.value):
                continue
            value = float(cell.value)
            if scenario.duration_scale != 1.0:
                value *= scenario.duration_scale
            if header in multipliers:
                value *= multipliers[header]
            cell.value = round(max(value, 0.02), 3)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{scenario.order_no}_焊接件明细.xlsx"
    workbook.save(output_path)
    return output_path


async def _active_mapping_rules() -> dict[str, dict]:
    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(OperationMappingRule)
            .where(OperationMappingRule.status == "active")
            .options(selectinload(OperationMappingRule.work_center))
        )
        return {
            rule.source_name: {
                "is_external": rule.is_external,
                "work_center_id": rule.work_center_id,
                "default_duration_hours": (
                    rule.work_center.default_duration_hours if rule.work_center else None
                ),
            }
            for rule in result.scalars().all()
        }


async def preview_generated_files(files: dict[str, Path]) -> dict[str, dict]:
    mapping_rules = await _active_mapping_rules()
    previews: dict[str, dict] = {}
    for order_no, path in files.items():
        preview = await parse_work_order_workbook(path.read_bytes(), path.name, mapping_rules)
        external_ops = len([operation for operation in preview.operations if operation.is_external])
        previews[order_no] = {
            "part_count": preview.summary.get("part_count"),
            "operation_count": preview.summary.get("operation_count"),
            "work_center_count": preview.summary.get("work_center_count"),
            "total_hours": preview.summary.get("total_hours"),
            "error_count": preview.summary.get("error_count"),
            "warning_count": preview.summary.get("warning_count"),
            "external_operation_count": external_ops,
        }
        if preview.summary.get("error_count"):
            messages = [issue.message for issue in preview.issues if issue.severity == "error"]
            raise ValueError(f"{path.name} 预览仍有错误：{messages[:5]}")
    return previews


def due_datetime(today: date, scenario: Scenario) -> datetime:
    return datetime.combine(today + timedelta(days=scenario.due_days), time(scenario.due_hour, 0))


def write_manifest(
    output_dir: Path,
    source: Path,
    files: dict[str, Path],
    previews: dict[str, dict],
    schedule_id: int | None = None,
    dashboard: dict | None = None,
) -> Path:
    today = date.today()
    orders = []
    for scenario in SCENARIOS:
        orders.append(
            {
                "order_no": scenario.order_no,
                "customer": scenario.customer,
                "product_name": scenario.product_name,
                "priority": scenario.priority,
                "quantity": 1,
                "due_date": due_datetime(today, scenario).isoformat(timespec="minutes"),
                "file": str(files[scenario.order_no]),
                "expected_focus": scenario.expected_focus,
                "preview": previews.get(scenario.order_no, {}),
            }
        )

    payload = {
        "source": str(source),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "schedule_start": datetime.combine(today, time(8, 0)).isoformat(timespec="minutes"),
        "manual_flow": [
            "在工单导入页逐个上传 file 字段中的 Excel。",
            "按同一条目中的 order_no/customer/product_name/priority/due_date 填写订单信息。",
            "确认导入 5 张订单后，在排产驾驶台选择这 5 张订单并从 schedule_start 开始排产。",
            "进入 /management-dashboard 验收五类风险、筛选、状态备注、下钻和导出。",
        ],
        "orders": orders,
        "schedule_id": schedule_id,
        "dashboard": dashboard,
    }
    manifest_path = output_dir / "FUBEI_经营看板验收_manifest.json"
    manifest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest_path


def _write_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="DDEAE4")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _fit_columns(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 36)
            sheet.column_dimensions[column_cells[0].column_letter].width = width


async def export_schedule_timetable(db, schedule_id: int, output_dir: Path) -> Path:
    overview = await get_production_scheduling_overview(db, schedule_id)
    items = await _load_schedule_items(db, schedule_id)
    workbook = Workbook()

    summary = workbook.active
    summary.title = "订单排产时间表"
    _write_header(
        summary,
        [
            "排产方案",
            "订单号",
            "客户",
            "产品",
            "交期",
            "计划开始",
            "计划完成",
            "延期天数",
            "状态",
            "最后完成工序",
            "最后工段",
            "最后设备",
        ],
    )
    by_order = {}
    for item in items:
        by_order.setdefault(item.work_order_id, []).append(item)
    for row in overview.orders:
        order_items = by_order.get(row.work_order_id, [])
        latest_item = max(order_items, key=lambda value: value.end_time) if order_items else None
        summary.append(
            [
                overview.schedule_no,
                row.order_no,
                row.customer_name,
                row.product_name,
                row.due_date,
                row.planned_start_time,
                row.planned_end_time,
                row.delay_days,
                row.status,
                latest_item.operation.name if latest_item else "",
                latest_item.operation.work_center.name if latest_item else "",
                latest_item.machine.name if latest_item and latest_item.machine else ("外协" if latest_item and latest_item.is_external else ""),
            ]
        )

    detail = workbook.create_sheet("工序排产明细")
    _write_header(
        detail,
        [
            "订单号",
            "客户",
            "零件NO",
            "图号",
            "零件名称",
            "工序",
            "工段",
            "设备",
            "计划开始",
            "计划完成",
            "工时",
            "外协",
            "资源顺序",
        ],
    )
    for item in sorted(items, key=lambda value: (value.start_time, value.work_order_id, value.sequence_on_resource)):
        operation = item.operation
        detail.append(
            [
                operation.work_order.order_no,
                operation.work_order.customer,
                operation.part.no,
                operation.part.drawing_no,
                operation.part.name,
                operation.name,
                operation.work_center.name,
                item.machine.name if item.machine else "外协",
                item.start_time,
                item.end_time,
                operation.duration_hours,
                "是" if item.is_external else "否",
                item.sequence_on_resource,
            ]
        )

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"
    _fit_columns(workbook)

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"排产时间表_{overview.schedule_no}_{date.today():%Y%m%d}.xlsx"
    workbook.save(path)
    return path


def update_manifest_timetable_path(output_dir: Path, timetable_path: Path) -> None:
    manifest_path = output_dir / "FUBEI_经营看板验收_manifest.json"
    if not manifest_path.exists():
        return
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    dashboard = payload.setdefault("dashboard", {})
    dashboard["schedule_timetable_path"] = str(timetable_path)
    manifest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


async def commit_and_schedule(files: dict[str, Path]) -> tuple[int, dict]:
    mapping_rules = await _active_mapping_rules()
    order_ids: list[int] = []
    today = date.today()
    async with AsyncSessionLocal() as db:
        for scenario in SCENARIOS:
            existing = (
                await db.execute(select(WorkOrder).where(WorkOrder.order_no == scenario.order_no))
            ).scalars().first()
            if existing:
                order_ids.append(existing.id)
                continue
            path = files[scenario.order_no]
            preview = await parse_work_order_workbook(path.read_bytes(), path.name, mapping_rules)
            payload = ImportCommitRequest(
                order=WorkOrderCreate(
                    order_no=scenario.order_no,
                    customer=scenario.customer,
                    product_name=scenario.product_name,
                    quantity=1,
                    priority=scenario.priority,
                    due_date=due_datetime(today, scenario),
                ),
                preview=preview,
                create_missing_work_centers=False,
            )
            result = await commit_import(db, payload)
            order_ids.append(result.work_order.id)

        schedule_start = datetime.combine(today, time(8, 0))
        result = await run_production_scheduling(
            db,
            start_time=schedule_start,
            work_order_ids=order_ids,
            keep_locked=False,
        )
        dashboard = await get_management_dashboard(db, schedule_id=result.schedule.id, horizon_days=30)
        if dashboard.issues:
            first_issue = dashboard.issues[0]
            await update_management_issue_state(
                db,
                ManagementIssueStateUpdate(
                    schedule_id=result.schedule.id,
                    issue_key=first_issue.issue_key,
                    status="processing",
                    note="验收脚本写入：状态备注刷新保留检查。",
                ),
            )
            dashboard = await get_management_dashboard(db, schedule_id=result.schedule.id, horizon_days=30)
        content, filename = await export_management_dashboard_to_excel(
            db,
            schedule_id=result.schedule.id,
            horizon_days=30,
        )
        export_path = DEFAULT_OUTPUT_DIR / filename
        export_path.write_bytes(content)
        timetable_path = await export_schedule_timetable(db, result.schedule.id, DEFAULT_OUTPUT_DIR)

    risk_types = sorted({issue.risk_type for issue in dashboard.issues})
    risk_levels = {
        risk_type: sorted({issue.risk_level for issue in dashboard.issues if issue.risk_type == risk_type})
        for risk_type in risk_types
    }
    expected = {
        "order_delay",
        "due_soon",
        "resource_bottleneck",
        "operation_blocking",
        "external_risk",
    }
    return result.schedule.id, {
        "summary": dashboard.summary.model_dump(),
        "risk_types": risk_types,
        "risk_levels": risk_levels,
        "missing_risk_types": sorted(expected - set(risk_types)),
        "export_path": str(export_path),
        "schedule_timetable_path": str(timetable_path),
    }


async def run(args: argparse.Namespace) -> None:
    source = Path(args.source)
    output_dir = Path(args.output_dir)
    if not source.exists():
        raise FileNotFoundError(f"源文件不存在：{source}")

    if args.export_schedule_id:
        async with AsyncSessionLocal() as db:
            path = await export_schedule_timetable(db, args.export_schedule_id, output_dir)
        update_manifest_timetable_path(output_dir, path)
        print(f"排产时间表已导出：{path}")
        return

    headers = load_primary_headers(source)
    if not args.skip_db:
        await prepare_foundation_data(headers)

    files = {
        scenario.order_no: generate_scenario_workbook(source, output_dir, scenario, headers)
        for scenario in SCENARIOS
    }
    previews = await preview_generated_files(files) if not args.skip_db else {}
    schedule_id = None
    dashboard = None
    if args.commit_and_schedule:
        if args.skip_db:
            raise ValueError("--commit-and-schedule 不能与 --skip-db 同时使用。")
        schedule_id, dashboard = await commit_and_schedule(files)
    manifest = write_manifest(output_dir, source, files, previews, schedule_id=schedule_id, dashboard=dashboard)

    print("经营问题看板验收数据已准备完成。")
    print(f"输出目录：{output_dir}")
    print(f"说明文件：{manifest}")
    for scenario in SCENARIOS:
        print(f"- {scenario.order_no}: {files[scenario.order_no]}")
    if dashboard:
        print("看板验收摘要：")
        print(json.dumps(dashboard, ensure_ascii=False, indent=2))


def main() -> None:
    parser = argparse.ArgumentParser(description="准备 FUBEI 经营问题看板验收数据。")
    parser.add_argument("--source", default=str(DEFAULT_SOURCE), help="真实 FUBEI 工艺表路径。")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="派生验收 Excel 输出目录。")
    parser.add_argument("--skip-db", action="store_true", help="只生成 Excel，不准备基础数据/映射。")
    parser.add_argument(
        "--commit-and-schedule",
        action="store_true",
        help="生成 Excel 后直接导入 5 单、运行排产、导出经营看板 Excel。",
    )
    parser.add_argument(
        "--export-schedule-id",
        type=int,
        default=None,
        help="只为已有排产方案导出订单排产时间表，不生成验收 Excel。",
    )
    args = parser.parse_args()
    asyncio.run(run(args))


if __name__ == "__main__":
    main()
